"""Focused tests for safe XLSX validation and business-header inspection."""

from __future__ import annotations

import struct
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from services.workbook_validation import (
    MappingStatus,
    WorkbookValidationError,
    normalize_header,
    validate_and_inspect_workbook,
    validate_generated_workbook,
)


def save_workbook(path: Path, rows: list[list[object]] | None = None) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bảng giá"
    for row in rows or []:
        sheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def inspect(path: Path, *, max_rows: int = 100, max_columns: int = 20):
    return validate_and_inspect_workbook(
        path,
        max_rows=max_rows,
        max_columns=max_columns,
    )


def test_maps_approved_vietnamese_headers_and_returns_safe_metadata(
    tmp_path: Path,
) -> None:
    path = save_workbook(
        tmp_path / "prices.xlsx",
        [
            ["BÁO CÁO THÁNG 7"],
            ["Họ tên", "Mã đặt-chỗ", "Số vé", "Giá hệ thống", "Giá thu"],
            ["NGUYEN VAN A", "ABC123", "738001", 1_000_000, 1_100_000],
        ],
    )

    result = inspect(path)

    assert result.sheet_count == 1
    sheet = result.sheets[0]
    assert sheet.name == "Bảng giá"
    assert sheet.header_row_number == 2
    assert sheet.mapping_status is MappingStatus.READY
    assert sheet.column_mapping == {
        "passenger_name": 1,
        "pnr": 2,
        "ticket_number": 3,
        "net_price": 4,
        "selling_price": 5,
    }
    assert sheet.missing_required_fields == ()
    assert sheet.ambiguous_fields == {}


def test_normalizes_accents_case_whitespace_and_punctuation() -> None:
    assert normalize_header("  GIÁ__BÁN / ") == "gia ban"
    assert normalize_header("Booking---Reference") == "booking reference"
    assert normalize_header(None) == ""


def test_chooses_earliest_highest_scoring_unambiguous_non_empty_row(
    tmp_path: Path,
) -> None:
    rows: list[list[object]] = [[None]] * 4
    rows.extend(
        [
            ["Cost", "Selling Price"],
            ["Passenger Name", "PNR", "Cost Price", "Sale Price"],
            ["Passenger Name", "PNR", "Net Price", "Customer Price"],
        ]
    )
    path = save_workbook(tmp_path / "headers.xlsx", rows)

    sheet = inspect(path).sheets[0]

    assert sheet.header_row_number == 6
    assert sheet.column_mapping["passenger_name"] == 1
    assert sheet.column_mapping["pnr"] == 2


def test_only_first_twenty_five_non_empty_rows_are_header_candidates(
    tmp_path: Path,
) -> None:
    rows = [[f"unrelated {number}"] for number in range(25)]
    rows.append(["Cost Price", "Selling Price"])
    path = save_workbook(tmp_path / "late-header.xlsx", rows)

    sheet = inspect(path).sheets[0]

    assert sheet.header_row_number == 1
    assert sheet.mapping_status is MappingStatus.MAPPING_INCOMPLETE
    assert sheet.missing_required_fields == ("net_price", "selling_price")


def test_reports_missing_required_mapping_without_guessing(tmp_path: Path) -> None:
    path = save_workbook(
        tmp_path / "missing.xlsx",
        [["Passenger Name", "PNR", "Cost Price"], ["A", "ABC123", 10]],
    )

    sheet = inspect(path).sheets[0]

    assert sheet.mapping_status is MappingStatus.MAPPING_INCOMPLETE
    assert sheet.column_mapping["net_price"] == 3
    assert "selling_price" not in sheet.column_mapping
    assert sheet.missing_required_fields == ("selling_price",)


def test_reports_duplicate_alias_as_ambiguous_without_guessing(tmp_path: Path) -> None:
    path = save_workbook(
        tmp_path / "ambiguous.xlsx",
        [["Cost", "Net Price", "Selling Price"], [1, 2, 3]],
    )

    sheet = inspect(path).sheets[0]

    assert sheet.mapping_status is MappingStatus.AMBIGUOUS_MAPPING
    assert "net_price" not in sheet.column_mapping
    assert sheet.ambiguous_fields == {"net_price": (1, 2)}
    assert sheet.missing_required_fields == ()


def test_highest_scoring_ambiguous_row_is_not_hidden_by_partial_row(
    tmp_path: Path,
) -> None:
    path = save_workbook(
        tmp_path / "ambiguous-priority.xlsx",
        [
            ["Cost", "Net Price", "Selling Price", "Passenger Name"],
            ["Cost"],
        ],
    )

    sheet = inspect(path).sheets[0]

    assert sheet.header_row_number == 1
    assert sheet.mapping_status is MappingStatus.AMBIGUOUS_MAPPING
    assert sheet.ambiguous_fields == {"net_price": (1, 2)}


def test_ready_header_is_preferred_over_higher_scoring_ambiguous_row(
    tmp_path: Path,
) -> None:
    path = save_workbook(
        tmp_path / "ready-priority.xlsx",
        [
            ["Passenger Name", "PNR", "Cost", "Net Price", "Selling Price"],
            ["Cost Price", "Selling Price"],
        ],
    )

    sheet = inspect(path).sheets[0]

    assert sheet.header_row_number == 2
    assert sheet.mapping_status is MappingStatus.READY
    assert sheet.column_mapping == {"net_price": 1, "selling_price": 2}


def test_inspects_visible_sheets_only(tmp_path: Path) -> None:
    path = tmp_path / "visible.xlsx"
    workbook = Workbook()
    visible = workbook.active
    visible.title = "Visible"
    visible.append(["Cost Price", "Selling Price"])
    hidden = workbook.create_sheet("Hidden")
    hidden.append(["Cost Price", "Selling Price"])
    hidden.sheet_state = "hidden"
    workbook.save(path)
    workbook.close()

    result = inspect(path)

    assert result.sheet_count == 1
    assert [sheet.name for sheet in result.sheets] == ["Visible"]


@pytest.mark.parametrize("suffix", [".xls", ".xlsm", ".csv", ""])
def test_rejects_non_xlsx_extensions(tmp_path: Path, suffix: str) -> None:
    path = tmp_path / f"workbook{suffix}"
    path.write_bytes(b"not relevant")

    with pytest.raises(WorkbookValidationError) as captured:
        inspect(path)

    assert captured.value.code == "UNSUPPORTED_FILE_TYPE"


def test_rejects_corrupt_and_non_zip_files_without_exposing_path(tmp_path: Path) -> None:
    path = tmp_path / "secret-customer-name.xlsx"
    path.write_bytes(b"not a zip")

    with pytest.raises(WorkbookValidationError) as captured:
        inspect(path)

    error = captured.value
    assert error.code == "INVALID_XLSX"
    assert "secret-customer-name" not in str(error)
    assert "secret-customer-name" not in repr(error.details)


def test_rejects_corrupt_xml_as_a_safe_domain_error(tmp_path: Path) -> None:
    source = save_workbook(
        tmp_path / "source.xlsx", [["Cost Price", "Selling Price"]]
    )
    path = tmp_path / "corrupt-content.xlsx"
    with zipfile.ZipFile(source) as original, zipfile.ZipFile(
        path, "w", compression=zipfile.ZIP_DEFLATED
    ) as corrupted:
        for member in original.infolist():
            content = original.read(member)
            if member.filename == "xl/workbook.xml":
                content = b"<workbook><broken>"
            corrupted.writestr(member, content)

    with pytest.raises(WorkbookValidationError) as captured:
        inspect(path)

    assert captured.value.code == "INVALID_XLSX"
    assert captured.value.details == {"reason": "INVALID_WORKBOOK_CONTENT"}


def test_rejects_empty_workbook(tmp_path: Path) -> None:
    path = save_workbook(tmp_path / "empty.xlsx")

    with pytest.raises(WorkbookValidationError) as captured:
        inspect(path)

    assert captured.value.code == "INVALID_XLSX"
    assert captured.value.details == {"reason": "EMPTY_WORKBOOK"}


@pytest.mark.parametrize(
    ("rows", "max_rows", "max_columns"),
    [
        ([["Cost", "Selling Price"], [1, 2], [3, 4]], 2, 20),
        ([["Cost", "Selling Price", "extra"]], 20, 2),
    ],
)
def test_rejects_configured_sheet_limits(
    tmp_path: Path,
    rows: list[list[object]],
    max_rows: int,
    max_columns: int,
) -> None:
    path = save_workbook(tmp_path / "large.xlsx", rows)

    with pytest.raises(WorkbookValidationError) as captured:
        inspect(path, max_rows=max_rows, max_columns=max_columns)

    assert captured.value.code == "WORKBOOK_LIMIT_EXCEEDED"
    assert captured.value.details["max_rows"] == max_rows
    assert captured.value.details["max_columns"] == max_columns


def test_ignores_formatting_only_trailing_xlsx_columns(tmp_path: Path) -> None:
    path = tmp_path / "styled-trailing-columns.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Cost Price", "Selling Price"])
    sheet.append([1_000_000, 1_200_000])
    sheet.cell(1, 256).number_format = "0.00"
    workbook.save(path)
    workbook.close()

    result = inspect(path, max_columns=2)

    assert result.sheets[0].max_column == 2


def test_rejects_unsafe_zip_member_path_without_disclosing_member(tmp_path: Path) -> None:
    path = save_workbook(
        tmp_path / "unsafe.xlsx", [["Cost Price", "Selling Price"]]
    )
    with zipfile.ZipFile(path, "a") as archive:
        archive.writestr("../../private/customer.txt", "secret")

    with pytest.raises(WorkbookValidationError) as captured:
        inspect(path)

    error = captured.value
    assert error.code == "UNSAFE_XLSX_ARCHIVE"
    assert error.details == {"reason": "UNSAFE_MEMBER_PATH"}
    assert "customer" not in str(error)
    assert "customer" not in repr(error.details)


def test_rejects_suspicious_compression_ratio(tmp_path: Path) -> None:
    path = tmp_path / "compressed.xlsx"
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("xl/repeated.xml", b"0" * (1024 * 1024))

    with pytest.raises(WorkbookValidationError) as captured:
        inspect(path)

    assert captured.value.code == "UNSAFE_XLSX_ARCHIVE"
    assert captured.value.details == {"reason": "SUSPICIOUS_COMPRESSION"}


def test_rejects_macro_content_even_with_xlsx_extension(tmp_path: Path) -> None:
    path = save_workbook(
        tmp_path / "macro.xlsx", [["Cost Price", "Selling Price"]]
    )
    with zipfile.ZipFile(path, "a") as archive:
        archive.writestr("xl/vbaProject.bin", b"macro")

    with pytest.raises(WorkbookValidationError) as captured:
        inspect(path)

    assert captured.value.code == "UNSUPPORTED_FILE_TYPE"
    assert captured.value.details == {"reason": "MACRO_CONTENT"}


def test_rejects_encrypted_archive_member(tmp_path: Path) -> None:
    path = save_workbook(
        tmp_path / "encrypted.xlsx", [["Cost Price", "Selling Price"]]
    )
    content = bytearray(path.read_bytes())
    offset = 0
    while True:
        offset = content.find(b"PK\x03\x04", offset)
        if offset < 0:
            break
        flags = struct.unpack_from("<H", content, offset + 6)[0]
        struct.pack_into("<H", content, offset + 6, flags | 0x1)
        offset += 4
    offset = 0
    while True:
        offset = content.find(b"PK\x01\x02", offset)
        if offset < 0:
            break
        flags = struct.unpack_from("<H", content, offset + 8)[0]
        struct.pack_into("<H", content, offset + 8, flags | 0x1)
        offset += 4
    path.write_bytes(content)

    with pytest.raises(WorkbookValidationError) as captured:
        inspect(path)

    assert captured.value.code == "INVALID_XLSX"
    assert captured.value.details == {"reason": "ENCRYPTED_WORKBOOK"}


def test_generated_workbook_validation_reopens_valid_output(tmp_path: Path) -> None:
    path = save_workbook(tmp_path / "generated.xlsx", [["any", "content"]])

    assert validate_generated_workbook(path) is None


def test_generated_workbook_validation_rejects_corruption(tmp_path: Path) -> None:
    path = tmp_path / "generated.xlsx"
    path.write_bytes(b"corrupt")

    with pytest.raises(WorkbookValidationError) as captured:
        validate_generated_workbook(path)

    assert captured.value.code == "INVALID_XLSX"


def test_positive_complexity_limits_are_required(tmp_path: Path) -> None:
    path = save_workbook(tmp_path / "valid.xlsx", [["Cost", "Selling Price"]])

    with pytest.raises(ValueError, match="positive"):
        inspect(path, max_rows=0)

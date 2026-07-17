"""Focused safety tests for workbook price mutation."""

from __future__ import annotations

import hashlib
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Protection

from services.workbook_mutation import (
    MAX_SAFE_VND,
    PriceChange,
    WorkbookMutationError,
    add_workbook_column,
    apply_price_changes,
    remove_workbook_column,
    update_workbook_column,
)


def _make_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Tickets"
    worksheet.append(["Passenger", "PNR", "Cost Price", "Selling Price", "Notes"])
    worksheet.append(["Nguyen A", "ABC123", 1_000_000, 1_200_000, "keep me"])
    worksheet.append(["Tran B", "DEF456", 2_000_000, 2_300_000, "untouched"])
    workbook.save(path)
    workbook.close()


def _apply(source: Path, output: Path, changes: list[PriceChange]):
    return apply_price_changes(
        source,
        output,
        sheet_name="Tickets",
        header_row_number=1,
        column_mapping={"net_price": 3, "selling_price": 4},
        changes=changes,
    )


def test_changes_only_approved_cells_and_preserves_source_bytes(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "version-2.xlsx"
    _make_workbook(source)
    source_before = source.read_bytes()
    source_checksum = hashlib.sha256(source_before).hexdigest()

    result = _apply(
        source,
        output,
        [PriceChange(row_number=2, net_price=1_050_000, selling_price=1_250_000)],
    )

    assert source.read_bytes() == source_before
    assert hashlib.sha256(source.read_bytes()).hexdigest() == source_checksum
    assert result.changed_cell_count == 2
    assert [(change.field, change.old_value, change.new_value) for change in result.changes] == [
        ("net_price", 1_000_000, 1_050_000),
        ("selling_price", 1_200_000, 1_250_000),
    ]
    workbook = load_workbook(output, data_only=False)
    worksheet = workbook["Tickets"]
    assert [worksheet.cell(2, column).value for column in range(1, 6)] == [
        "Nguyen A", "ABC123", 1_050_000, 1_250_000, "keep me"
    ]
    assert [worksheet.cell(3, column).value for column in range(1, 6)] == [
        "Tran B", "DEF456", 2_000_000, 2_300_000, "untouched"
    ]
    workbook.close()


def test_output_is_a_reopenable_workbook(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_workbook(source)
    _apply(source, output, [PriceChange(row_number=2, selling_price=1_300_000)])

    workbook = load_workbook(output, read_only=False, data_only=False)
    assert workbook["Tickets"]["D2"].value == 1_300_000
    workbook.close()


def test_preserves_other_sheets_formulas_and_styles(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_workbook(source)
    workbook = load_workbook(source)
    worksheet = workbook["Tickets"]
    worksheet["E2"] = "=C2+D2"
    worksheet["E2"].fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    summary = workbook.create_sheet("Summary")
    summary["A1"] = "=Tickets!D2"
    workbook.save(source)
    workbook.close()

    _apply(source, output, [PriceChange(row_number=2, net_price=1_050_000)])

    generated = load_workbook(output, data_only=False)
    assert generated.sheetnames == ["Tickets", "Summary"]
    assert generated["Tickets"]["E2"].value == "=C2+D2"
    assert generated["Tickets"]["E2"].fill.fgColor.rgb == "00FFF2CC"
    assert generated["Summary"]["A1"].value == "=Tickets!D2"
    generated.close()


def test_generic_changes_validate_types_and_preserve_formats(tmp_path: Path) -> None:
    source = tmp_path / "generic-source.xlsx"
    output = tmp_path / "generic-output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(["Text", "Number", "Currency", "Date", "Boolean", "Blank"])
    worksheet.append(["old", 1, 12.5, date(2026, 1, 1), False, "clear me"])
    worksheet["C2"].number_format = '$#,##0.00'
    worksheet["D2"].number_format = "dd/mm/yyyy"
    worksheet["A2"].fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    workbook.save(source)
    workbook.close()

    column_config = [
        {"id": "source-text", "column_number": 1, "data_type": "text"},
        {"id": "source-number", "column_number": 2, "data_type": "number"},
        {"id": "source-currency", "column_number": 3, "data_type": "currency"},
        {"id": "source-date", "column_number": 4, "data_type": "date"},
        {"id": "source-boolean", "column_number": 5, "data_type": "boolean"},
        {"id": "source-blank", "column_number": 6, "data_type": "text"},
    ]
    result = apply_price_changes(
        source,
        output,
        sheet_name="Data",
        header_row_number=1,
        column_mapping={},
        column_config=column_config,
        changes=[
            PriceChange(
                row_number=2,
                values={
                    "source-text": "updated",
                    "source-number": -2.75,
                    "source-currency": -1234.56,
                    "source-date": "2026-07-14",
                    "source-boolean": True,
                    "source-blank": None,
                },
            )
        ],
    )

    assert result.changed_cell_count == 6
    generated = load_workbook(output, data_only=False)
    row = generated["Data"]
    assert row["A2"].value == "updated"
    assert row["B2"].value == -2.75
    assert row["C2"].value == -1234.56
    assert row["D2"].value.date() == date(2026, 7, 14)
    assert row["E2"].value is True
    assert row["F2"].value is None
    assert row["C2"].number_format == '$#,##0.00'
    assert row["D2"].number_format == "dd/mm/yyyy"
    assert row["A2"].fill.fgColor.rgb == "00FFF2CC"
    generated.close()


@pytest.mark.parametrize("value", ["=1+1", '=HYPERLINK("https://example.com", "link")'])
def test_configured_text_changes_reject_formula_injection(
    tmp_path: Path,
    value: str,
) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_workbook(source)

    with pytest.raises(WorkbookMutationError) as raised:
        apply_price_changes(
            source,
            output,
            sheet_name="Tickets",
            header_row_number=1,
            column_mapping={},
            column_config=[
                {"id": "source-notes", "column_number": 5, "data_type": "text"}
            ],
            changes=[PriceChange(row_number=2, values={"source-notes": value})],
        )

    assert raised.value.code == "INVALID_CELL_VALUE"
    assert not output.exists()


def test_configured_text_changes_preserve_normal_text(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_workbook(source)

    apply_price_changes(
        source,
        output,
        sheet_name="Tickets",
        header_row_number=1,
        column_mapping={},
        column_config=[
            {"id": "source-notes", "column_number": 5, "data_type": "text"}
        ],
        changes=[
            PriceChange(
                row_number=2,
                values={"source-notes": "Need invoice = confirmed"},
            )
        ],
    )

    workbook = load_workbook(output, data_only=False)
    assert workbook["Tickets"]["E2"].value == "Need invoice = confirmed"
    workbook.close()


def test_configured_changes_keep_imported_formula_cells_non_editable(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_workbook(source)
    workbook = load_workbook(source)
    workbook["Tickets"]["E2"] = "=C2+D2"
    workbook.save(source)
    workbook.close()

    with pytest.raises(WorkbookMutationError) as raised:
        apply_price_changes(
            source,
            output,
            sheet_name="Tickets",
            header_row_number=1,
            column_mapping={},
            column_config=[
                {"id": "source-notes", "column_number": 5, "data_type": "text"}
            ],
            changes=[
                PriceChange(row_number=2, values={"source-notes": "updated"})
            ],
        )

    assert raised.value.code == "CELL_NOT_EDITABLE"
    assert not output.exists()


@pytest.mark.parametrize(
    ("column_id", "semantic_field", "value"),
    [
        ("source-cost", "net_price", 0),
        ("source-sale", "selling_price", 1_250_000),
    ],
)
def test_semantic_vnd_prices_accept_whole_non_negative_values(
    tmp_path: Path,
    column_id: str,
    semantic_field: str,
    value: int,
) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_workbook(source)
    column_number = 3 if semantic_field == "net_price" else 4

    apply_price_changes(
        source,
        output,
        sheet_name="Tickets",
        header_row_number=1,
        column_mapping={},
        column_config=[
            {
                "id": column_id,
                "column_number": column_number,
                "data_type": "currency",
                "semantic_field": semantic_field,
            }
        ],
        changes=[PriceChange(row_number=2, values={column_id: value})],
    )

    workbook = load_workbook(output, data_only=False)
    assert workbook["Tickets"].cell(2, column_number).value == value
    workbook.close()


@pytest.mark.parametrize("semantic_field", ["net_price", "selling_price"])
@pytest.mark.parametrize("value", [-1, 1.5, MAX_SAFE_VND + 1])
def test_semantic_vnd_prices_reject_invalid_values(
    tmp_path: Path,
    semantic_field: str,
    value: object,
) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_workbook(source)

    with pytest.raises(WorkbookMutationError) as raised:
        apply_price_changes(
            source,
            output,
            sheet_name="Tickets",
            header_row_number=1,
            column_mapping={},
            column_config=[
                {
                    "id": f"source-{semantic_field}",
                    "column_number": 3 if semantic_field == "net_price" else 4,
                    "data_type": "currency",
                    "semantic_field": semantic_field,
                }
            ],
            changes=[
                PriceChange(
                    row_number=2,
                    values={f"source-{semantic_field}": value},
                )
            ],
        )

    assert raised.value.code == "INVALID_CELL_VALUE"
    assert not output.exists()


def test_rejects_duplicate_configured_cell_targets_after_alias_resolution(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_workbook(source)

    with pytest.raises(WorkbookMutationError) as raised:
        apply_price_changes(
            source,
            output,
            sheet_name="Tickets",
            header_row_number=1,
            column_mapping={},
            column_config=[
                {
                    "id": "source-cost",
                    "column_number": 3,
                    "data_type": "currency",
                    "semantic_field": "net_price",
                }
            ],
            changes=[
                PriceChange(
                    row_number=2,
                    values={"source-cost": 1_050_000, "net_price": 1_100_000},
                )
            ],
        )

    assert raised.value.code == "INVALID_CELL_VALUE"
    assert not output.exists()


@pytest.mark.parametrize(
    ("data_type", "value"),
    [
        ("date", "2026-07-14T10:00:00Z"),
        ("text", "x" * 32_768),
        ("number", 1.234567890123456),
    ],
)
def test_generic_changes_reject_unrepresentable_excel_values(
    tmp_path: Path,
    data_type: str,
    value: object,
) -> None:
    source = tmp_path / "generic-source.xlsx"
    output = tmp_path / "generic-output.xlsx"
    _make_workbook(source)

    with pytest.raises(WorkbookMutationError) as raised:
        apply_price_changes(
            source,
            output,
            sheet_name="Tickets",
            header_row_number=1,
            column_mapping={},
            column_config=[
                {"id": "source-value", "column_number": 3, "data_type": data_type}
            ],
            changes=[PriceChange(row_number=2, values={"source-value": value})],
        )

    assert raised.value.code == "INVALID_CELL_VALUE"
    assert not output.exists()


def test_generic_changes_reject_type_mismatches(tmp_path: Path) -> None:
    source = tmp_path / "generic-source.xlsx"
    output = tmp_path / "generic-output.xlsx"
    _make_workbook(source)

    with pytest.raises(WorkbookMutationError) as raised:
        apply_price_changes(
            source,
            output,
            sheet_name="Tickets",
            header_row_number=1,
            column_mapping={},
            column_config=[
                {"id": "source-number", "column_number": 3, "data_type": "number"}
            ],
            changes=[PriceChange(row_number=2, values={"source-number": "not a number"})],
        )

    assert raised.value.code == "INVALID_CELL_VALUE"
    assert not output.exists()


@pytest.mark.parametrize("field", ["net_price", "selling_price"])
def test_rejects_formula_price_cells(tmp_path: Path, field: str) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_workbook(source)
    workbook = load_workbook(source)
    column = 3 if field == "net_price" else 4
    workbook["Tickets"].cell(2, column).value = "=1+1"
    workbook.save(source)
    workbook.close()

    change = PriceChange(row_number=2, **{field: 2_000_000})
    with pytest.raises(WorkbookMutationError) as raised:
        _apply(source, output, [change])
    assert raised.value.code == "CELL_NOT_EDITABLE"
    assert not output.exists()


def test_rejects_merged_price_cell(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_workbook(source)
    workbook = load_workbook(source)
    workbook["Tickets"].merge_cells("C2:C3")
    workbook.save(source)
    workbook.close()

    with pytest.raises(WorkbookMutationError) as raised:
        _apply(source, output, [PriceChange(row_number=2, net_price=2_000_000)])
    assert raised.value.code == "CELL_NOT_EDITABLE"


def test_rejects_locked_cell_only_when_sheet_is_protected(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_workbook(source)
    workbook = load_workbook(source)
    worksheet = workbook["Tickets"]
    worksheet["C2"].protection = Protection(locked=True)
    worksheet.protection.sheet = True
    workbook.save(source)
    workbook.close()

    with pytest.raises(WorkbookMutationError) as raised:
        _apply(source, output, [PriceChange(row_number=2, net_price=2_000_000)])
    assert raised.value.code == "CELL_NOT_EDITABLE"


@pytest.mark.parametrize(
    "value",
    [-1, 1.5, float("nan"), float("inf"), MAX_SAFE_VND + 1, True, "1000"],
)
def test_rejects_invalid_price_values(tmp_path: Path, value: object) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_workbook(source)

    with pytest.raises(WorkbookMutationError) as raised:
        _apply(source, output, [PriceChange(row_number=2, net_price=value)])  # type: ignore[arg-type]
    assert raised.value.code == "INVALID_CELL_VALUE"
    assert not output.exists()


def test_rejects_duplicate_and_invalid_rows(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _make_workbook(source)

    with pytest.raises(WorkbookMutationError) as duplicate:
        _apply(
            source,
            tmp_path / "duplicate.xlsx",
            [PriceChange(2, net_price=10), PriceChange(2, selling_price=20)],
        )
    assert duplicate.value.code == "INVALID_ROW"

    with pytest.raises(WorkbookMutationError) as header:
        _apply(source, tmp_path / "header.xlsx", [PriceChange(1, net_price=10)])
    assert header.value.code == "INVALID_ROW"


def test_max_changes_limits_changed_cells_not_only_rows(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_workbook(source)

    with pytest.raises(WorkbookMutationError) as raised:
        apply_price_changes(
            source,
            output,
            sheet_name="Tickets",
            header_row_number=1,
            column_mapping={"net_price": 3, "selling_price": 4},
            changes=[PriceChange(2, net_price=10, selling_price=20)],
            max_changes=1,
        )

    assert raised.value.code == "INVALID_ROW"
    assert not output.exists()


def test_rejects_existing_output_without_overwrite(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_workbook(source)
    output.write_bytes(b"existing immutable version")

    with pytest.raises(WorkbookMutationError) as raised:
        _apply(source, output, [PriceChange(row_number=2, net_price=2_000_000)])
    assert raised.value.code == "STORAGE_WRITE_FAILED"
    assert output.read_bytes() == b"existing immutable version"


def test_failure_cleans_temporary_files(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_workbook(source)

    def fail_link(source_path: Path, output_path: Path) -> None:
        del source_path, output_path
        raise OSError("publish failed")

    monkeypatch.setattr("services.workbook_mutation.os.link", fail_link)
    with pytest.raises(WorkbookMutationError) as raised:
        _apply(source, output, [PriceChange(row_number=2, net_price=2_000_000)])
    assert raised.value.code == "STORAGE_WRITE_FAILED"
    assert not output.exists()
    assert sorted(path.name for path in tmp_path.iterdir()) == ["source.xlsx"]


def test_temp_cleanup_failure_does_not_report_published_output_as_failed(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_workbook(source)
    original_unlink = Path.unlink

    def fail_temp_unlink(path: Path, *args, **kwargs) -> None:
        if path.name.startswith(".output.xlsx."):
            raise OSError("cleanup failed")
        original_unlink(path, *args, **kwargs)

    monkeypatch.setattr(Path, "unlink", fail_temp_unlink)

    result = _apply(
        source,
        output,
        [PriceChange(row_number=2, net_price=1_100_000)],
    )

    assert result.changed_cell_count == 1
    assert output.exists()


def test_output_directory_failure_uses_safe_domain_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "private-customer" / "output.xlsx"
    _make_workbook(source)

    def fail_mkdir(*_args, **_kwargs) -> None:
        raise OSError(f"permission denied: {output.parent}")

    monkeypatch.setattr(Path, "mkdir", fail_mkdir)

    with pytest.raises(WorkbookMutationError) as raised:
        _apply(source, output, [PriceChange(row_number=2, net_price=1_100_000)])

    assert raised.value.code == "STORAGE_WRITE_FAILED"
    assert "private-customer" not in str(raised.value)


def test_structural_changes_regenerate_formula_references_and_calc_flags(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    with_formula = tmp_path / "with-formula.xlsx"
    moved = tmp_path / "moved.xlsx"
    updated = tmp_path / "updated.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(["Unused", "Fare", "Sale"])
    worksheet.append(["x", 100, 150])
    workbook.save(source)
    workbook.close()

    formula = {
        "schema_version": 1,
        "expression": {
            "type": "binary",
            "operator": "-",
            "left": {"type": "column", "column_id": "sale"},
            "right": {"type": "column", "column_id": "fare"},
        },
    }
    before = [
        {"id": "unused", "label": "Unused", "column_number": 1, "data_type": "text"},
        {"id": "fare", "label": "Fare", "column_number": 2, "data_type": "currency"},
        {"id": "sale", "label": "Sale", "column_number": 3, "data_type": "currency"},
        {"id": "profit", "label": "Profit", "column_number": 4, "data_type": "currency", "formula": formula},
    ]
    add_workbook_column(
        source,
        with_formula,
        sheet_name="Data",
        header_row_number=1,
        label="Profit",
        column_config=before,
    )
    generated = load_workbook(with_formula, data_only=False)
    assert generated["Data"]["D2"].value == "=(C2 - B2)"
    assert generated.calculation.calcMode == "auto"
    assert generated.calculation.fullCalcOnLoad is True
    generated.close()
    after_remove = [
        {**item, "column_number": int(item["column_number"]) - 1}
        for item in before
        if item["id"] != "unused"
    ]
    remove_workbook_column(
        with_formula,
        moved,
        sheet_name="Data",
        column_number=1,
        header_row_number=1,
        column_config=after_remove,
    )
    generated = load_workbook(moved, data_only=False)
    assert generated["Data"]["C2"].value == "=(B2 - A2)"
    generated.close()

    updated_formula = {
        "schema_version": 1,
        "expression": {
            "type": "function",
            "function": "MAX",
            "arguments": [
                {"type": "column", "column_id": "fare"},
                {"type": "column", "column_id": "sale"},
                {"type": "constant", "value": "0"},
            ],
        },
    }
    after_remove[-1] = {**after_remove[-1], "label": "Maximum", "formula": updated_formula}
    update_workbook_column(
        moved,
        updated,
        sheet_name="Data",
        header_row_number=1,
        column_number=3,
        label="Maximum",
        column_config=after_remove,
    )
    generated = load_workbook(updated, data_only=False)
    assert generated["Data"]["C1"].value == "Maximum"
    assert generated["Data"]["C2"].value == "=MAX(A2,B2,0)"
    generated.close()


def test_add_column_uses_meaningful_bound_not_formatting_dimension(tmp_path: Path) -> None:
    source = tmp_path / "formatted-source.xlsx"
    output = tmp_path / "bounded-output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(["Name", "Value"])
    worksheet.append(["A", 1])
    worksheet.cell(1, 256).fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    workbook.save(source)
    workbook.close()

    result = add_workbook_column(
        source,
        output,
        sheet_name="Data",
        header_row_number=1,
        label="Note",
        meaningful_max_row=2,
        meaningful_max_column=2,
    )

    assert result.column_number == 3
    generated = load_workbook(output)
    assert generated["Data"]["C1"].value == "Note"
    generated.close()

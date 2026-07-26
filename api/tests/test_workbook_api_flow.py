"""End-to-end HTTP flow for Workbook Editor V2 backend integration."""

from __future__ import annotations

import io
import uuid
from pathlib import Path
from collections.abc import Generator

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlmodel import Session, SQLModel, create_engine

from core.auth import get_current_user
from database import get_session
from models import User, UserRole
from routes import workbooks as workbook_routes
from storage.workbooks import LocalWorkbookStorage


def _workbook_bytes() -> bytes:
    stream = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Tickets"
    worksheet.append(["Hành khách", "Mã chỗ", "Giá gốc", "Giá bán"])
    worksheet.append(["NGUYỄN VĂN A", "ABC123", 1_000_000, 1_200_000])
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def test_authenticated_upload_edit_and_download_flow(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    engine = create_engine(
        f"sqlite:///{tmp_path / 'workbook-api.db'}",
        connect_args={"check_same_thread": False},
    )
    SQLModel.metadata.create_all(engine)
    actor = User(
        id=uuid.uuid4(),
        username="workbook-api-user",
        hashed_password="test-only",
        role=UserRole.STAFF,
    )
    other_staff = User(
        id=uuid.uuid4(),
        username="other-workbook-user",
        hashed_password="test-only",
        role=UserRole.STAFF,
    )
    admin = User(
        id=uuid.uuid4(),
        username="workbook-admin",
        hashed_password="test-only",
        role=UserRole.ADMIN,
    )
    with Session(engine) as db:
        db.add_all([actor, other_staff, admin])
        db.commit()
        for user in (actor, other_staff, admin):
            db.refresh(user)

    def override_session() -> Generator[Session, None, None]:
        with Session(engine) as db:
            yield db

    application = FastAPI()
    application.include_router(
        workbook_routes.router,
        prefix="/api/v1/workbooks",
    )
    application.dependency_overrides[get_session] = override_session
    current_actor = [actor]
    application.dependency_overrides[get_current_user] = lambda: current_actor[0]
    storage = LocalWorkbookStorage(tmp_path / "storage")
    monkeypatch.setattr(workbook_routes, "_workbook_storage", lambda: storage)
    client = TestClient(application)
    original = _workbook_bytes()

    upload = client.post(
        "/api/v1/workbooks/uploads",
        files={
            "file": (
                "prices.xlsx",
                original,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload.status_code == 201
    uploaded = upload.json()["data"]
    assert uploaded["sheets"][0]["mapping_status"] == "READY"

    create_session = client.post(
        "/api/v1/workbooks/sessions",
        json={
            "workbook_id": uploaded["id"],
            "sheet_name": "Tickets",
        },
    )
    assert create_session.status_code == 201
    session_data = create_session.json()["data"]

    current_actor[0] = other_staff
    for method, path in (
        ("get", f"/api/v1/workbooks/sessions/{session_data['id']}/records"),
        ("get", f"/api/v1/workbooks/sessions/{session_data['id']}/download"),
    ):
        assert getattr(client, method)(path).status_code == 404

    current_actor[0] = admin
    assert client.get(
        f"/api/v1/workbooks/sessions/{session_data['id']}"
    ).status_code == 200
    current_actor[0] = actor

    records = client.get(
        f"/api/v1/workbooks/sessions/{session_data['id']}/records",
        params={"search": "nguyen van a"},
    )
    assert records.status_code == 200
    assert records.json()["data"]["items"][0]["row_number"] == 2

    save = client.post(
        f"/api/v1/workbooks/sessions/{session_data['id']}/saves",
        json={
            "request_id": str(uuid.uuid4()),
            "base_version": 1,
            "changes": [
                {
                    "row_number": 2,
                    "values": {
                        "net_price": 1_050_000,
                        "selling_price": 1_250_000,
                    },
                }
            ],
        },
    )
    assert save.status_code == 201
    assert save.json()["data"]["current_version"] == 2

    download = client.get(
        f"/api/v1/workbooks/sessions/{session_data['id']}/download"
    )
    assert download.status_code == 200
    assert download.headers["x-workbook-version"] == "2"

    edited = load_workbook(io.BytesIO(download.content), data_only=False)
    assert edited["Tickets"]["C2"].value == 1_050_000
    assert edited["Tickets"]["D2"].value == 1_250_000
    edited.close()

"""Focused tests for readable XLSX download formatting."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from services.workbook_export import format_workbook_for_export


def _column_config() -> list[dict[str, object]]:
    return [
        {
            "id": "name",
            "label": "Passenger name",
            "column_number": 1,
            "data_type": "text",
        },
        {
            "id": "ticket",
            "label": "Ticket number",
            "column_number": 2,
            "data_type": "text",
        },
        {
            "id": "amount",
            "label": "Total amount",
            "column_number": 3,
            "data_type": "currency",
        },
        {
            "id": "remark",
            "label": "Remark",
            "column_number": 4,
            "data_type": "text",
        },
    ]


def test_formats_download_copy_without_changing_values_or_other_sheets(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Tickets"
    worksheet.append(["Passenger name", "Ticket number", "Total amount", "Remark"])
    worksheet.append(["Nguyễn An", 7_382_321_366_982, 1_200_000, "Paid"])
    worksheet.append(
        [
            "HIRABAYASHI/YOSHITAKA",
            7_382_321_455_459,
            -250_000,
            (
                "A long finance note that should wrap inside a capped column "
                "instead of making the worksheet extremely wide."
            ),
        ]
    )
    worksheet["A3"].font = Font(color="FF0000")
    worksheet.column_dimensions["B"].hidden = True
    summary = workbook.create_sheet("Summary")
    summary["A1"] = "=Tickets!C2"
    workbook.save(source)
    workbook.close()

    format_workbook_for_export(
        source,
        output,
        sheet_name="Tickets",
        header_row_number=1,
        column_config=_column_config(),
        meaningful_max_row=3,
        meaningful_max_column=4,
    )

    generated = load_workbook(output, data_only=False)
    tickets = generated["Tickets"]
    assert tickets["A2"].value == "Nguyễn An"
    assert tickets["B2"].value == 7_382_321_366_982
    assert tickets["C3"].value == -250_000
    assert generated["Summary"]["A1"].value == "=Tickets!C2"
    assert generated["Summary"].sheet_view.showGridLines is not False

    assert tickets["A1"].fill.fgColor.rgb == "001B61C9"
    assert tickets["A1"].font.bold is True
    assert tickets["A1"].font.color.rgb == "00FFFFFF"
    assert tickets["A2"].font.name == "Aptos"
    assert tickets["A3"].font.color.rgb == "00FF0000"
    assert tickets["A3"].fill.fgColor.rgb == "00F7F9FC"
    assert tickets["D3"].alignment.wrap_text is True
    assert tickets.row_dimensions[3].height > 20

    assert 20 < tickets.column_dimensions["A"].width <= 44
    assert tickets.column_dimensions["B"].width >= 15
    assert tickets.column_dimensions["B"].hidden is True
    assert tickets.column_dimensions["D"].width == 44
    assert tickets["B2"].number_format == "@"
    assert tickets["C2"].number_format == "#,##0"
    assert tickets.freeze_panes == "A2"
    assert tickets.auto_filter.ref == "A1:D3"
    assert tickets.sheet_view.showGridLines is False
    assert tickets.sheet_view.zoomScale == 90
    generated.close()


def test_styles_grouped_headers_and_freezes_below_the_header_band(
    tmp_path: Path,
) -> None:
    source = tmp_path / "grouped-source.xlsx"
    output = tmp_path / "grouped-output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Tickets"
    worksheet.merge_cells("A1:A2")
    worksheet.merge_cells("B1:C1")
    worksheet["A1"] = "Passenger"
    worksheet["B1"] = "Prices"
    worksheet["B2"] = "Net"
    worksheet["C2"] = "Selling"
    worksheet.append(["Nguyễn An", 1_000_000, 1_200_000])
    workbook.save(source)
    workbook.close()

    config = [
        {
            "id": "passenger",
            "label": "Passenger",
            "column_number": 1,
            "data_type": "text",
        },
        {
            "id": "net",
            "label": "Net",
            "column_number": 2,
            "data_type": "currency",
        },
        {
            "id": "selling",
            "label": "Selling",
            "column_number": 3,
            "data_type": "currency",
        },
    ]
    format_workbook_for_export(
        source,
        output,
        sheet_name="Tickets",
        header_row_number=1,
        column_config=config,
        meaningful_max_row=3,
        meaningful_max_column=3,
    )

    generated = load_workbook(output)
    worksheet = generated["Tickets"]
    assert worksheet["A1"].fill.fgColor.rgb == "001B61C9"
    assert worksheet["B1"].fill.fgColor.rgb == "001B61C9"
    assert worksheet["B2"].fill.fgColor.rgb == "00EDF5FF"
    assert worksheet["B2"].font.color.rgb == "00181D26"
    assert worksheet.freeze_panes == "A3"
    assert worksheet.auto_filter.ref is None
    assert worksheet.print_title_rows == "$1:$2"
    generated.close()


def test_formats_each_detected_data_sheet(tmp_path: Path) -> None:
    source = tmp_path / "multi-sheet-source.xlsx"
    output = tmp_path / "multi-sheet-output.xlsx"
    workbook = Workbook()
    selected = workbook.active
    selected.title = "Tickets"
    selected.append(["Passenger name", "Ticket number", "Total amount", "Remark"])
    selected.append(["Nguyễn An", "738001", 1_200_000, "Paid"])
    archive = workbook.create_sheet("Archive")
    archive.append(["Name", "TKT.Nbr.", "Total VND"])
    archive.append(
        [
            "A passenger name that needs a wider column",
            "7382321366982",
            2_300_000,
        ]
    )
    workbook.save(source)
    workbook.close()

    format_workbook_for_export(
        source,
        output,
        sheet_name="Tickets",
        header_row_number=1,
        column_config=_column_config(),
        meaningful_max_row=2,
        meaningful_max_column=4,
        sheet_metadata=[
            {
                "name": "Tickets",
                "header_row_number": 1,
                "max_row": 2,
                "max_column": 4,
                "detected_headers": [
                    "Passenger name",
                    "Ticket number",
                    "Total amount",
                    "Remark",
                ],
            },
            {
                "name": "Archive",
                "header_row_number": 1,
                "max_row": 2,
                "max_column": 3,
                "detected_headers": ["Name", "TKT.Nbr.", "Total VND"],
            },
        ],
    )

    generated = load_workbook(output)
    archive = generated["Archive"]
    assert archive["A1"].fill.fgColor.rgb == "001B61C9"
    assert archive["B2"].number_format == "@"
    assert archive["C2"].number_format == "#,##0"
    assert archive.column_dimensions["A"].width > 35
    assert archive.freeze_panes == "A2"
    assert archive.sheet_view.showGridLines is False
    generated.close()

from pathlib import Path

from openpyxl import Workbook, load_workbook

from services.workbook_mutation import PriceChange, add_workbook_column, apply_price_changes
from services.workbook_reader import read_workbook_records


def test_source_cells_are_editable_and_formula_is_row_local(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    with_formula = tmp_path / "formula.xlsx"
    edited = tmp_path / "edited.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Base", "Rate"])
    sheet.append([200, 10])
    workbook.save(source)
    workbook.close()

    add_workbook_column(
        source,
        with_formula,
        sheet_name="Data",
        header_row_number=1,
        label="Commission",
        formula={"left_column_number": 1, "operator": "%", "right_column_number": 2},
    )
    config = [
        {"id": "source-a", "column_number": 1, "origin": "source", "data_type": "number"},
        {"id": "source-b", "column_number": 2, "origin": "source", "data_type": "number"},
        {
            "id": "user-result",
            "column_number": 3,
            "origin": "user",
            "data_type": "number",
            "formula": {"left_column_id": "source-a", "operator": "%", "right_column_id": "source-b"},
        },
    ]
    page = read_workbook_records(
        with_formula,
        sheet_name="Data",
        header_row_number=1,
        column_mapping={},
        column_config=config,
    )
    assert page.records[0].values == {"source-a": 200, "source-b": 10, "user-result": 20}
    assert page.records[0].editable == {"source-a": True, "source-b": True}

    apply_price_changes(
        with_formula,
        edited,
        sheet_name="Data",
        header_row_number=1,
        column_mapping={},
        column_config=config,
        changes=[PriceChange(row_number=2, values={"source-a": 300})],
    )
    updated = read_workbook_records(
        edited, sheet_name="Data", header_row_number=1, column_mapping={}, column_config=config
    )
    assert updated.records[0].values["user-result"] == 30
    generated = load_workbook(edited, data_only=False)
    assert generated["Data"]["C2"].value == "=A2*B2/100"
    generated.close()

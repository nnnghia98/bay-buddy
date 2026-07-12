"""Tests for server-side Workbook Editor V2 record reads."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from services.workbook_reader import WorkbookReadError, read_workbook_records


MAPPING = {
    "passenger_name": 1,
    "pnr": 2,
    "ticket_number": 3,
    "net_price": 4,
    "selling_price": 5,
}


@pytest.fixture()
def workbook_path(tmp_path: Path) -> Path:
    path = tmp_path / "records.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Vé tháng 7"
    worksheet.append(["BÁO CÁO GIÁ VÉ"])
    worksheet.append(["Hành khách", "PNR", "Số vé", "Giá gốc", "Giá bán"])
    worksheet.append(["Nguyễn Ánh", "ABC123", "738001", 1_000_000, 1_200_000])
    worksheet.append([None, None, None, None, None])
    worksheet.append(["Đỗ Bình", "DEF456", "738002", "=600000+500000", 1_300_000])
    worksheet.append(["Tran An", "ZZZ999", "738003", 900_000, 1_100_000])
    worksheet.append(["Trần An", "AAA111", "738004", 900_000, 1_150_000])
    worksheet.append(["Lê Cường", "CCC222", "738005", 1_400_000, None])
    workbook.save(path)
    workbook.close()
    return path


def read(path: Path, **overrides):
    parameters = {
        "sheet_name": "Vé tháng 7",
        "header_row_number": 2,
        "column_mapping": MAPPING,
    }
    parameters.update(overrides)
    return read_workbook_records(path, **parameters)


def test_reads_headers_rows_totals_and_skips_fully_blank_rows(
    workbook_path: Path,
) -> None:
    result = read(workbook_path, page=1, page_size=2)

    assert [column.field for column in result.columns] == list(MAPPING)
    assert [column.header for column in result.columns] == [
        "Hành khách",
        "PNR",
        "Số vé",
        "Giá gốc",
        "Giá bán",
    ]
    assert [record.row_number for record in result.records] == [3, 5]
    assert result.pagination.page == 1
    assert result.pagination.page_size == 2
    assert result.pagination.total == 5
    assert result.pagination.total_pages == 3


def test_preserves_source_header_text_without_semantic_translation(tmp_path: Path) -> None:
    path = tmp_path / "original-headers.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Mixed"
    worksheet.append(["  Hành khách  ", "Booking Code", "Giá NET", None, 2026])
    worksheet.append(["Nguyễn An", "ABC123", 1_000_000, "note", "value"])
    workbook.save(path)
    workbook.close()

    result = read_workbook_records(
        path,
        sheet_name="Mixed",
        header_row_number=1,
        column_mapping={"passenger_name": 1, "pnr": 2, "net_price": 3},
    )

    assert [column.header for column in result.columns] == [
        "  Hành khách  ",
        "Booking Code",
        "Giá NET",
        "",
        "2026",
    ]


def test_preserves_group_headers_and_starts_records_after_child_band(tmp_path: Path) -> None:
    path = tmp_path / "grouped-headers.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Grouped"
    worksheet.append(["Carrier", "ICT2", None, "ICT3", None])
    worksheet.append([None, "%", "Comm", "%", "Comm"])
    worksheet.append(["VN", 5, 100_000, 7, 120_000])
    workbook.save(path)
    workbook.close()

    result = read_workbook_records(
        path,
        sheet_name="Grouped",
        header_row_number=1,
        column_mapping={},
    )

    assert result.header_row_count == 2
    assert [column.header for column in result.columns] == [
        "Carrier", "%", "Comm", "%", "Comm"
    ]
    assert [column.group_label for column in result.columns] == [
        None, "ICT2", "ICT2", "ICT3", "ICT3"
    ]
    assert [column.header_row_span for column in result.columns] == [2, 1, 1, 1, 1]
    assert [record.row_number for record in result.records] == [3]


def test_formula_price_cell_is_not_editable(workbook_path: Path) -> None:
    result = read(workbook_path)
    formula_record = next(record for record in result.records if record.row_number == 5)

    assert formula_record.values["net_price"] == "=600000+500000"
    assert formula_record.editable == {
        "net_price": False,
        "selling_price": True,
    }


def test_merged_price_cells_are_not_editable(tmp_path: Path) -> None:
    path = tmp_path / "merged.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Tickets"
    worksheet.append(["Cost Price", "Selling Price"])
    worksheet.append([1_000_000, 1_200_000])
    worksheet.append([None, 1_300_000])
    worksheet.merge_cells("A2:A3")
    workbook.save(path)
    workbook.close()

    result = read_workbook_records(
        path,
        sheet_name="Tickets",
        header_row_number=1,
        column_mapping={"net_price": 1, "selling_price": 2},
    )

    assert result.records[0].editable["net_price"] is False
    assert result.records[1].editable["net_price"] is False


def test_search_falls_back_to_prices_when_identity_fields_are_unmapped(
    workbook_path: Path,
) -> None:
    result = read_workbook_records(
        workbook_path,
        sheet_name="Vé tháng 7",
        header_row_number=2,
        column_mapping={"net_price": 4, "selling_price": 5},
        search="1200000",
    )

    assert [record.row_number for record in result.records] == [3]


def test_reads_incomplete_mapping_without_price_columns(workbook_path: Path) -> None:
    result = read_workbook_records(
        workbook_path,
        sheet_name="Vé tháng 7",
        header_row_number=2,
        column_mapping={"passenger_name": 1, "pnr": 2},
        search="abc123",
    )

    assert [column.field for column in result.columns] == [
        "passenger_name",
        "pnr",
        "column_3",
        "column_4",
        "column_5",
    ]
    assert result.records[0].values == {
        "passenger_name": "Nguyễn Ánh",
        "pnr": "ABC123",
        "column_3": "738001",
        "column_4": 1_000_000,
        "column_5": 1_200_000,
    }
    assert result.records[0].editable == {}


@pytest.mark.parametrize("search", ["nguyen anh", "NGUYỄN ÁNH", "abc123"])
def test_search_is_unicode_normalized_and_covers_identity_values(
    workbook_path: Path,
    search: str,
) -> None:
    result = read(workbook_path, search=search)

    assert [record.row_number for record in result.records] == [3]
    assert result.pagination.total == 1


def test_search_is_applied_before_pagination(workbook_path: Path) -> None:
    result = read(workbook_path, search="tran an", page=2, page_size=1)

    assert result.pagination.total == 2
    assert result.pagination.total_pages == 2
    assert [record.row_number for record in result.records] == [7]


def test_sort_is_stable_by_physical_row_then_paginated(workbook_path: Path) -> None:
    first_page = read(
        workbook_path,
        sort_by="net_price",
        sort_direction="asc",
        page=1,
        page_size=2,
    )
    second_page = read(
        workbook_path,
        sort_by="net_price",
        sort_direction="asc",
        page=2,
        page_size=2,
    )

    assert [record.row_number for record in first_page.records] == [6, 7]
    assert [record.row_number for record in second_page.records] == [3, 8]


def test_descending_sort_keeps_equal_values_in_physical_order(
    workbook_path: Path,
) -> None:
    result = read(
        workbook_path,
        sort_by="passenger_name",
        sort_direction="desc",
    )

    tran_rows = [
        record.row_number
        for record in result.records
        if record.values["passenger_name"] in {"Tran An", "Trần An"}
    ]
    assert tran_rows == [6, 7]


def test_descending_sort_keeps_blank_values_last(workbook_path: Path) -> None:
    result = read(
        workbook_path,
        sort_by="selling_price",
        sort_direction="desc",
    )

    assert result.records[-1].row_number == 8


@pytest.mark.parametrize(
    ("overrides", "code"),
    [
        ({"page": 0}, "INVALID_PAGINATION"),
        ({"page_size": 201}, "INVALID_PAGINATION"),
        ({"max_page_size": 0}, "INVALID_PAGINATION"),
        ({"header_row_number": 0}, "INVALID_ROW"),
        ({"header_row_number": 99}, "INVALID_ROW"),
        ({"sheet_name": "Không tồn tại"}, "SHEET_NOT_FOUND"),
        ({"sort_by": "unknown"}, "INVALID_SORT"),
        ({"sort_direction": "sideways"}, "INVALID_SORT"),
        (
            {
                "column_mapping": {
                    "net_price": 4,
                    "selling_price": 4,
                }
            },
            "AMBIGUOUS_MAPPING",
        ),
        (
            {
                "column_mapping": {
                    "net_price": 4,
                    "selling_price": 5,
                    "unsupported": 3,
                }
            },
            "INVALID_MAPPING",
        ),
        (
            {
                "column_mapping": {
                    "net_price": 4,
                    "selling_price": 50,
                }
            },
            "INVALID_MAPPING",
        ),
    ],
)
def test_rejects_bad_inputs(
    workbook_path: Path,
    overrides: dict,
    code: str,
) -> None:
    with pytest.raises(WorkbookReadError) as error:
        read(workbook_path, **overrides)

    assert error.value.code == code
    assert str(workbook_path) not in str(error.value)


def test_rejects_invalid_workbook_without_exposing_path(tmp_path: Path) -> None:
    invalid_path = tmp_path / "private-name.xlsx"
    invalid_path.write_bytes(b"not an xlsx")

    with pytest.raises(WorkbookReadError) as error:
        read(invalid_path)

    assert error.value.code == "INVALID_XLSX"
    assert str(invalid_path) not in str(error.value)

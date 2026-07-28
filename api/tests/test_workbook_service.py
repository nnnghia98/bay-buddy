"""Integration tests for Workbook Editor V2 application orchestration."""

from __future__ import annotations

import hashlib
import io
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook as OpenpyxlWorkbook, load_workbook
from sqlalchemy import event
from sqlmodel import Session, SQLModel, create_engine, select

from models import (
    User,
    UserRole,
    Workbook,
    WorkbookOperation,
    WorkbookSession,
    WorkbookSessionStatus,
    WorkbookVersion,
)
from services.workbook_mutation import PriceChange
from services.workbook_reader import WorkbookCellReference
import services.workbook_service as workbook_service
from services.workbook_service import (
    XLS_MIME_TYPE,
    XLSX_MIME_TYPE,
    WorkbookServiceError,
    add_session_column,
    create_editing_session,
    discard_editing_session,
    get_current_download,
    get_editing_session,
    get_latest_editing_session,
    list_editing_sessions,
    lookup_session_cell_values,
    preview_session_formula,
    read_session_records,
    rename_editing_session,
    remove_session_column,
    save_session_changes,
    update_session_column,
    update_session_column_configuration,
    upload_workbook,
)
from services.workbook_validation import MappingStatus
from storage.workbooks import LocalWorkbookStorage


@pytest.fixture()
def engine():
    database = create_engine("sqlite://", connect_args={"check_same_thread": False})

    @event.listens_for(database, "connect")
    def enable_foreign_keys(dbapi_connection, _connection_record) -> None:
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

    SQLModel.metadata.create_all(database)
    return database


@pytest.fixture()
def storage(tmp_path: Path) -> LocalWorkbookStorage:
    return LocalWorkbookStorage(tmp_path / "objects")


def make_user(db: Session, role: UserRole = UserRole.STAFF) -> User:
    user = User(
        username=f"user-{uuid.uuid4()}",
        hashed_password="test-only",
        role=role,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


def workbook_bytes(*, mapped: bool = True) -> bytes:
    stream = io.BytesIO()
    workbook = OpenpyxlWorkbook()
    worksheet = workbook.active
    worksheet.title = "Tickets"
    if mapped:
        worksheet.append(
            ["Passenger Name", "PNR", "Ticket Number", "Cost Price", "Selling Price"]
        )
        worksheet.append(["Nguyễn An", "ABC123", "738001", 1_000_000, 1_200_000])
        worksheet.append(["Trần Bình", "DEF456", "738002", 2_000_000, 2_300_000])
    else:
        worksheet.append(["Passenger Name", "PNR"])
        worksheet.append(["Nguyễn An", "ABC123"])
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def generic_workbook_bytes() -> bytes:
    stream = io.BytesIO()
    workbook = OpenpyxlWorkbook()
    worksheet = workbook.active
    worksheet.title = "Inventory"
    worksheet.append(["Quarterly inventory"])
    worksheet.append(["Item", "Quantity", "Active", "Date"])
    worksheet.append(["A-100", 12.5, True, date(2026, 1, 1)])
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def grouped_header_workbook_bytes() -> bytes:
    stream = io.BytesIO()
    workbook = OpenpyxlWorkbook()
    worksheet = workbook.active
    worksheet.title = "Metrics"
    worksheet.append(["Carrier", "Metrics", None])
    worksheet.append([None, "Quantity", "Active"])
    worksheet.append(["VN", 12.5, True])
    worksheet.merge_cells("B1:C1")
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def legacy_grouped_workbook_bytes(*, hidden_tail: bool = True) -> bytes:
    stream = io.BytesIO()
    workbook = OpenpyxlWorkbook()
    worksheet = workbook.active
    worksheet.title = "Legacy invoice"
    headers = [f"Column {column}" for column in range(1, 38)]
    headers[0] = "No."
    headers[15] = "ICT2"
    headers[16] = None
    headers[17] = "ICT3"
    headers[18] = None
    worksheet.append(headers)
    child_headers = [None] * 37
    child_headers[15:19] = ["%", "Comm", "%", "Comm"]
    worksheet.append(child_headers)
    worksheet.append([1, "01/06/2026", "Passenger", *range(4, 38)])
    worksheet.merge_cells("A1:A2")
    for column in range(2, 16):
        worksheet.merge_cells(start_row=1, start_column=column, end_row=2, end_column=column)
    worksheet.merge_cells("P1:Q1")
    worksheet.merge_cells("R1:S1")
    for column in range(20, 38):
        worksheet.merge_cells(start_row=1, start_column=column, end_row=2, end_column=column)
    if hidden_tail:
        for column in range(20, 33):
            worksheet.column_dimensions[worksheet.cell(1, column).column_letter].hidden = True
    worksheet["I3"].number_format = "#,##0.00"
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def workbook_with_blank_header_bytes() -> bytes:
    stream = io.BytesIO()
    workbook = OpenpyxlWorkbook()
    worksheet = workbook.active
    worksheet.title = "Tickets"
    worksheet.append(["Passenger Name", "PNR", None, "Cost Price", "Selling Price"])
    worksheet.append(["Nguyễn An", "ABC123", "note", 1_000_000, 1_200_000])
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def upload(
    db: Session,
    storage: LocalWorkbookStorage,
    actor: User,
    *,
    content: bytes | None = None,
):
    return upload_workbook(
        db,
        storage,
        actor=actor,
        filename="Bảng giá tháng 7.xlsx",
        mime_type=XLSX_MIME_TYPE,
        source=io.BytesIO(content if content is not None else workbook_bytes()),
        max_upload_bytes=5 * 1024 * 1024,
        max_rows=100,
        max_columns=20,
    )


def test_upload_persists_immutable_original_and_mapping(engine, storage) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        content = workbook_bytes()
        result = upload(db, storage, actor, content=content)

        persisted = db.get(Workbook, result.id)
        assert persisted is not None
        assert result.checksum == hashlib.sha256(content).hexdigest()
        assert persisted.original_relative_path.startswith("originals/")
        assert Path(persisted.original_relative_path).is_absolute() is False
        assert result.sheets[0].mapping_status is MappingStatus.READY
        assert result.sheets[0].column_mapping["net_price"] == 4
        with storage.open_read(key=persisted.original_relative_path) as stored:
            assert stored.read() == content


def test_upload_normalizes_legacy_xls_before_publishing(
    engine,
    storage,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    converted = workbook_bytes()

    def convert(
        source: Path,
        target: Path,
        *,
        max_rows: int,
        max_columns: int,
    ) -> None:
        assert source.suffix == ".xls"
        assert target.suffix == ".xlsx"
        assert (max_rows, max_columns) == (100, 20)
        target.write_bytes(converted)

    monkeypatch.setattr(workbook_service, "convert_xls_to_xlsx", convert)

    with Session(engine) as db:
        actor = make_user(db)
        result = upload_workbook(
            db,
            storage,
            actor=actor,
            filename="Bảng giá tháng 7.xls",
            mime_type=XLS_MIME_TYPE,
            source=io.BytesIO(b"legacy-xls-content"),
            max_upload_bytes=5 * 1024 * 1024,
            max_rows=100,
            max_columns=20,
        )

        persisted = db.get(Workbook, result.id)
        assert persisted is not None
        assert result.original_filename.endswith(".xls")
        assert result.mime_type == XLSX_MIME_TYPE
        assert result.checksum == hashlib.sha256(converted).hexdigest()
        with storage.open_read(key=persisted.original_relative_path) as stored:
            assert stored.read() == converted


def test_legacy_grouped_session_uses_main_header_band(
    engine,
    storage,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    converted = legacy_grouped_workbook_bytes()

    def convert(
        _source: Path,
        target: Path,
        *,
        max_rows: int,
        max_columns: int,
    ) -> None:
        assert (max_rows, max_columns) == (100, 50)
        target.write_bytes(converted)

    monkeypatch.setattr(workbook_service, "convert_xls_to_xlsx", convert)

    with Session(engine) as db:
        actor = make_user(db)
        uploaded = upload_workbook(
            db,
            storage,
            actor=actor,
            filename="legacy-invoice.xls",
            mime_type=XLS_MIME_TYPE,
            source=io.BytesIO(b"legacy"),
            max_upload_bytes=5 * 1024 * 1024,
            max_rows=100,
            max_columns=50,
        )
        session = create_editing_session(
            db,
            storage,
            actor=actor,
            workbook_id=uploaded.id,
            sheet_name="Legacy invoice",
        )
        records = read_session_records(
            db,
            storage,
            actor=actor,
            session_id=session.id,
        )

        visible = [column for column in records.page.columns if not column.hidden]
        top_level_headers = []
        for column in visible:
            label = column.group_label or column.header
            if not top_level_headers or top_level_headers[-1] != label:
                top_level_headers.append(label)

        assert len(visible) == 19
        assert len(top_level_headers) == 17
        assert visible[0].header == "No."
        assert visible[0].hidden is False
        assert [column.group_label for column in visible[15:19]] == [
            "ICT2",
            "ICT2",
            "ICT3",
            "ICT3",
        ]
        assert all(column.hidden for column in records.page.columns[19:])

        persisted = db.get(WorkbookSession, session.id)
        assert persisted is not None
        persisted.column_config = [
            {
                **column,
                "hidden": str(column.get("label", "")).strip().casefold()
                == "no.",
            }
            for column in persisted.column_config
        ]
        db.add(persisted)
        db.commit()

        reopened = get_editing_session(
            db,
            actor=actor,
            session_id=session.id,
            storage=storage,
        )

        assert reopened.column_config[0]["hidden"] is False
        assert all(
            column["hidden"] for column in reopened.column_config[19:]
        )


def test_legacy_session_created_before_visibility_preservation_is_repaired(
    engine,
    storage,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    converted = legacy_grouped_workbook_bytes(hidden_tail=False)

    def convert(
        _source: Path,
        target: Path,
        *,
        max_rows: int,
        max_columns: int,
    ) -> None:
        assert (max_rows, max_columns) == (100, 50)
        target.write_bytes(converted)

    monkeypatch.setattr(workbook_service, "convert_xls_to_xlsx", convert)

    with Session(engine) as db:
        actor = make_user(db)
        uploaded = upload_workbook(
            db,
            storage,
            actor=actor,
            filename="legacy-before-fix.xls",
            mime_type=XLS_MIME_TYPE,
            source=io.BytesIO(b"legacy"),
            max_upload_bytes=5 * 1024 * 1024,
            max_rows=100,
            max_columns=50,
        )
        session = create_editing_session(
            db,
            storage,
            actor=actor,
            workbook_id=uploaded.id,
            sheet_name="Legacy invoice",
        )
        persisted = db.get(WorkbookSession, session.id)
        assert persisted is not None
        persisted.column_config = [
            {
                **column,
                "hidden": str(column.get("label", "")).strip().casefold()
                == "no.",
                "sticky": column.get("column_number") == 2,
            }
            for column in persisted.column_config
        ]
        db.add(persisted)
        db.commit()

        reopened = get_editing_session(
            db,
            actor=actor,
            session_id=session.id,
            storage=storage,
        )

        assert reopened.column_config[0]["hidden"] is False
        assert all(
            not column["hidden"] for column in reopened.column_config[:19]
        )
        assert all(column["hidden"] for column in reopened.column_config[19:])
        assert reopened.column_config[1]["sticky"] is True


@pytest.mark.parametrize(
    ("filename", "mime_type", "limit", "code", "status"),
    [
        ("prices.csv", XLSX_MIME_TYPE, 10_000, "UNSUPPORTED_FILE_TYPE", 415),
        ("prices.xlsx", "text/plain", 10_000, "UNSUPPORTED_FILE_TYPE", 415),
        ("prices.xlsx", XLSX_MIME_TYPE, 10, "FILE_TOO_LARGE", 413),
    ],
)
def test_upload_rejects_type_and_size(
    engine,
    storage,
    filename: str,
    mime_type: str,
    limit: int,
    code: str,
    status: int,
) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        with pytest.raises(WorkbookServiceError) as captured:
            upload_workbook(
                db,
                storage,
                actor=actor,
                filename=filename,
                mime_type=mime_type,
                source=io.BytesIO(workbook_bytes()),
                max_upload_bytes=limit,
                max_rows=100,
                max_columns=20,
            )
        assert captured.value.code == code
        assert captured.value.status_code == status
        assert db.exec(select(Workbook)).all() == []


def test_incomplete_mapping_warns_but_can_create_session(engine, storage) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        result = upload(db, storage, actor, content=workbook_bytes(mapped=False))
        assert result.sheets[0].mapping_status is MappingStatus.MAPPING_INCOMPLETE

        session = create_editing_session(
            db,
            storage,
            actor=actor,
            workbook_id=result.id,
            sheet_name="Tickets",
        )
        assert session.selected_sheet_name == "Tickets"
        assert session.column_mapping == {"passenger_name": 1, "pnr": 2}

        before = read_session_records(
            db,
            storage,
            actor=actor,
            session_id=session.id,
        )
        assert [column.header for column in before.page.columns] == [
            "Passenger Name",
            "PNR",
        ]


def test_automatic_header_detection_creates_generic_session(engine, storage) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        result = upload(db, storage, actor, content=generic_workbook_bytes())

        session = create_editing_session(
            db,
            storage,
            actor=actor,
            workbook_id=result.id,
            sheet_name="Inventory",
        )

        assert session.header_row_number == 2
        assert session.column_mapping == {}
        assert [column["label"] for column in session.column_config] == [
            "Item",
            "Quantity",
            "Active",
            "Date",
        ]
        assert [column["data_type"] for column in session.column_config] == [
            "text",
            "number",
            "boolean",
            "date",
        ]
        records = read_session_records(
            db,
            storage,
            actor=actor,
            session_id=session.id,
        )
        assert [record.row_number for record in records.page.records] == [3]

        config_by_label = {
            column["label"]: column["id"] for column in session.column_config
        }
        saved = save_session_changes(
            db,
            storage,
            actor=actor,
            session_id=session.id,
            request_id=uuid.uuid4(),
            base_version=1,
            changes=[
                PriceChange(
                    row_number=3,
                    values={
                        config_by_label["Item"]: "A-101",
                        config_by_label["Quantity"]: -2.75,
                        config_by_label["Active"]: False,
                        config_by_label["Date"]: "2026-07-14",
                    },
                )
            ],
        )
        assert saved.changed_cells == 4
        operation = db.exec(
            select(WorkbookOperation).where(
                WorkbookOperation.session_id == session.id
            )
        ).one()
        assert operation.operation_payload["changes"][3]["new_value"] == "2026-07-14"


def test_type_inference_starts_after_grouped_header_band(engine, storage) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        result = upload(db, storage, actor, content=grouped_header_workbook_bytes())

        session = create_editing_session(
            db,
            storage,
            actor=actor,
            workbook_id=result.id,
            sheet_name="Metrics",
        )

        assert [column["data_type"] for column in session.column_config] == [
            "text",
            "number",
            "boolean",
        ]


def test_staff_ownership_is_hidden_while_admin_can_access(engine, storage) -> None:
    with Session(engine) as db:
        owner = make_user(db)
        other = make_user(db)
        admin = make_user(db, UserRole.ADMIN)
        uploaded = upload(db, storage, owner)
        editing = create_editing_session(
            db,
            storage,
            actor=owner,
            workbook_id=uploaded.id,
            sheet_name="Tickets",
        )

        with pytest.raises(WorkbookServiceError) as captured:
            get_editing_session(db, actor=other, session_id=editing.id)
        assert captured.value.code == "SESSION_NOT_FOUND"
        assert get_editing_session(
            db, actor=admin, session_id=editing.id
        ).id == editing.id

        with pytest.raises(WorkbookServiceError) as workbook_error:
            create_editing_session(
                db,
                storage,
                actor=other,
                workbook_id=uploaded.id,
                sheet_name="Tickets",
            )
        assert workbook_error.value.code == "WORKBOOK_NOT_FOUND"


def test_session_library_is_owner_scoped_paginated_and_searchable(
    engine, storage
) -> None:
    with Session(engine) as db:
        owner = make_user(db)
        other = make_user(db)
        admin = make_user(db, UserRole.ADMIN)
        uploaded = upload(db, storage, owner)
        first = create_editing_session(
            db, storage, actor=owner, workbook_id=uploaded.id, sheet_name="Tickets"
        )
        second = create_editing_session(
            db, storage, actor=owner, workbook_id=uploaded.id, sheet_name="Tickets"
        )
        other_upload = upload(db, storage, other)
        create_editing_session(
            db,
            storage,
            actor=other,
            workbook_id=other_upload.id,
            sheet_name="Tickets",
        )
        admin_upload = upload(db, storage, admin)
        admin_session = create_editing_session(
            db,
            storage,
            actor=admin,
            workbook_id=admin_upload.id,
            sheet_name="Tickets",
        )

        rename_editing_session(
            db,
            actor=owner,
            session_id=first.id,
            display_name="  July supplier prices  ",
        )
        completed = db.get(WorkbookSession, second.id)
        assert completed is not None
        completed.status = WorkbookSessionStatus.COMPLETED
        db.add(completed)
        db.commit()

        page = list_editing_sessions(db, actor=owner, page=1, page_size=1)
        assert page.total == 2
        assert page.total_pages == 2
        assert len(page.items) == 1

        search = list_editing_sessions(
            db,
            actor=owner,
            search="supplier",
        )
        assert [item.id for item in search.items] == [first.id]
        assert search.items[0].display_name == "July supplier prices"

        completed_only = list_editing_sessions(
            db,
            actor=owner,
            session_status=WorkbookSessionStatus.COMPLETED,
        )
        assert [item.id for item in completed_only.items] == [second.id]

        assert list_editing_sessions(db, actor=other).total == 1
        admin_list = list_editing_sessions(db, actor=admin)
        assert admin_list.total == 4
        assert {item.id for item in admin_list.items} == {
            first.id,
            second.id,
            admin_session.id,
            db.exec(
                select(WorkbookSession).where(
                    WorkbookSession.created_by == other.id
                )
            ).one().id,
        }


def test_rename_requires_owner_and_active_session(engine, storage) -> None:
    with Session(engine) as db:
        owner = make_user(db)
        other = make_user(db)
        uploaded = upload(db, storage, owner)
        editing = create_editing_session(
            db, storage, actor=owner, workbook_id=uploaded.id, sheet_name="Tickets"
        )

        with pytest.raises(WorkbookServiceError) as hidden:
            rename_editing_session(
                db,
                actor=other,
                session_id=editing.id,
                display_name="Not allowed",
            )
        assert hidden.value.code == "SESSION_NOT_FOUND"

        renamed = rename_editing_session(
            db,
            actor=owner,
            session_id=editing.id,
            display_name="  Agency July workbook  ",
        )
        assert renamed.display_name == "Agency July workbook"

        persisted = db.get(WorkbookSession, editing.id)
        assert persisted is not None
        persisted.status = WorkbookSessionStatus.COMPLETED
        db.add(persisted)
        db.commit()
        with pytest.raises(WorkbookServiceError) as inactive:
            rename_editing_session(
                db,
                actor=owner,
                session_id=editing.id,
                display_name="Too late",
            )
        assert inactive.value.code == "SESSION_NOT_ACTIVE"


def test_soft_discard_marks_timestamp_and_hides_session_by_default(
    engine, storage
) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        uploaded = upload(db, storage, actor)
        editing = create_editing_session(
            db, storage, actor=actor, workbook_id=uploaded.id, sheet_name="Tickets"
        )

        discarded = discard_editing_session(
            db,
            actor=actor,
            session_id=editing.id,
        )

        assert discarded.status == WorkbookSessionStatus.DISCARDED
        persisted = db.get(WorkbookSession, editing.id)
        assert persisted is not None
        assert persisted.discarded_at is not None
        assert persisted.updated_at == persisted.discarded_at
        assert list_editing_sessions(db, actor=actor).total == 0
        discarded_list = list_editing_sessions(
            db,
            actor=actor,
            session_status=WorkbookSessionStatus.DISCARDED,
        )
        assert [item.id for item in discarded_list.items] == [editing.id]

        with pytest.raises(WorkbookServiceError) as inactive:
            discard_editing_session(db, actor=actor, session_id=editing.id)
        assert inactive.value.code == "SESSION_NOT_ACTIVE"


def test_latest_session_restores_most_recent_active_session(engine, storage) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        uploaded = upload(db, storage, actor)
        first = create_editing_session(
            db, storage, actor=actor, workbook_id=uploaded.id, sheet_name="Tickets"
        )
        second = create_editing_session(
            db, storage, actor=actor, workbook_id=uploaded.id, sheet_name="Tickets"
        )

        restored = get_latest_editing_session(db, actor=actor)

        assert restored.id == second.id
        assert restored.id != first.id


def test_latest_session_is_scoped_to_current_user(engine) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        with pytest.raises(WorkbookServiceError) as captured:
            get_latest_editing_session(db, actor=actor)
        assert captured.value.code == "SESSION_NOT_FOUND"


def test_blank_source_headers_are_preserved_when_restored(engine, storage) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        uploaded = upload(
            db, storage, actor, content=workbook_with_blank_header_bytes()
        )

        editing = create_editing_session(
            db, storage, actor=actor, workbook_id=uploaded.id, sheet_name="Tickets"
        )
        restored = get_latest_editing_session(db, actor=actor)

        assert editing.column_config[2]["label"] == ""
        assert restored.column_config[2]["label"] == ""


def test_sessions_are_independent_and_records_read_server_side(engine, storage) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        uploaded = upload(db, storage, actor)
        first = create_editing_session(
            db,
            storage,
            actor=actor,
            workbook_id=uploaded.id,
            sheet_name="Tickets",
        )
        second = create_editing_session(
            db,
            storage,
            actor=actor,
            workbook_id=uploaded.id,
            sheet_name="Tickets",
        )
        versions = db.exec(select(WorkbookVersion)).all()
        assert len(versions) == 2
        assert versions[0].relative_path != versions[1].relative_path
        assert versions[0].checksum == uploaded.checksum == versions[1].checksum

        records = read_session_records(
            db,
            storage,
            actor=actor,
            session_id=first.id,
            search="nguyen an",
            page_size=1,
        )
        assert records.page.pagination.total == 1
        assert records.page.records[0].row_number == 2

        save_session_changes(
            db,
            storage,
            actor=actor,
            session_id=first.id,
            request_id=uuid.uuid4(),
            base_version=1,
            changes=[PriceChange(row_number=2, selling_price=1_300_000)],
        )
        untouched = read_session_records(
            db, storage, actor=actor, session_id=second.id
        )
        selling_column = next(
            column
            for column in untouched.page.columns
            if column.semantic_field == "selling_price"
        )
        assert untouched.page.records[0].values[selling_column.field] == 1_200_000


def test_cell_value_lookup_enforces_owner_version_and_does_not_mutate(
    engine, storage
) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        other = make_user(db)
        admin = make_user(db, UserRole.ADMIN)
        uploaded = upload(db, storage, actor)
        editing = create_editing_session(
            db,
            storage,
            actor=actor,
            workbook_id=uploaded.id,
            sheet_name="Tickets",
        )
        selling_id = next(
            column["id"]
            for column in editing.column_config
            if column.get("semantic_field") == "selling_price"
        )
        versions_before = len(db.exec(select(WorkbookVersion)).all())
        operations_before = len(db.exec(select(WorkbookOperation)).all())

        owner_result = lookup_session_cell_values(
            db,
            storage,
            actor=actor,
            session_id=editing.id,
            base_version=1,
            cells=[WorkbookCellReference(2, selling_id)],
        )
        assert owner_result.cells[0].value == 1_200_000
        admin_result = lookup_session_cell_values(
            db,
            storage,
            actor=admin,
            session_id=editing.id,
            base_version=1,
            cells=[WorkbookCellReference(2, selling_id)],
        )
        assert admin_result.cells[0].value == 1_200_000
        with pytest.raises(WorkbookServiceError) as hidden:
            lookup_session_cell_values(
                db,
                storage,
                actor=other,
                session_id=editing.id,
                base_version=1,
                cells=[WorkbookCellReference(2, selling_id)],
            )
        assert hidden.value.code == "SESSION_NOT_FOUND"
        assert len(db.exec(select(WorkbookVersion)).all()) == versions_before
        assert len(db.exec(select(WorkbookOperation)).all()) == operations_before

        save_session_changes(
            db,
            storage,
            actor=actor,
            session_id=editing.id,
            request_id=uuid.uuid4(),
            base_version=1,
            changes=[PriceChange(row_number=2, selling_price=1_300_000)],
        )
        with pytest.raises(WorkbookServiceError) as stale:
            lookup_session_cell_values(
                db,
                storage,
                actor=actor,
                session_id=editing.id,
                base_version=1,
                cells=[WorkbookCellReference(2, selling_id)],
            )
        assert stale.value.code == "VERSION_CONFLICT"
        assert stale.value.details == {"current_version": 2}
        current = lookup_session_cell_values(
            db,
            storage,
            actor=actor,
            session_id=editing.id,
            base_version=2,
            cells=[WorkbookCellReference(2, selling_id)],
        )
        assert current.cells[0].value == 1_300_000


def test_save_versions_audits_replays_conflicts_and_downloads(engine, storage) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        original = workbook_bytes()
        uploaded = upload(db, storage, actor, content=original)
        editing = create_editing_session(
            db,
            storage,
            actor=actor,
            workbook_id=uploaded.id,
            sheet_name="Tickets",
        )
        request_id = uuid.uuid4()
        changes = [PriceChange(row_number=2, net_price=1_050_000)]
        saved = save_session_changes(
            db,
            storage,
            actor=actor,
            session_id=editing.id,
            request_id=request_id,
            base_version=1,
            changes=changes,
        )
        assert saved.current_version == 2
        assert saved.changed_cells == 1
        assert saved.replayed is False
        assert saved.checksum != uploaded.checksum

        session_row = db.get(WorkbookSession, editing.id)
        assert session_row is not None and session_row.current_version == 2
        versions = db.exec(
            select(WorkbookVersion).where(
                WorkbookVersion.session_id == editing.id
            )
        ).all()
        operations = db.exec(
            select(WorkbookOperation).where(
                WorkbookOperation.session_id == editing.id
            )
        ).all()
        assert [version.version_number for version in versions] == [1, 2]
        assert len(operations) == 1
        assert operations[0].operation_payload["changes"][0]["old_value"] == 1_000_000
        assert operations[0].operation_payload["changes"][0]["new_value"] == 1_050_000

        replay = save_session_changes(
            db,
            storage,
            actor=actor,
            session_id=editing.id,
            request_id=request_id,
            base_version=1,
            changes=changes,
        )
        assert replay.replayed is True
        assert replay.current_version == 2
        assert len(db.exec(select(WorkbookOperation)).all()) == 1

        with pytest.raises(WorkbookServiceError) as reused:
            save_session_changes(
                db,
                storage,
                actor=actor,
                session_id=editing.id,
                request_id=request_id,
                base_version=1,
                changes=[PriceChange(row_number=2, net_price=1_060_000)],
            )
        assert reused.value.code == "IDEMPOTENCY_KEY_REUSED"
        assert reused.value.status_code == 409

        with pytest.raises(WorkbookServiceError) as stale:
            save_session_changes(
                db,
                storage,
                actor=actor,
                session_id=editing.id,
                request_id=uuid.uuid4(),
                base_version=1,
                changes=[PriceChange(row_number=2, net_price=1_060_000)],
            )
        assert stale.value.code == "VERSION_CONFLICT"
        assert stale.value.details == {"current_version": 2}

        descriptor = get_current_download(
            db, storage, actor=actor, session_id=editing.id
        )
        try:
            downloaded = descriptor.stream.read()
        finally:
            descriptor.stream.close()
        assert descriptor.filename.endswith("-edited-v2.xlsx")
        assert descriptor.version == 2
        assert descriptor.file_size == len(downloaded)
        assert descriptor.checksum == hashlib.sha256(downloaded).hexdigest()
        exported_workbook = load_workbook(io.BytesIO(downloaded))
        exported_sheet = exported_workbook["Tickets"]
        assert exported_sheet["A1"].fill.fgColor.rgb == "001B61C9"
        assert exported_sheet.freeze_panes == "A2"
        assert exported_sheet.sheet_view.showGridLines is False
        exported_workbook.close()
        assert hashlib.sha256(original).hexdigest() == uploaded.checksum
        workbook_row = db.get(Workbook, uploaded.id)
        assert workbook_row is not None
        with storage.open_read(key=workbook_row.original_relative_path) as source:
            assert source.read() == original


def test_failed_save_rolls_back_version_operation_and_session(engine, storage) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        uploaded = upload(db, storage, actor)
        editing = create_editing_session(
            db,
            storage,
            actor=actor,
            workbook_id=uploaded.id,
            sheet_name="Tickets",
        )

        with pytest.raises(WorkbookServiceError) as captured:
            save_session_changes(
                db,
                storage,
                actor=actor,
                session_id=editing.id,
                request_id=uuid.uuid4(),
                base_version=1,
                changes=[PriceChange(row_number=2, values={"net_price": "invalid"})],
            )

        assert captured.value.code == "INVALID_CELL_VALUE"
        session_row = db.get(WorkbookSession, editing.id)
        assert session_row is not None and session_row.current_version == 1
        assert len(db.exec(select(WorkbookVersion)).all()) == 1
        assert db.exec(select(WorkbookOperation)).all() == []


def test_download_rejects_corrupted_stored_version(engine, storage) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        uploaded = upload(db, storage, actor)
        editing = create_editing_session(
            db,
            storage,
            actor=actor,
            workbook_id=uploaded.id,
            sheet_name="Tickets",
        )
        version = db.exec(
            select(WorkbookVersion).where(
                WorkbookVersion.session_id == editing.id,
                WorkbookVersion.version_number == 1,
            )
        ).one()
        (storage._root / version.relative_path).write_bytes(b"corrupt")

        with pytest.raises(WorkbookServiceError) as captured:
            get_current_download(
                db,
                storage,
                actor=actor,
                session_id=editing.id,
            )

        assert captured.value.code == "STORAGE_OBJECT_MISSING"


def test_sqlite_concurrent_replay_creates_one_version(
    tmp_path: Path,
) -> None:
    database = create_engine(
        f"sqlite:///{tmp_path / 'concurrent.db'}",
        connect_args={"check_same_thread": False},
    )
    SQLModel.metadata.create_all(database)
    local_storage = LocalWorkbookStorage(tmp_path / "concurrent-storage")
    with Session(database) as setup_db:
        actor = make_user(setup_db)
        uploaded = upload(setup_db, local_storage, actor)
        editing = create_editing_session(
            setup_db,
            local_storage,
            actor=actor,
            workbook_id=uploaded.id,
            sheet_name="Tickets",
        )
        actor_id = actor.id

    request_id = uuid.uuid4()

    def save_once() -> bool:
        with Session(database) as worker_db:
            worker = worker_db.get(User, actor_id)
            assert worker is not None
            result = save_session_changes(
                worker_db,
                local_storage,
                actor=worker,
                session_id=editing.id,
                request_id=request_id,
                base_version=1,
                changes=[PriceChange(row_number=2, net_price=1_050_000)],
            )
            return result.replayed

    with ThreadPoolExecutor(max_workers=2) as executor:
        replayed = list(executor.map(lambda _index: save_once(), range(2)))

    assert sorted(replayed) == [False, True]
    with Session(database) as check_db:
        assert len(check_db.exec(select(WorkbookOperation)).all()) == 1
        assert len(check_db.exec(select(WorkbookVersion)).all()) == 2


def test_formula_preview_add_update_and_dependency_removal(engine, storage) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        uploaded = upload(db, storage, actor)
        editing = create_editing_session(
            db,
            storage,
            actor=actor,
            workbook_id=uploaded.id,
            sheet_name="Tickets",
        )
        by_semantic = {
            column.get("semantic_field"): column["id"]
            for column in editing.column_config
            if column.get("semantic_field")
        }
        profit_formula = {
            "schema_version": 1,
            "expression": {
                "type": "binary",
                "operator": "-",
                "left": {"type": "column", "column_id": by_semantic["selling_price"]},
                "right": {"type": "column", "column_id": by_semantic["net_price"]},
            },
        }

        preview = preview_session_formula(
            db,
            storage,
            actor=actor,
            session_id=editing.id,
            base_version=1,
            formula=profit_formula,
            output_type="currency",
        )
        assert preview.valid is True
        assert [row.value for row in preview.results] == [200_000, 300_000]
        assert preview.readable_expression == "(Selling Price - Cost Price)"

        added = add_session_column(
            db,
            storage,
            actor=actor,
            session_id=editing.id,
            base_version=1,
            label="Profit",
            data_type="currency",
            formula=profit_formula,
        )
        formula_column = next(column for column in added.column_config if column["label"] == "Profit")
        assert added.current_version == 2
        assert formula_column["formula"]["schema_version"] == 1

        rounded_formula = {
            "schema_version": 1,
            "expression": {
                "type": "round",
                "value": {"type": "column", "column_id": formula_column["id"]},
                "digits": 0,
            },
        }
        rounded = add_session_column(
            db,
            storage,
            actor=actor,
            session_id=editing.id,
            base_version=2,
            label="Rounded profit",
            data_type="currency",
            formula=rounded_formula,
        )
        rounded_column = next(column for column in rounded.column_config if column["label"] == "Rounded profit")
        assert rounded.current_version == 3

        with pytest.raises(WorkbookServiceError) as in_use:
            remove_session_column(
                db,
                storage,
                actor=actor,
                session_id=editing.id,
                column_id=formula_column["id"],
                base_version=3,
            )
        assert in_use.value.code == "COLUMN_IN_USE"
        assert in_use.value.details["dependent_column_ids"] == [rounded_column["id"]]

        updated = update_session_column(
            db,
            storage,
            actor=actor,
            session_id=editing.id,
            column_id=rounded_column["id"],
            base_version=3,
            label="Rounded margin",
            formula=None,
            formula_was_provided=True,
        )
        assert updated.current_version == 4
        updated_column = next(column for column in updated.column_config if column["id"] == rounded_column["id"])
        assert updated_column["label"] == "Rounded margin"
        assert updated_column["formula"] is None
        assert len(db.exec(select(WorkbookVersion)).all()) == 4


def test_column_configuration_is_version_checked_and_versionless(engine, storage) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        uploaded = upload(db, storage, actor)
        editing = create_editing_session(
            db,
            storage,
            actor=actor,
            workbook_id=uploaded.id,
            sheet_name="Tickets",
        )
        first_id = editing.column_config[0]["id"]

        configured = update_session_column_configuration(
            db,
            actor=actor,
            session_id=editing.id,
            base_version=1,
            hidden_column_ids=[first_id],
            sticky_column_ids=[],
        )
        assert configured.current_version == 1
        assert configured.column_config[0]["hidden"] is True
        assert len(db.exec(select(WorkbookVersion)).all()) == 1

        added = add_session_column(
            db,
            storage,
            actor=actor,
            session_id=editing.id,
            base_version=1,
            label="Note",
            data_type="text",
        )
        with pytest.raises(WorkbookServiceError) as stale:
            update_session_column_configuration(
                db,
                actor=actor,
                session_id=editing.id,
                base_version=1,
                hidden_column_ids=[],
                sticky_column_ids=[],
            )
        assert stale.value.code == "VERSION_CONFLICT"
        persisted = db.get(WorkbookSession, editing.id)
        assert persisted is not None
        assert any(item["id"] == added.column_config[-1]["id"] for item in persisted.column_config)


def test_column_type_change_rejects_incompatible_populated_values(engine, storage) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        uploaded = upload(db, storage, actor)
        editing = create_editing_session(
            db,
            storage,
            actor=actor,
            workbook_id=uploaded.id,
            sheet_name="Tickets",
        )
        added = add_session_column(
            db,
            storage,
            actor=actor,
            session_id=editing.id,
            base_version=1,
            label="Note",
            data_type="text",
        )
        note_id = added.column_config[-1]["id"]
        saved = save_session_changes(
            db,
            storage,
            actor=actor,
            session_id=editing.id,
            request_id=uuid.uuid4(),
            base_version=2,
            changes=[PriceChange(row_number=2, values={note_id: "not a number"})],
        )

        with pytest.raises(WorkbookServiceError) as incompatible:
            update_session_column(
                db,
                storage,
                actor=actor,
                session_id=editing.id,
                column_id=note_id,
                base_version=saved.current_version,
                data_type="number",
            )
        assert incompatible.value.code == "INVALID_COLUMN_TYPE"


def test_formula_preview_and_update_enforce_base_version(engine, storage) -> None:
    with Session(engine) as db:
        actor = make_user(db)
        uploaded = upload(db, storage, actor)
        editing = create_editing_session(
            db,
            storage,
            actor=actor,
            workbook_id=uploaded.id,
            sheet_name="Tickets",
        )
        formula = {
            "schema_version": 1,
            "expression": {"type": "constant", "value": "1"},
        }
        with pytest.raises(WorkbookServiceError) as preview_conflict:
            preview_session_formula(
                db,
                storage,
                actor=actor,
                session_id=editing.id,
                base_version=2,
                formula=formula,
                output_type="number",
            )
        assert preview_conflict.value.code == "VERSION_CONFLICT"

        user_column = add_session_column(
            db,
            storage,
            actor=actor,
            session_id=editing.id,
            base_version=1,
            label="Manual",
            data_type="number",
        ).column_config[-1]
        with pytest.raises(WorkbookServiceError) as update_conflict:
            update_session_column(
                db,
                storage,
                actor=actor,
                session_id=editing.id,
                column_id=user_column["id"],
                base_version=1,
                label="Stale rename",
            )
        assert update_conflict.value.code == "VERSION_CONFLICT"

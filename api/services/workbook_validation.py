"""Safe validation and business-header inspection for Workbook Editor V2."""

from __future__ import annotations

import re
import unicodedata
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from enum import Enum
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, range_boundaries


_HEADER_SCAN_NON_EMPTY_ROWS = 25
_MAX_ARCHIVE_MEMBERS = 10_000
_MAX_ARCHIVE_UNCOMPRESSED_BYTES = 200 * 1024 * 1024
_MAX_MEMBER_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
_MAX_COMPRESSION_RATIO = 250

_REQUIRED_FIELDS = ("net_price", "selling_price")
_ALIASES = {
    "net_price": (
        "Giá gốc",
        "Giá hệ thống",
        "Giá net",
        "Net Price",
        "Cost Price",
        "Cost",
    ),
    "selling_price": (
        "Giá bán",
        "Giá thu",
        "Selling Price",
        "Sale Price",
        "Customer Price",
    ),
    "passenger_name": (
        "Nội dung",
        "Hành khách",
        "Tên hành khách",
        "Họ tên",
        "Passenger Name",
    ),
    "pnr": (
        "Mã chỗ",
        "Mã đặt chỗ",
        "PNR",
        "Booking Code",
        "Booking Reference",
    ),
    "ticket_number": (
        "Số vé",
        "Ticket Number",
        "Ticket No",
    ),
}


class MappingStatus(str, Enum):
    """Business-column mapping result for one visible worksheet."""

    READY = "READY"
    MAPPING_INCOMPLETE = "MAPPING_INCOMPLETE"
    AMBIGUOUS_MAPPING = "AMBIGUOUS_MAPPING"


class WorkbookValidationError(ValueError):
    """A safe, machine-readable workbook validation failure."""

    def __init__(
        self,
        code: str,
        message: str,
        *,
        details: dict[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = details or {}


@dataclass(frozen=True, slots=True)
class _DetectedHeader:
    """The worksheet header selected by bounded automatic inspection."""

    row_number: int
    detected_headers: tuple[str, ...]
    column_mapping: dict[str, int]
    mapping_status: MappingStatus
    missing_required_fields: tuple[str, ...]
    ambiguous_fields: dict[str, tuple[int, ...]]


@dataclass(frozen=True, slots=True)
class WorksheetInspection:
    """Safe worksheet metadata and best-effort business-column mapping."""

    name: str
    max_row: int
    max_column: int
    header_row_number: int | None
    detected_headers: tuple[str, ...]
    column_mapping: dict[str, int]
    mapping_status: MappingStatus
    missing_required_fields: tuple[str, ...]
    ambiguous_fields: dict[str, tuple[int, ...]]


@dataclass(frozen=True, slots=True)
class WorkbookInspection:
    """Inspection result for all visible worksheets in an XLSX workbook."""

    sheet_count: int
    sheets: tuple[WorksheetInspection, ...]


@dataclass(frozen=True, slots=True)
class _HeaderCandidate:
    row_number: int
    headers: tuple[str, ...]
    mapping: dict[str, int]
    ambiguous: dict[str, tuple[int, ...]]
    text_cell_count: int
    non_empty_cell_count: int

    @property
    def score(self) -> int:
        return len(self.mapping) + len(self.ambiguous)

    @property
    def ready(self) -> bool:
        return not self.ambiguous and all(
            field in self.mapping for field in _REQUIRED_FIELDS
        )


def normalize_header(value: object) -> str:
    """Normalize a header for accent/case/punctuation-insensitive matching."""

    if value is None:
        return ""
    decomposed = unicodedata.normalize("NFKD", str(value))
    without_accents = "".join(
        character
        for character in decomposed
        if not unicodedata.combining(character)
    )
    punctuation_as_spaces = re.sub(r"[^a-z0-9]+", " ", without_accents.casefold())
    return " ".join(punctuation_as_spaces.split())


_NORMALIZED_ALIASES = {
    field: frozenset(normalize_header(alias) for alias in aliases)
    for field, aliases in _ALIASES.items()
}


def validate_and_inspect_workbook(
    path: str | Path,
    *,
    max_rows: int,
    max_columns: int,
) -> WorkbookInspection:
    """Validate an XLSX and inspect visible sheets for supported headers."""

    if max_rows <= 0 or max_columns <= 0:
        raise ValueError("Workbook row and column limits must be positive.")

    workbook_path = Path(path)
    _require_xlsx_extension(workbook_path)
    _validate_archive(workbook_path)

    workbook = _open_workbook(workbook_path)
    try:
        visible_sheets = tuple(
            worksheet
            for worksheet in workbook.worksheets
            if worksheet.sheet_state == "visible"
        )
        if not visible_sheets:
            raise WorkbookValidationError(
                "INVALID_XLSX",
                "Workbook must contain at least one visible worksheet.",
                details={"reason": "NO_VISIBLE_WORKSHEETS"},
            )

        inspections: list[WorksheetInspection] = []
        workbook_has_content = False
        for worksheet in visible_sheets:
            row_count, column_count = _meaningful_worksheet_bounds(
                workbook_path, worksheet
            )
            if row_count > max_rows or column_count > max_columns:
                raise WorkbookValidationError(
                    "WORKBOOK_LIMIT_EXCEEDED",
                    "Workbook exceeds the supported worksheet size.",
                    details={
                        "sheet_name": worksheet.title,
                        "rows": row_count,
                        "columns": column_count,
                        "max_rows": max_rows,
                        "max_columns": max_columns,
                    },
                )

            non_empty_rows = _first_non_empty_rows(
                worksheet,
                max_row=row_count,
                max_column=column_count,
            )
            if non_empty_rows:
                workbook_has_content = True
            inspections.append(
                _inspect_sheet(
                    name=worksheet.title,
                    max_row=row_count,
                    max_column=column_count,
                    non_empty_rows=non_empty_rows,
                )
            )

        if not workbook_has_content:
            raise WorkbookValidationError(
                "INVALID_XLSX",
                "Workbook does not contain any data.",
                details={"reason": "EMPTY_WORKBOOK"},
            )

        return WorkbookInspection(
            sheet_count=len(inspections),
            sheets=tuple(inspections),
        )
    finally:
        workbook.close()


def validate_generated_workbook(path: str | Path) -> None:
    """Reopen a generated XLSX to ensure it is safe and structurally readable."""

    workbook_path = Path(path)
    _require_xlsx_extension(workbook_path)
    _validate_archive(workbook_path)
    workbook = _open_workbook(workbook_path)
    try:
        if not workbook.worksheets:
            raise WorkbookValidationError(
                "INVALID_XLSX",
                "Generated workbook does not contain a worksheet.",
                details={"reason": "NO_WORKSHEETS"},
            )
    finally:
        workbook.close()


def _require_xlsx_extension(path: Path) -> None:
    if path.suffix.casefold() != ".xlsx":
        raise WorkbookValidationError(
            "UNSUPPORTED_FILE_TYPE",
            "Only .xlsx workbooks are supported.",
        )


def _validate_archive(path: Path) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
            if not members:
                raise WorkbookValidationError(
                    "INVALID_XLSX",
                    "Workbook package is empty.",
                    details={"reason": "EMPTY_PACKAGE"},
                )
            if len(members) > _MAX_ARCHIVE_MEMBERS:
                _raise_unsafe_archive("TOO_MANY_MEMBERS")

            total_uncompressed = 0
            macro_content = False
            for member in members:
                _validate_member_name(member.filename)
                if member.flag_bits & 0x1:
                    raise WorkbookValidationError(
                        "INVALID_XLSX",
                        "Encrypted workbooks are not supported.",
                        details={"reason": "ENCRYPTED_WORKBOOK"},
                    )

                total_uncompressed += member.file_size
                if (
                    member.file_size > _MAX_MEMBER_UNCOMPRESSED_BYTES
                    or total_uncompressed > _MAX_ARCHIVE_UNCOMPRESSED_BYTES
                ):
                    _raise_unsafe_archive("UNCOMPRESSED_SIZE_LIMIT")
                if member.file_size and (
                    member.compress_size == 0
                    or member.file_size / member.compress_size
                    > _MAX_COMPRESSION_RATIO
                ):
                    _raise_unsafe_archive("SUSPICIOUS_COMPRESSION")

                normalized_name = member.filename.casefold()
                if normalized_name.endswith("vbaproject.bin"):
                    macro_content = True

            if macro_content or _content_types_include_macros(archive):
                raise WorkbookValidationError(
                    "UNSUPPORTED_FILE_TYPE",
                    "Macro-enabled workbooks are not supported.",
                    details={"reason": "MACRO_CONTENT"},
                )
    except WorkbookValidationError:
        raise
    except (FileNotFoundError, PermissionError, OSError, zipfile.BadZipFile) as exc:
        raise WorkbookValidationError(
            "INVALID_XLSX",
            "Workbook is not a readable XLSX file.",
            details={"reason": "UNREADABLE_PACKAGE"},
        ) from exc


def _validate_member_name(name: str) -> None:
    if not name or "\x00" in name or "\\" in name:
        _raise_unsafe_archive("UNSAFE_MEMBER_PATH")
    member_path = PurePosixPath(name)
    if (
        member_path.is_absolute()
        or any(part in {"", ".", ".."} for part in member_path.parts)
        or (member_path.parts and ":" in member_path.parts[0])
    ):
        _raise_unsafe_archive("UNSAFE_MEMBER_PATH")


def _raise_unsafe_archive(reason: str) -> None:
    raise WorkbookValidationError(
        "UNSAFE_XLSX_ARCHIVE",
        "Workbook package failed safety validation.",
        details={"reason": reason},
    )


def _content_types_include_macros(archive: zipfile.ZipFile) -> bool:
    try:
        content_types = archive.read("[Content_Types].xml")
    except KeyError:
        return False
    lowered = content_types.lower()
    return b"macroenabled" in lowered or b"vnd.ms-office.vba" in lowered


def _open_workbook(path: Path):
    try:
        return load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        # openpyxl can surface several XML parser implementations and relationship
        # errors for malformed packages. This untrusted-file boundary deliberately
        # converts all parser failures into one safe domain error.
        raise WorkbookValidationError(
            "INVALID_XLSX",
            "Workbook content is corrupt or unsupported.",
            details={"reason": "INVALID_WORKBOOK_CONTENT"},
        ) from exc


def _meaningful_worksheet_bounds(path: Path, worksheet: Any) -> tuple[int, int]:
    """Measure populated XLSX cells, excluding formatting-only dimensions."""

    worksheet_path = getattr(worksheet, "_worksheet_path", None)
    if not worksheet_path:
        return int(worksheet.max_row or 0), int(worksheet.max_column or 0)

    merged_ranges: list[tuple[int, int, int, int]] = []
    with zipfile.ZipFile(path) as archive, archive.open(worksheet_path) as worksheet_xml:
        for _event, element in ET.iterparse(worksheet_xml, events=("end",)):
            if element.tag.endswith("mergeCell"):
                reference = element.attrib.get("ref")
                if reference:
                    merged_ranges.append(range_boundaries(reference))
            element.clear()

    merged_anchors = {
        (min_row, min_column)
        for min_column, min_row, _max_column, _max_row in merged_ranges
    }
    meaningful_merged_anchors: set[tuple[int, int]] = set()
    max_row = 0
    max_column = 0
    with zipfile.ZipFile(path) as archive, archive.open(worksheet_path) as worksheet_xml:
        for _event, element in ET.iterparse(worksheet_xml, events=("end",)):
            if element.tag.endswith("}c") or element.tag == "c":
                reference = element.attrib.get("r")
                if reference:
                    has_formula = any(
                        child.tag.endswith("}f") or child.tag == "f"
                        for child in element
                    )
                    has_value = any(
                        (child.tag.endswith("}v") or child.tag == "v" or child.tag.endswith("}t") or child.tag == "t")
                        and child.text not in {None, ""}
                        for child in element.iter()
                    )
                    if has_formula or has_value:
                        column_letters, row_number = coordinate_from_string(reference)
                        column_number = column_index_from_string(column_letters)
                        if (row_number, column_number) in merged_anchors:
                            meaningful_merged_anchors.add(
                                (row_number, column_number)
                            )
                        max_row = max(max_row, row_number)
                        max_column = max(max_column, column_number)
                element.clear()

    for min_column, min_row, max_merged_column, max_merged_row in merged_ranges:
        if (min_row, min_column) not in meaningful_merged_anchors:
            continue
        max_row = max(max_row, max_merged_row)
        max_column = max(max_column, max_merged_column)
    return max_row, max_column


def _first_non_empty_rows(
    worksheet: Any,
    *,
    max_row: int,
    max_column: int,
) -> list[tuple[int, tuple[object, ...]]]:
    rows: list[tuple[int, tuple[object, ...]]] = []
    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_column,
            values_only=True,
        ),
        start=1,
    ):
        values = tuple(row)
        if not any(value is not None and str(value).strip() for value in values):
            continue
        rows.append((row_number, values))
        if len(rows) == _HEADER_SCAN_NON_EMPTY_ROWS:
            break
    return rows


def _candidate_inspection(candidate: _HeaderCandidate) -> _DetectedHeader:
    mapping = dict(candidate.mapping)
    ambiguous = dict(candidate.ambiguous)
    missing = tuple(
        field
        for field in _REQUIRED_FIELDS
        if field not in mapping and field not in ambiguous
    )
    if ambiguous:
        status = MappingStatus.AMBIGUOUS_MAPPING
    elif missing:
        status = MappingStatus.MAPPING_INCOMPLETE
    else:
        status = MappingStatus.READY
    return _DetectedHeader(
        row_number=candidate.row_number,
        detected_headers=candidate.headers,
        column_mapping=mapping,
        mapping_status=status,
        missing_required_fields=missing,
        ambiguous_fields=ambiguous,
    )


def _inspect_sheet(
    *,
    name: str,
    max_row: int,
    max_column: int,
    non_empty_rows: list[tuple[int, tuple[object, ...]]],
) -> WorksheetInspection:
    candidates = [
        _header_candidate(row_number, values)
        for row_number, values in non_empty_rows
    ]
    mapped_candidates = [
        candidate for candidate in candidates if candidate.score > 0
    ]
    ready_candidates = [
        candidate for candidate in mapped_candidates if candidate.ready
    ]
    mapped_selection = ready_candidates or mapped_candidates
    if mapped_selection:
        selected = min(
            mapped_selection,
            key=lambda item: (-item.score, item.row_number),
        )
    elif candidates:
        selected = max(
            candidates,
            key=lambda item: (
                item.text_cell_count,
                item.non_empty_cell_count,
                -item.row_number,
            ),
        )
    else:
        selected = None
    selected_inspection = _candidate_inspection(selected) if selected else None

    return WorksheetInspection(
        name=name,
        max_row=max_row,
        max_column=max_column,
        header_row_number=(
            selected_inspection.row_number if selected_inspection else None
        ),
        detected_headers=(
            selected_inspection.detected_headers if selected_inspection else ()
        ),
        column_mapping=(
            selected_inspection.column_mapping if selected_inspection else {}
        ),
        mapping_status=(
            selected_inspection.mapping_status
            if selected_inspection
            else MappingStatus.MAPPING_INCOMPLETE
        ),
        missing_required_fields=(
            selected_inspection.missing_required_fields
            if selected_inspection
            else _REQUIRED_FIELDS
        ),
        ambiguous_fields=(
            selected_inspection.ambiguous_fields if selected_inspection else {}
        ),
    )


def _header_candidate(
    row_number: int,
    values: tuple[object, ...],
) -> _HeaderCandidate:
    headers = tuple("" if value is None else str(value).strip() for value in values)
    non_empty_values = tuple(
        value for value in values if value is not None and str(value).strip()
    )
    matches: dict[str, list[int]] = {}
    for column_number, header in enumerate(headers, start=1):
        normalized = normalize_header(header)
        if not normalized:
            continue
        for field, aliases in _NORMALIZED_ALIASES.items():
            if normalized in aliases:
                matches.setdefault(field, []).append(column_number)

    mapping = {
        field: columns[0]
        for field, columns in matches.items()
        if len(columns) == 1
    }
    ambiguous = {
        field: tuple(columns) for field, columns in matches.items() if len(columns) > 1
    }
    return _HeaderCandidate(
        row_number,
        headers,
        mapping,
        ambiguous,
        sum(isinstance(value, str) for value in non_empty_values),
        len(non_empty_values),
    )

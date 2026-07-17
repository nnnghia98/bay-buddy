"""Safe typed cell mutations for Workbook Editor V2.

Configured sessions edit supported unlocked cells by stable column ID. The
legacy price-only path remains available for compatibility with older callers.
"""

from __future__ import annotations

import math
import os
import tempfile
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from services.workbook_formula import (
    WorkbookFormulaError,
    normalize_column_formulas,
    render_excel_formula,
)
from services.workbook_validation import (
    WorkbookValidationError,
    validate_generated_workbook,
)


MAX_SAFE_VND = 1_000_000_000_000
MAX_EXCEL_TEXT_LENGTH = 32_767
MAX_EXCEL_SIGNIFICANT_DIGITS = 15
_EDITABLE_FIELDS = ("net_price", "selling_price")


@dataclass(frozen=True, slots=True)
class PriceChange:
    """Requested price values for one physical worksheet row."""

    row_number: int
    net_price: int | float | Decimal | None = None
    selling_price: int | float | Decimal | None = None
    values: dict[str, object] | None = None


@dataclass(frozen=True, slots=True)
class AuditedPriceChange:
    """One changed cell recorded with its prior and resulting values."""

    row_number: int
    field: str
    column_number: int
    old_value: object
    new_value: object


@dataclass(frozen=True, slots=True)
class WorkbookMutationResult:
    """Summary of the successfully published workbook mutation."""

    changed_cell_count: int
    changes: tuple[AuditedPriceChange, ...]


@dataclass(frozen=True, slots=True)
class StructuralColumnResult:
    column_number: int


def _publish_workbook(workbook, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            dir=output.parent, prefix=f".{output.name}.", suffix=".tmp.xlsx", delete=False
        ) as temporary:
            temporary_path = Path(temporary.name)
        workbook.save(temporary_path)
        validate_generated_workbook(temporary_path)
        os.link(temporary_path, output)
        temporary_path.unlink()
        temporary_path = None
        _fsync_directory(output.parent)
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)


def _enable_full_calculation(workbook: object) -> None:
    calculation = getattr(workbook, "calculation", None)
    if calculation is not None:
        calculation.calcMode = "auto"
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True


def _regenerate_formula_cells(
    worksheet: Worksheet,
    *,
    header_row_number: int,
    column_config: list[dict[str, object]],
    meaningful_max_row: int | None = None,
) -> None:
    try:
        normalized, formula_order = normalize_column_formulas(column_config)
    except WorkbookFormulaError as exc:
        raise _reject(exc.code, str(exc)) from exc
    if not formula_order:
        return
    by_id = {str(item["id"]): item for item in normalized}
    formula_numbers = {
        int(by_id[column_id]["column_number"]) for column_id in formula_order
    }
    data_numbers = [
        int(item["column_number"])
        for item in normalized
        if int(item["column_number"]) not in formula_numbers
    ]
    max_row = meaningful_max_row or int(worksheet.max_row or 0)
    header_end_row = _header_end_row(
        worksheet, header_row_number, max_row=max_row
    )
    for row_number in range(header_end_row + 1, max_row + 1):
        has_record = any(
            worksheet.cell(row_number, column_number).value not in (None, "")
            for column_number in data_numbers
        )
        for column_id in formula_order:
            item = by_id[column_id]
            cell = worksheet.cell(row_number, int(item["column_number"]))
            cell.value = (
                render_excel_formula(
                    item["formula"], normalized, row_number=row_number
                )
                if has_record
                else None
            )


def add_workbook_column(
    source_path: str | Path,
    output_path: str | Path,
    *,
    sheet_name: str,
    header_row_number: int,
    label: str,
    formula: dict[str, object] | None = None,
    column_config: list[dict[str, object]] | None = None,
    meaningful_max_row: int | None = None,
    meaningful_max_column: int | None = None,
) -> StructuralColumnResult:
    """Append one user-owned column without changing source columns."""
    source, output = Path(source_path), Path(output_path)
    _validate_paths(source, output)
    workbook = load_workbook(source, read_only=False, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise _reject("SHEET_NOT_FOUND", "Selected worksheet does not exist.")
        worksheet = workbook[sheet_name]
        max_row = meaningful_max_row or int(worksheet.max_row or 0)
        max_column = meaningful_max_column or int(worksheet.max_column or 0)
        column_number = max_column + 1
        header_end_row = _header_end_row(
            worksheet,
            header_row_number,
            max_row=max_row,
            max_column=max_column,
        )
        worksheet.cell(header_row_number, column_number, label)
        if header_end_row > header_row_number:
            worksheet.merge_cells(
                start_row=header_row_number,
                start_column=column_number,
                end_row=header_end_row,
                end_column=column_number,
            )
        if column_config is not None:
            _regenerate_formula_cells(
                worksheet,
                header_row_number=header_row_number,
                column_config=column_config,
                meaningful_max_row=max_row,
            )
            _enable_full_calculation(workbook)
        elif formula:
            left = int(formula["left_column_number"])
            right = int(formula["right_column_number"])
            operator = str(formula["operator"])
            if operator not in {"+", "-", "*", "/", "%"}:
                raise _reject("INVALID_FORMULA", "Formula operator is not supported.")
            excel_operator = "*" if operator == "%" else operator
            suffix = "/100" if operator == "%" else ""
            for row_number in range(header_end_row + 1, max_row + 1):
                if all(
                    worksheet.cell(row_number, column).value is None
                    for column in range(1, max_column + 1)
                ):
                    continue
                left_ref = f"{get_column_letter(left)}{row_number}"
                right_ref = f"{get_column_letter(right)}{row_number}"
                worksheet.cell(row_number, column_number, f"={left_ref}{excel_operator}{right_ref}{suffix}")
        _publish_workbook(workbook, output)
        return StructuralColumnResult(column_number=column_number)
    except WorkbookMutationError:
        raise
    except Exception as exc:
        raise _reject("STORAGE_WRITE_FAILED", "Workbook column could not be added.") from exc
    finally:
        workbook.close()


def remove_workbook_column(
    source_path: str | Path,
    output_path: str | Path,
    *,
    sheet_name: str,
    column_number: int,
    header_row_number: int | None = None,
    column_config: list[dict[str, object]] | None = None,
    meaningful_max_row: int | None = None,
    meaningful_max_column: int | None = None,
) -> StructuralColumnResult:
    """Remove one user-owned physical column."""
    source, output = Path(source_path), Path(output_path)
    _validate_paths(source, output)
    workbook = load_workbook(source, read_only=False, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise _reject("SHEET_NOT_FOUND", "Selected worksheet does not exist.")
        worksheet = workbook[sheet_name]
        max_column = meaningful_max_column or int(worksheet.max_column or 0)
        if column_number < 1 or column_number > max_column:
            raise _reject("COLUMN_NOT_FOUND", "Column was not found.")
        worksheet.delete_cols(column_number, 1)
        if column_config is not None:
            if header_row_number is None:
                raise _reject("INVALID_ROW", "Header row is required for formula regeneration.")
            _regenerate_formula_cells(
                worksheet,
                header_row_number=header_row_number,
                column_config=column_config,
                meaningful_max_row=meaningful_max_row,
            )
            _enable_full_calculation(workbook)
        _publish_workbook(workbook, output)
        return StructuralColumnResult(column_number=column_number)
    except WorkbookMutationError:
        raise
    except Exception as exc:
        raise _reject("STORAGE_WRITE_FAILED", "Workbook column could not be removed.") from exc
    finally:
        workbook.close()


def update_workbook_column(
    source_path: str | Path,
    output_path: str | Path,
    *,
    sheet_name: str,
    header_row_number: int,
    column_number: int,
    label: str,
    column_config: list[dict[str, object]],
    clear_column_values: bool = False,
    meaningful_max_row: int | None = None,
    meaningful_max_column: int | None = None,
) -> StructuralColumnResult:
    """Update one user column header and regenerate every managed formula."""

    source, output = Path(source_path), Path(output_path)
    _validate_paths(source, output)
    workbook = load_workbook(source, read_only=False, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise _reject("SHEET_NOT_FOUND", "Selected worksheet does not exist.")
        worksheet = workbook[sheet_name]
        max_row = meaningful_max_row or int(worksheet.max_row or 0)
        max_column = meaningful_max_column or int(worksheet.max_column or 0)
        if column_number < 1 or column_number > max_column:
            raise _reject("COLUMN_NOT_FOUND", "Column was not found.")
        worksheet.cell(header_row_number, column_number, label)
        if clear_column_values:
            header_end_row = _header_end_row(
                worksheet,
                header_row_number,
                max_row=max_row,
                max_column=max_column,
            )
            for row_number in range(header_end_row + 1, max_row + 1):
                worksheet.cell(row_number, column_number).value = None
        _regenerate_formula_cells(
            worksheet,
            header_row_number=header_row_number,
            column_config=column_config,
            meaningful_max_row=max_row,
        )
        _enable_full_calculation(workbook)
        _publish_workbook(workbook, output)
        return StructuralColumnResult(column_number=column_number)
    except WorkbookMutationError:
        raise
    except Exception as exc:
        raise _reject("STORAGE_WRITE_FAILED", "Workbook column could not be updated.") from exc
    finally:
        workbook.close()


class WorkbookMutationError(Exception):
    """Expected mutation rejection with a stable, client-safe error code."""

    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code


def _reject(code: str, message: str) -> WorkbookMutationError:
    return WorkbookMutationError(code, message)


def _normalize_vnd(value: object) -> int:
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        raise _reject("INVALID_CELL_VALUE", "Price values must be whole VND amounts.")
    if isinstance(value, float) and not math.isfinite(value):
        raise _reject("INVALID_CELL_VALUE", "Price values must be finite.")

    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise _reject("INVALID_CELL_VALUE", "Price value is invalid.") from exc

    if not decimal_value.is_finite():
        raise _reject("INVALID_CELL_VALUE", "Price values must be finite.")
    if decimal_value != decimal_value.to_integral_value():
        raise _reject("INVALID_CELL_VALUE", "Price values must use whole VND.")
    if decimal_value < 0 or decimal_value > MAX_SAFE_VND:
        raise _reject(
            "INVALID_CELL_VALUE",
            f"Price values must be between 0 and {MAX_SAFE_VND} VND.",
        )
    return int(decimal_value)


def _normalize_numeric(value: object, *, label: str) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        raise _reject("INVALID_CELL_VALUE", f"{label} columns require numeric values or null.")
    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise _reject("INVALID_CELL_VALUE", f"{label} value is invalid.") from exc
    if not decimal_value.is_finite():
        raise _reject("INVALID_CELL_VALUE", f"{label} values must be finite.")
    normalized = decimal_value.normalize()
    if len(normalized.as_tuple().digits) > MAX_EXCEL_SIGNIFICANT_DIGITS:
        raise _reject(
            "INVALID_CELL_VALUE",
            f"{label} values may use at most {MAX_EXCEL_SIGNIFICANT_DIGITS} significant digits.",
        )
    if normalized and not -307 <= normalized.adjusted() <= 307:
        raise _reject(
            "INVALID_CELL_VALUE",
            f"{label} value is outside Excel's supported numeric range.",
        )
    if decimal_value == decimal_value.to_integral_value():
        return int(decimal_value)
    return float(decimal_value)


def _normalize_date(value: object) -> date | datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is not None and value.utcoffset() is not None:
            raise _reject(
                "INVALID_CELL_VALUE",
                "Date columns do not support timezone-aware datetimes.",
            )
        return value
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        raise _reject("INVALID_CELL_VALUE", "Date columns require an ISO date or null.")
    normalized = value.strip()
    if not normalized:
        return None
    try:
        if "T" in normalized or " " in normalized:
            parsed = datetime.fromisoformat(normalized.replace("Z", "+00:00"))
            if parsed.tzinfo is not None and parsed.utcoffset() is not None:
                raise _reject(
                    "INVALID_CELL_VALUE",
                    "Date columns do not support timezone-aware datetimes.",
                )
            return parsed
        return date.fromisoformat(normalized)
    except ValueError as exc:
        raise _reject("INVALID_CELL_VALUE", "Date columns require a valid ISO date.") from exc


def _normalize_configured_value(
    data_type: object,
    value: object,
    *,
    semantic_field: object = None,
) -> object:
    """Normalize a configured value using its semantic rule when present."""
    if semantic_field in _EDITABLE_FIELDS:
        return _normalize_vnd(value)
    if data_type == "text":
        if value is None:
            return None
        if not isinstance(value, str):
            raise _reject("INVALID_CELL_VALUE", "Text columns require text or null.")
        if value.startswith("="):
            raise _reject(
                "INVALID_CELL_VALUE",
                "Text values cannot begin with '='. Use a managed formula column instead.",
            )
        if len(value) > MAX_EXCEL_TEXT_LENGTH:
            raise _reject(
                "INVALID_CELL_VALUE",
                f"Text values may contain at most {MAX_EXCEL_TEXT_LENGTH} characters.",
            )
        return value
    if data_type == "number":
        return _normalize_numeric(value, label="Number")
    if data_type == "currency":
        return _normalize_numeric(value, label="Currency")
    if data_type == "date":
        return _normalize_date(value)
    if data_type == "boolean":
        if value is None or isinstance(value, bool):
            return value
        raise _reject("INVALID_CELL_VALUE", "Boolean columns require true, false, or null.")
    raise _reject("INVALID_CELL_VALUE", "Column data type is not supported.")


def validate_workbook_column_values(
    source_path: str | Path,
    *,
    sheet_name: str,
    header_row_number: int,
    column_number: int,
    data_type: str,
    semantic_field: str | None = None,
    meaningful_max_row: int | None = None,
    meaningful_max_column: int | None = None,
) -> None:
    """Ensure a metadata type change matches every populated physical value."""

    workbook = load_workbook(Path(source_path), read_only=False, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise _reject("SHEET_NOT_FOUND", "Selected worksheet does not exist.")
        worksheet = workbook[sheet_name]
        max_row = meaningful_max_row or int(worksheet.max_row or 0)
        max_column = meaningful_max_column or int(worksheet.max_column or 0)
        if column_number < 1 or column_number > max_column:
            raise _reject("COLUMN_NOT_FOUND", "Column was not found.")
        header_end_row = _header_end_row(
            worksheet,
            header_row_number,
            max_row=max_row,
            max_column=max_column,
        )
        for row_number in range(header_end_row + 1, max_row + 1):
            value = worksheet.cell(row_number, column_number).value
            if value is None:
                continue
            try:
                _normalize_configured_value(
                    data_type,
                    value,
                    semantic_field=semantic_field,
                )
            except WorkbookMutationError as exc:
                raise _reject(
                    "INVALID_COLUMN_TYPE",
                    f"Column contains a value incompatible with {data_type} at row {row_number}.",
                ) from exc
    finally:
        workbook.close()


def _mapped_columns(column_mapping: dict[str, int]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for field in _EDITABLE_FIELDS:
        value = column_mapping.get(field)
        if isinstance(value, bool) or not isinstance(value, int) or value < 1:
            raise _reject(
                "MAPPING_INCOMPLETE", "Both price fields require valid mapped columns."
            )
        columns[field] = value
    if columns["net_price"] == columns["selling_price"]:
        raise _reject("MAPPING_INCOMPLETE", "Price fields must map to different columns.")
    return columns


def _validate_paths(source_path: Path, output_path: Path) -> None:
    try:
        if source_path.resolve() == output_path.resolve():
            raise _reject(
                "STORAGE_WRITE_FAILED",
                "Output must not replace the source workbook.",
            )
        if not source_path.is_file():
            raise _reject("INVALID_XLSX", "Source workbook is unavailable.")
        if output_path.exists():
            raise _reject("STORAGE_WRITE_FAILED", "Output workbook already exists.")
    except WorkbookMutationError:
        raise
    except OSError as exc:
        raise _reject(
            "STORAGE_WRITE_FAILED",
            "Workbook paths could not be accessed.",
        ) from exc


def _is_merged(worksheet: Worksheet, cell: Cell) -> bool:
    return any(cell.coordinate in merged_range for merged_range in worksheet.merged_cells.ranges)


def _validate_editable_cell(worksheet: Worksheet, cell: Cell) -> None:
    if cell.data_type == "f" or (
        isinstance(cell.value, str) and cell.value.startswith("=")
    ):
        raise _reject("CELL_NOT_EDITABLE", "Formula price cells cannot be edited.")
    if _is_merged(worksheet, cell):
        raise _reject("CELL_NOT_EDITABLE", "Merged price cells cannot be edited.")
    if worksheet.protection.sheet and cell.protection.locked:
        raise _reject("CELL_NOT_EDITABLE", "Locked price cells cannot be edited.")


def _prepare_changes(
    worksheet: Worksheet,
    *,
    header_row_number: int,
    columns: dict[str, int],
    changes: Sequence[PriceChange],
    max_changes: int,
    meaningful_max_row: int | None = None,
    meaningful_max_column: int | None = None,
) -> list[tuple[Cell, AuditedPriceChange]]:
    if isinstance(max_changes, bool) or max_changes < 1:
        raise _reject("INVALID_ROW", "Maximum change count must be positive.")
    if not changes or len(changes) > max_changes:
        raise _reject("INVALID_ROW", f"Provide between 1 and {max_changes} row changes.")
    max_row = meaningful_max_row or int(worksheet.max_row or 0)
    max_column = meaningful_max_column or int(worksheet.max_column or 0)
    if header_row_number < 1 or header_row_number > max_row:
        raise _reject("INVALID_ROW", "Header row is outside the worksheet.")
    if any(column > max_column for column in columns.values()):
        raise _reject("MAPPING_INCOMPLETE", "Mapped price column is outside the worksheet.")

    prepared: list[tuple[Cell, AuditedPriceChange]] = []
    seen_rows: set[int] = set()
    for change in changes:
        row_number = change.row_number
        if isinstance(row_number, bool) or not isinstance(row_number, int):
            raise _reject("INVALID_ROW", "Physical row numbers must be integers.")
        if row_number in seen_rows:
            raise _reject("INVALID_ROW", "Duplicate physical row numbers are not allowed.")
        seen_rows.add(row_number)
        if row_number <= header_row_number or row_number > max_row:
            raise _reject("INVALID_ROW", "Physical row is outside the workbook data area.")
        if all(
            worksheet.cell(row=row_number, column=column).value is None
            for column in range(1, max_column + 1)
        ):
            raise _reject("INVALID_ROW", "Blank worksheet rows cannot be edited.")

        requested = {
            "net_price": change.net_price,
            "selling_price": change.selling_price,
        }
        if all(value is None for value in requested.values()):
            raise _reject("INVALID_CELL_VALUE", "Each row must include a price value.")

        for field, value in requested.items():
            if value is None:
                continue
            new_value = _normalize_vnd(value)
            column_number = columns[field]
            cell = worksheet.cell(row=row_number, column=column_number)
            _validate_editable_cell(worksheet, cell)
            if cell.value == new_value:
                continue
            prepared.append(
                (
                    cell,
                    AuditedPriceChange(
                        row_number=row_number,
                        field=field,
                        column_number=column_number,
                        old_value=cell.value,
                        new_value=new_value,
                    ),
                )
            )

    if not prepared:
        raise _reject("NO_CHANGES", "At least one price must differ from the workbook.")
    if len(prepared) > max_changes:
        raise _reject(
            "INVALID_ROW",
            f"A save may change at most {max_changes} price cells.",
        )
    return prepared


def _header_end_row(
    worksheet: Worksheet,
    header_row_number: int,
    *,
    max_row: int | None = None,
    max_column: int | None = None,
) -> int:
    max_row = max_row or int(worksheet.max_row or 0)
    max_column = max_column or int(worksheet.max_column or 0)
    header_end_row = max(
        (
            merged.max_row
            for merged in worksheet.merged_cells.ranges
            if merged.min_row <= header_row_number <= merged.max_row
        ),
        default=header_row_number,
    )
    header_end_row = min(header_end_row, max_row)
    if header_end_row == header_row_number and header_row_number < max_row:
        next_values = [
            worksheet.cell(header_row_number + 1, column).value
            for column in range(1, max_column + 1)
        ]
        top_values = [
            worksheet.cell(header_row_number, column).value
            for column in range(1, max_column + 1)
        ]
        populated = [value for value in next_values if value not in (None, "")]
        if (
            populated
            and len(populated) <= max(4, max_column // 2)
            and all(isinstance(value, str) for value in populated)
            and any(value in (None, "") for value in top_values)
        ):
            header_end_row += 1
    return header_end_row


def _prepare_configured_changes(
    worksheet: Worksheet,
    *,
    header_row_number: int,
    column_config: list[dict[str, object]],
    changes: Sequence[PriceChange],
    max_changes: int,
    meaningful_max_row: int | None = None,
    meaningful_max_column: int | None = None,
) -> list[tuple[Cell, AuditedPriceChange]]:
    if isinstance(max_changes, bool) or max_changes < 1:
        raise _reject("INVALID_ROW", "Maximum change count must be positive.")
    if not changes or len(changes) > max_changes:
        raise _reject("INVALID_ROW", f"Provide between 1 and {max_changes} row changes.")
    max_row = meaningful_max_row or int(worksheet.max_row or 0)
    max_column = meaningful_max_column or int(worksheet.max_column or 0)
    if header_row_number < 1 or header_row_number > max_row:
        raise _reject("INVALID_ROW", "Header row is outside the worksheet.")

    by_id: dict[str, dict[str, object]] = {}
    by_semantic: dict[str, dict[str, object]] = {}
    for raw_item in column_config:
        item = dict(raw_item)
        column_id = item.get("id")
        column_number = item.get("column_number")
        if (
            not isinstance(column_id, str)
            or not column_id
            or isinstance(column_number, bool)
            or not isinstance(column_number, int)
            or column_number < 1
            or column_number > max_column
        ):
            raise _reject("INVALID_MAPPING", "Column configuration is invalid.")
        if column_id in by_id:
            raise _reject("INVALID_MAPPING", "Column IDs must be unique.")
        by_id[column_id] = item
        semantic_field = item.get("semantic_field")
        if isinstance(semantic_field, str) and semantic_field:
            by_semantic[semantic_field] = item

    prepared: list[tuple[Cell, AuditedPriceChange]] = []
    seen_rows: set[int] = set()
    header_end_row = _header_end_row(
        worksheet,
        header_row_number,
        max_row=max_row,
        max_column=max_column,
    )
    for change in changes:
        row_number = change.row_number
        if isinstance(row_number, bool) or not isinstance(row_number, int):
            raise _reject("INVALID_ROW", "Physical row numbers must be integers.")
        if row_number in seen_rows:
            raise _reject("INVALID_ROW", "Duplicate physical row numbers are not allowed.")
        seen_rows.add(row_number)
        if row_number <= header_end_row or row_number > max_row:
            raise _reject("INVALID_ROW", "Physical row is outside the workbook data area.")
        if all(
            worksheet.cell(
                row=row_number,
                column=int(item["column_number"]),
            ).value is None
            for item in by_id.values()
        ):
            raise _reject("INVALID_ROW", "Blank worksheet rows cannot be edited.")

        requested = list((change.values or {}).items())
        if change.net_price is not None:
            requested.append(("net_price", change.net_price))
        if change.selling_price is not None:
            requested.append(("selling_price", change.selling_price))
        if not requested:
            raise _reject("INVALID_CELL_VALUE", "Each row must include at least one cell value.")

        seen_targets: set[tuple[int, int]] = set()
        for requested_field, value in requested:
            item = by_id.get(requested_field) or by_semantic.get(requested_field)
            if item is None or item.get("formula"):
                raise _reject("COLUMN_NOT_EDITABLE", "Column is not editable.")
            column_number = int(item["column_number"])
            target = (row_number, column_number)
            if target in seen_targets:
                raise _reject(
                    "INVALID_CELL_VALUE",
                    "A cell may be updated only once per save request.",
                )
            seen_targets.add(target)
            cell = worksheet.cell(row=row_number, column=column_number)
            _validate_editable_cell(worksheet, cell)
            new_value = _normalize_configured_value(
                item.get("data_type", "text"),
                value,
                semantic_field=item.get("semantic_field"),
            )
            if cell.value == new_value:
                continue
            prepared.append(
                (
                    cell,
                    AuditedPriceChange(
                        row_number=row_number,
                        field=requested_field,
                        column_number=column_number,
                        old_value=cell.value,
                        new_value=new_value,
                    ),
                )
            )

    if not prepared:
        raise _reject("NO_CHANGES", "At least one cell must differ from the workbook.")
    if len(prepared) > max_changes:
        raise _reject(
            "INVALID_ROW",
            f"A save may change at most {max_changes} cells.",
        )
    return prepared


def _fsync_directory(directory: Path) -> None:
    descriptor = os.open(directory, os.O_RDONLY)
    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def _cell_values_equal(expected: object, generated: object) -> bool:
    if isinstance(expected, date) and not isinstance(expected, datetime) and isinstance(generated, datetime):
        return generated.time().replace(tzinfo=None) == datetime.min.time() and generated.date() == expected
    return expected == generated


def _validate_generated_workbook(
    path: Path,
    *,
    expected_workbook,
    sheet_name: str,
    audited_changes: Sequence[AuditedPriceChange],
) -> None:
    try:
        validate_generated_workbook(path)
    except WorkbookValidationError as exc:
        raise _reject(
            "STORAGE_WRITE_FAILED",
            "Generated workbook could not be validated.",
        ) from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        raise _reject("STORAGE_WRITE_FAILED", "Generated workbook could not be validated.") from exc
    try:
        if workbook.sheetnames != expected_workbook.sheetnames:
            raise _reject(
                "STORAGE_WRITE_FAILED",
                "Generated workbook changed worksheet structure.",
            )
        for expected_sheet, generated_sheet in zip(
            expected_workbook.worksheets, workbook.worksheets
        ):
            if (
                expected_sheet.max_row != generated_sheet.max_row
                or expected_sheet.max_column != generated_sheet.max_column
            ):
                raise _reject(
                    "STORAGE_WRITE_FAILED",
                    "Generated workbook changed worksheet structure.",
                )
        if sheet_name not in workbook.sheetnames:
            raise _reject("STORAGE_WRITE_FAILED", "Generated workbook lost its worksheet.")
        worksheet = workbook[sheet_name]
        for change in audited_changes:
            if not _cell_values_equal(
                change.new_value,
                worksheet.cell(change.row_number, change.column_number).value,
            ):
                raise _reject("STORAGE_WRITE_FAILED", "Generated workbook failed validation.")
    finally:
        workbook.close()


def apply_price_changes(
    source_path: str | Path,
    output_path: str | Path,
    *,
    sheet_name: str,
    header_row_number: int,
    column_mapping: dict[str, int],
    changes: Sequence[PriceChange],
    max_changes: int = 500,
    column_config: list[dict[str, object]] | None = None,
    meaningful_max_row: int | None = None,
    meaningful_max_column: int | None = None,
) -> WorkbookMutationResult:
    """Apply validated price changes and atomically publish a new workbook."""

    source = Path(source_path)
    output = Path(output_path)
    _validate_paths(source, output)
    columns = _mapped_columns(column_mapping) if not column_config else {}
    try:
        output.parent.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise _reject(
            "STORAGE_WRITE_FAILED",
            "Workbook output directory is unavailable.",
        ) from exc

    try:
        workbook = load_workbook(source, read_only=False, data_only=False)
    except Exception as exc:
        raise _reject("INVALID_XLSX", "Source workbook could not be opened.") from exc

    temporary_path: Path | None = None
    try:
        if sheet_name not in workbook.sheetnames:
            raise _reject("SHEET_NOT_FOUND", "Selected worksheet does not exist.")
        worksheet = workbook[sheet_name]
        if column_config:
            prepared = _prepare_configured_changes(
                worksheet,
                header_row_number=header_row_number,
                column_config=column_config,
                changes=changes,
                max_changes=max_changes,
                meaningful_max_row=meaningful_max_row,
                meaningful_max_column=meaningful_max_column,
            )
        else:
            prepared = _prepare_changes(
                worksheet,
                header_row_number=header_row_number,
                columns=columns,
                changes=changes,
                max_changes=max_changes,
                meaningful_max_row=meaningful_max_row,
                meaningful_max_column=meaningful_max_column,
            )
        for cell, audit in prepared:
            cell.value = audit.new_value

        with tempfile.NamedTemporaryFile(
            dir=output.parent,
            prefix=f".{output.name}.",
            suffix=".tmp.xlsx",
            delete=False,
        ) as temporary:
            temporary_path = Path(temporary.name)
        workbook.save(temporary_path)
        with temporary_path.open("rb") as generated_file:
            os.fsync(generated_file.fileno())

        audited_changes = tuple(audit for _, audit in prepared)
        _validate_generated_workbook(
            temporary_path,
            expected_workbook=workbook,
            sheet_name=sheet_name,
            audited_changes=audited_changes,
        )
        try:
            os.link(temporary_path, output)
        except FileExistsError as exc:
            raise _reject("STORAGE_WRITE_FAILED", "Output workbook already exists.") from exc
        try:
            temporary_path.unlink()
            temporary_path = None
        except OSError:
            # Publication already succeeded. A leftover unreferenced temporary
            # file is preferable to reporting failure for a valid immutable output.
            pass
        _fsync_directory(output.parent)
        return WorkbookMutationResult(
            changed_cell_count=len(audited_changes),
            changes=audited_changes,
        )
    except WorkbookMutationError:
        raise
    except Exception as exc:
        raise _reject("STORAGE_WRITE_FAILED", "Workbook output could not be written.") from exc
    finally:
        workbook.close()
        if temporary_path is not None:
            try:
                temporary_path.unlink(missing_ok=True)
            except OSError:
                pass

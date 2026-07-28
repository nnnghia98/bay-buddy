"""Presentation-only formatting for downloadable Workbook Editor files."""

from __future__ import annotations

import math
import unicodedata
from copy import copy
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from services.workbook_reader import read_header_structure
from services.workbook_validation import (
    WorkbookValidationError,
    validate_generated_workbook,
)


_PRIMARY = "1B61C9"
_PRIMARY_SOFT = "EDF5FF"
_FOREGROUND = "181D26"
_ROW_ALT = "F7F9FC"
_WHITE = "FFFFFF"
_BODY_FONT = "Aptos"
_MIN_ROW_HEIGHT = 20
_MAX_ROW_HEIGHT = 48
_MAX_TEXT_COLUMN_WIDTH = 44
_MAX_TYPED_COLUMN_WIDTH = 22


class WorkbookExportError(RuntimeError):
    """Safe failure raised while preparing a presentation export."""


def format_workbook_for_export(
    source_path: str | Path,
    output_path: str | Path,
    *,
    sheet_name: str,
    header_row_number: int,
    column_config: list[dict[str, object]],
    meaningful_max_row: int,
    meaningful_max_column: int,
    sheet_metadata: list[dict[str, object]] | None = None,
) -> None:
    """Create a readable XLSX copy while preserving workbook values and formulas."""

    source = Path(source_path)
    output = Path(output_path)
    try:
        workbook = load_workbook(source, read_only=False, data_only=False)
    except Exception as exc:
        raise WorkbookExportError("Source workbook could not be opened.") from exc

    try:
        if sheet_name not in workbook.sheetnames:
            raise WorkbookExportError("Selected worksheet does not exist.")
        worksheet = workbook[sheet_name]
        _format_selected_worksheet(
            worksheet,
            header_row_number=header_row_number,
            column_config=column_config,
            meaningful_max_row=meaningful_max_row,
            meaningful_max_column=meaningful_max_column,
        )
        for raw_sheet in sheet_metadata or []:
            metadata = dict(raw_sheet)
            metadata_name = metadata.get("name")
            if metadata_name == sheet_name:
                continue
            if (
                not isinstance(metadata_name, str)
                or metadata_name not in workbook.sheetnames
            ):
                raise WorkbookExportError("Workbook sheet metadata is invalid.")
            _format_metadata_worksheet(
                workbook[metadata_name],
                metadata=metadata,
            )
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
    except WorkbookExportError:
        raise
    except Exception as exc:
        raise WorkbookExportError("Workbook export could not be formatted.") from exc
    finally:
        workbook.close()

    try:
        validate_generated_workbook(output)
    except (OSError, WorkbookValidationError) as exc:
        raise WorkbookExportError("Formatted workbook failed validation.") from exc


def _format_metadata_worksheet(
    worksheet: Worksheet,
    *,
    metadata: dict[str, object],
) -> None:
    header_row_number = metadata.get("header_row_number")
    meaningful_max_row = metadata.get("max_row")
    meaningful_max_column = metadata.get("max_column")
    detected_headers = metadata.get("detected_headers")
    if (
        isinstance(header_row_number, bool)
        or not isinstance(header_row_number, int)
        or isinstance(meaningful_max_row, bool)
        or not isinstance(meaningful_max_row, int)
        or isinstance(meaningful_max_column, bool)
        or not isinstance(meaningful_max_column, int)
        or not isinstance(detected_headers, list)
        or len(detected_headers) < meaningful_max_column
    ):
        raise WorkbookExportError("Workbook sheet metadata is invalid.")

    header_end_row, _structure = read_header_structure(
        worksheet,
        header_row_number,
        max_row=meaningful_max_row,
        max_column=meaningful_max_column,
    )
    inferred_config = [
        {
            "id": f"export-{column_number}",
            "label": str(detected_headers[column_number - 1] or ""),
            "column_number": column_number,
            "data_type": _infer_data_type(
                worksheet,
                column_number=column_number,
                label=str(detected_headers[column_number - 1] or ""),
                start_row=header_end_row + 1,
                max_row=meaningful_max_row,
            ),
        }
        for column_number in range(1, meaningful_max_column + 1)
    ]
    _format_selected_worksheet(
        worksheet,
        header_row_number=header_row_number,
        column_config=inferred_config,
        meaningful_max_row=meaningful_max_row,
        meaningful_max_column=meaningful_max_column,
    )


def _format_selected_worksheet(
    worksheet: Worksheet,
    *,
    header_row_number: int,
    column_config: list[dict[str, object]],
    meaningful_max_row: int,
    meaningful_max_column: int,
) -> None:
    if (
        isinstance(header_row_number, bool)
        or not isinstance(header_row_number, int)
        or header_row_number < 1
        or header_row_number > meaningful_max_row
        or meaningful_max_row > int(worksheet.max_row or 0)
        or meaningful_max_column > int(worksheet.max_column or 0)
    ):
        raise WorkbookExportError("Worksheet export bounds are invalid.")

    header_end_row, _structure = read_header_structure(
        worksheet,
        header_row_number,
        max_row=meaningful_max_row,
        max_column=meaningful_max_column,
    )
    config_by_column = _column_config_by_number(
        column_config,
        meaningful_max_column=meaningful_max_column,
    )
    wrapped_columns = _fit_column_widths(
        worksheet,
        header_row_number=header_row_number,
        header_end_row=header_end_row,
        max_row=meaningful_max_row,
        max_column=meaningful_max_column,
        config_by_column=config_by_column,
    )

    _style_header_band(
        worksheet,
        header_row_number=header_row_number,
        header_end_row=header_end_row,
        max_column=meaningful_max_column,
    )
    _style_data_rows(
        worksheet,
        start_row=header_end_row + 1,
        max_row=meaningful_max_row,
        max_column=meaningful_max_column,
        config_by_column=config_by_column,
        wrapped_columns=wrapped_columns,
    )

    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 90
    if worksheet.freeze_panes is None and header_end_row < meaningful_max_row:
        worksheet.freeze_panes = f"A{header_end_row + 1}"
    if worksheet.print_title_rows is None:
        worksheet.print_title_rows = (
            f"{header_row_number}:{header_end_row}"
        )
    if (
        header_end_row == header_row_number
        and not worksheet.auto_filter.ref
        and len(worksheet.tables) == 0
        and header_end_row < meaningful_max_row
    ):
        last_column = get_column_letter(meaningful_max_column)
        worksheet.auto_filter.ref = (
            f"A{header_end_row}:{last_column}{meaningful_max_row}"
        )


def _column_config_by_number(
    column_config: list[dict[str, object]],
    *,
    meaningful_max_column: int,
) -> dict[int, dict[str, object]]:
    configured: dict[int, dict[str, object]] = {}
    for raw_item in column_config:
        item = dict(raw_item)
        column_number = item.get("column_number")
        if (
            isinstance(column_number, bool)
            or not isinstance(column_number, int)
            or column_number < 1
            or column_number > meaningful_max_column
        ):
            raise WorkbookExportError("Column configuration is invalid.")
        configured[column_number] = item
    return configured


def _fit_column_widths(
    worksheet: Worksheet,
    *,
    header_row_number: int,
    header_end_row: int,
    max_row: int,
    max_column: int,
    config_by_column: dict[int, dict[str, object]],
) -> set[int]:
    wrapped_columns: set[int] = set()
    for column_number in range(1, max_column + 1):
        item = config_by_column.get(column_number, {})
        data_type = str(item.get("data_type") or "text")
        label = str(item.get("label") or "")
        observed_width = max(
            _visual_text_width(line) for line in label.splitlines() or [""]
        )
        has_line_break = "\n" in label

        for row_number in range(header_row_number, max_row + 1):
            cell = worksheet.cell(row_number, column_number)
            if isinstance(cell, MergedCell):
                continue
            if (
                row_number <= header_end_row
                and _is_horizontal_header_merge(
                    worksheet,
                    row_number=row_number,
                    column_number=column_number,
                )
            ):
                continue
            text = _display_text(cell.value, data_type=data_type)
            if not text:
                continue
            has_line_break = has_line_break or "\n" in text
            observed_width = max(
                observed_width,
                *(_visual_text_width(line) for line in text.splitlines()),
            )

        minimum_width, maximum_width = _width_limits(data_type)
        desired_width = max(minimum_width, observed_width + 2)
        if desired_width > maximum_width or has_line_break:
            wrapped_columns.add(column_number)
        dimension = worksheet.column_dimensions[get_column_letter(column_number)]
        dimension.width = min(desired_width, maximum_width)
        dimension.bestFit = True
    return wrapped_columns


def _style_header_band(
    worksheet: Worksheet,
    *,
    header_row_number: int,
    header_end_row: int,
    max_column: int,
) -> None:
    vertically_merged: set[tuple[int, int]] = set()
    for merged in worksheet.merged_cells.ranges:
        if (
            merged.min_row <= header_row_number
            and merged.max_row >= header_end_row
        ):
            vertically_merged.update(
                (row_number, column_number)
                for row_number in range(header_row_number, header_end_row + 1)
                for column_number in range(
                    max(1, merged.min_col),
                    min(max_column, merged.max_col) + 1,
                )
            )

    for row_number in range(header_row_number, header_end_row + 1):
        is_primary_row = row_number == header_row_number
        for column_number in range(1, max_column + 1):
            cell = worksheet.cell(row_number, column_number)
            use_primary = is_primary_row or (
                row_number,
                column_number,
            ) in vertically_merged
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=_PRIMARY if use_primary else _PRIMARY_SOFT,
            )
            cell.font = Font(
                name=_BODY_FONT,
                size=10,
                bold=True,
                color=_WHITE if use_primary else _FOREGROUND,
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        row_dimension = worksheet.row_dimensions[row_number]
        target_height = 26 if is_primary_row else 24
        row_dimension.height = max(row_dimension.height or 0, target_height)


def _style_data_rows(
    worksheet: Worksheet,
    *,
    start_row: int,
    max_row: int,
    max_column: int,
    config_by_column: dict[int, dict[str, object]],
    wrapped_columns: set[int],
) -> None:
    alternate_fill = PatternFill(fill_type="solid", fgColor=_ROW_ALT)
    for row_number in range(start_row, max_row + 1):
        row_height = _MIN_ROW_HEIGHT
        is_alternate = (row_number - start_row) % 2 == 1
        for column_number in range(1, max_column + 1):
            cell = worksheet.cell(row_number, column_number)
            if isinstance(cell, MergedCell):
                continue
            item = config_by_column.get(column_number, {})
            data_type = str(item.get("data_type") or "text")
            text = _display_text(cell.value, data_type=data_type)
            should_wrap = column_number in wrapped_columns and bool(text)

            font = copy(cell.font)
            font.name = _BODY_FONT
            font.sz = 10
            cell.font = font

            alignment = copy(cell.alignment)
            alignment.horizontal = _horizontal_alignment(
                cell.value,
                data_type=data_type,
            )
            alignment.vertical = "top" if should_wrap else "center"
            alignment.wrap_text = should_wrap
            cell.alignment = alignment

            if is_alternate and cell.fill.fill_type is None:
                cell.fill = alternate_fill
            _apply_readable_number_format(cell, data_type=data_type)

            if should_wrap:
                column_width = worksheet.column_dimensions[
                    get_column_letter(column_number)
                ].width or _MAX_TEXT_COLUMN_WIDTH
                line_count = max(
                    1,
                    sum(
                        max(1, math.ceil(_visual_text_width(line) / column_width))
                        for line in text.splitlines()
                    ),
                )
                row_height = max(
                    row_height,
                    min(_MAX_ROW_HEIGHT, 15 * min(line_count, 3) + 3),
                )

        row_dimension = worksheet.row_dimensions[row_number]
        row_dimension.height = max(row_dimension.height or 0, row_height)


def _apply_readable_number_format(cell: Any, *, data_type: str) -> None:
    if data_type == "text":
        cell.number_format = "@"
    elif cell.number_format == "General" and data_type == "currency":
        cell.number_format = "#,##0"
    elif cell.number_format == "General" and data_type == "number":
        cell.number_format = "#,##0.##"
    elif cell.number_format == "General" and data_type == "date":
        cell.number_format = "dd/mm/yyyy"


def _horizontal_alignment(value: object, *, data_type: str) -> str:
    if data_type in {"number", "currency"}:
        return "right"
    if data_type in {"date", "boolean"} or isinstance(
        value,
        (date, datetime, time),
    ):
        return "center"
    return "left"


def _infer_data_type(
    worksheet: Worksheet,
    *,
    column_number: int,
    label: str,
    start_row: int,
    max_row: int,
) -> str:
    normalized_label = " ".join(label.casefold().split())
    if normalized_label == "no.":
        return "number"
    if any(
        token in normalized_label
        for token in (
            "tkt",
            "ticket",
            "pnr",
            "phone",
            "doc.nbr",
            "rc nbr",
            "số hd",
        )
    ):
        return "text"
    if any(
        token in normalized_label
        for token in ("date", "ngày", "departure", "return")
    ):
        return "date"
    if any(
        token in normalized_label
        for token in (
            "amount",
            "fare",
            "vat",
            "tax",
            "charge",
            "fee",
            "penalty",
            "comm",
            "total",
            "pay",
        )
    ):
        return "currency"

    sampled = [
        worksheet.cell(row_number, column_number).value
        for row_number in range(start_row, min(max_row, start_row + 49) + 1)
        if worksheet.cell(row_number, column_number).value not in (None, "")
    ]
    if sampled and all(isinstance(value, bool) for value in sampled):
        return "boolean"
    if sampled and all(isinstance(value, (date, datetime, time)) for value in sampled):
        return "date"
    if sampled and all(
        isinstance(value, (int, float, Decimal)) and not isinstance(value, bool)
        for value in sampled
    ):
        return "number"
    return "text"


def _width_limits(data_type: str) -> tuple[int, int]:
    if data_type == "currency":
        return 14, _MAX_TYPED_COLUMN_WIDTH
    if data_type == "number":
        return 10, _MAX_TYPED_COLUMN_WIDTH
    if data_type == "date":
        return 12, 16
    if data_type == "boolean":
        return 10, 14
    return 10, _MAX_TEXT_COLUMN_WIDTH


def _display_text(value: object, *, data_type: str) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, timedelta):
        return str(value)
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, Decimal)) and not isinstance(value, bool):
        return f"{value:,}"
    if isinstance(value, float):
        if not math.isfinite(value):
            return str(value)
        if data_type == "text" and value.is_integer():
            return str(int(value))
        return f"{value:,.2f}".rstrip("0").rstrip(".")
    if isinstance(value, str) and value.startswith("="):
        return ""
    return str(value).strip()


def _visual_text_width(value: str) -> int:
    return sum(
        2 if unicodedata.east_asian_width(character) in {"F", "W"} else 1
        for character in value
    )


def _is_horizontal_header_merge(
    worksheet: Worksheet,
    *,
    row_number: int,
    column_number: int,
) -> bool:
    return any(
        merged.min_col < merged.max_col
        and merged.min_row <= row_number <= merged.max_row
        and merged.min_col <= column_number <= merged.max_col
        for merged in worksheet.merged_cells.ranges
    )

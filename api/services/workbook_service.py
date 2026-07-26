"""Application orchestration for the Workbook Editor V2 MVP."""

from __future__ import annotations

import hashlib
import json
import math
import re
import tempfile
import threading
import uuid
from contextlib import contextmanager
from dataclasses import asdict, dataclass, field
from decimal import Decimal
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, BinaryIO, Iterator, Sequence

from openpyxl import load_workbook
from sqlalchemy import func, or_
from sqlalchemy.exc import IntegrityError
from sqlmodel import Session, select

from models import (
    User,
    UserRole,
    Workbook,
    WorkbookOperation,
    WorkbookOperationType,
    WorkbookSession,
    WorkbookSessionStatus,
    WorkbookVersion,
)
from services.workbook_formula import (
    WorkbookFormulaError,
    dependent_formula_columns,
    evaluate_normalized_formula_columns,
    normalize_column_formulas,
    normalize_formula,
    referenced_column_ids,
    render_readable_expression,
)
from services.workbook_mutation import (
    PriceChange,
    WorkbookMutationError,
    apply_price_changes,
    add_workbook_column,
    remove_workbook_column,
    update_workbook_column,
    validate_workbook_column_values,
)
from services.workbook_reader import (
    WorkbookCellReference,
    WorkbookCellValueResult,
    WorkbookReadError,
    WorkbookRecordPage,
    read_header_structure,
    read_hidden_columns,
    read_workbook_cell_values,
    read_workbook_records,
)
from services.workbook_validation import (
    WorkbookValidationError,
    WorksheetInspection,
    validate_and_inspect_workbook,
)
from services.workbook_xls_conversion import convert_xls_to_xlsx
from storage.workbooks import WorkbookStorage
from models.workbook import utc_now


XLSX_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
XLS_MIME_TYPE = "application/vnd.ms-excel"


def _utc_datetime(value: datetime) -> datetime:
    """Normalize SQLite's timezone-naive DateTime reads to explicit UTC."""

    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)
_SUPPORTED_UPLOAD_MIME_TYPES = frozenset(
    {XLSX_MIME_TYPE, XLS_MIME_TYPE, "application/octet-stream"}
)
_COPY_CHUNK_SIZE = 1024 * 1024
_sqlite_save_locks: dict[uuid.UUID, threading.RLock] = {}
_sqlite_save_locks_guard = threading.Lock()
_verified_local_objects: set[tuple[int, int, int, int, str]] = set()
_verified_local_objects_guard = threading.Lock()


class WorkbookServiceError(RuntimeError):
    """Safe application error suitable for conversion to an API response."""

    def __init__(
        self,
        code: str,
        status_code: int,
        message: str,
        *,
        details: dict[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.status_code = status_code
        self.message = message
        self.details = details or {}


@dataclass(frozen=True, slots=True)
class WorkbookUploadResult:
    id: uuid.UUID
    original_filename: str
    mime_type: str
    file_size: int
    checksum: str
    sheet_count: int
    sheets: tuple[WorksheetInspection, ...]
    created_at: datetime


@dataclass(frozen=True, slots=True)
class EditingSessionDescriptor:
    id: uuid.UUID
    workbook_id: uuid.UUID
    original_filename: str
    selected_sheet_name: str
    header_row_number: int
    column_mapping: dict[str, int]
    current_version: int
    status: WorkbookSessionStatus
    created_at: datetime
    updated_at: datetime
    column_config: list[dict[str, Any]] = field(default_factory=list)


@dataclass(frozen=True, slots=True)
class SessionSummaryDescriptor:
    id: uuid.UUID
    display_name: str
    original_filename: str
    selected_sheet_name: str
    current_version: int
    status: WorkbookSessionStatus
    created_at: datetime
    updated_at: datetime


@dataclass(frozen=True, slots=True)
class SessionListResult:
    items: tuple[SessionSummaryDescriptor, ...]
    page: int
    page_size: int
    total: int
    total_pages: int


@dataclass(frozen=True, slots=True)
class SessionRecordsResult:
    session_id: uuid.UUID
    version: int
    sheet_name: str
    page: WorkbookRecordPage


@dataclass(frozen=True, slots=True)
class SessionCellValuesResult:
    session_id: uuid.UUID
    version: int
    cells: tuple[WorkbookCellValueResult, ...]


@dataclass(frozen=True, slots=True)
class FormulaPreviewRow:
    row_number: int
    value: int | float | None
    error_code: str | None = None
    error_message: str | None = None


@dataclass(frozen=True, slots=True)
class FormulaPreviewResult:
    valid: bool
    normalized_formula: dict[str, Any] | None
    readable_expression: str | None
    referenced_column_ids: tuple[str, ...]
    results: tuple[FormulaPreviewRow, ...]
    errors: tuple[dict[str, Any], ...] = ()
    warnings: tuple[dict[str, Any], ...] = ()


@dataclass(frozen=True, slots=True)
class WorkbookSaveResult:
    operation_id: uuid.UUID
    request_id: uuid.UUID
    previous_version: int
    current_version: int
    changed_cells: int
    saved_at: datetime
    checksum: str
    file_size: int
    replayed: bool = False


@dataclass(frozen=True, slots=True)
class WorkbookDownloadDescriptor:
    stream: BinaryIO
    filename: str
    mime_type: str
    checksum: str
    file_size: int
    version: int


class _DeletingBinaryStream:
    """Binary stream that removes its verified temporary file when closed."""

    def __init__(self, path: Path) -> None:
        self._path = path
        self._stream = path.open("rb")

    def read(self, size: int = -1) -> bytes:
        return self._stream.read(size)

    def close(self) -> None:
        self._stream.close()
        try:
            self._path.unlink(missing_ok=True)
        except OSError:
            pass


def _id(value: uuid.UUID | None) -> uuid.UUID:
    if value is None:  # pragma: no cover - persisted entities always have IDs
        raise WorkbookServiceError("INTERNAL_ERROR", 500, "Workbook state is invalid.")
    return value


def _safe_original_filename(filename: str) -> str:
    if not isinstance(filename, str):
        raise WorkbookServiceError(
            "UNSUPPORTED_FILE_TYPE", 415, "Only .xlsx and .xls workbooks are supported."
        )
    basename = filename.replace("\\", "/").rsplit("/", 1)[-1].strip()
    if (
        not basename
        or len(basename) > 255
        or Path(basename).suffix.casefold() not in {".xls", ".xlsx"}
    ):
        raise WorkbookServiceError(
            "UNSUPPORTED_FILE_TYPE", 415, "Only .xlsx and .xls workbooks are supported."
        )
    return basename


def _authorize_owner(*, actor: User, owner_id: uuid.UUID) -> None:
    if actor.role != UserRole.ADMIN and actor.id != owner_id:
        # Ownership failures deliberately look like missing records.
        raise WorkbookServiceError("SESSION_NOT_FOUND", 404, "Session was not found.")


def _error_status(code: str) -> int:
    if code == "FILE_TOO_LARGE":
        return 413
    if code == "UNSUPPORTED_FILE_TYPE":
        return 415
    if code in {"WORKBOOK_NOT_FOUND", "SESSION_NOT_FOUND", "SHEET_NOT_FOUND"}:
        return 404
    if code in {
        "SESSION_NOT_ACTIVE",
        "VERSION_CONFLICT",
        "IDEMPOTENCY_KEY_REUSED",
    }:
        return 409
    if code in {"STORAGE_OBJECT_MISSING", "STORAGE_WRITE_FAILED"}:
        return 500
    return 422


def _translate_domain_error(error: Exception) -> WorkbookServiceError:
    code = getattr(error, "code", "INVALID_XLSX")
    message = str(error) or "Workbook request could not be completed."
    details = getattr(error, "details", {})
    return WorkbookServiceError(
        code,
        _error_status(code),
        message,
        details=details if isinstance(details, dict) else {},
    )


def _sheet_metadata(sheet: WorksheetInspection) -> dict[str, Any]:
    return {
        "name": sheet.name,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "header_row_number": sheet.header_row_number,
        "detected_headers": list(sheet.detected_headers),
        "column_mapping": dict(sheet.column_mapping),
        "mapping_status": sheet.mapping_status.value,
        "missing_required_fields": list(sheet.missing_required_fields),
        "ambiguous_fields": {
            field: list(columns) for field, columns in sheet.ambiguous_fields.items()
        },
    }


def _metadata_sheet(workbook: Workbook, sheet_name: str) -> dict[str, Any]:
    for raw_sheet in workbook.sheet_metadata:
        if raw_sheet.get("name") == sheet_name:
            return raw_sheet
    raise WorkbookServiceError("SHEET_NOT_FOUND", 404, "Worksheet was not found.")


def _session_meaningful_bounds(
    editing_session: WorkbookSession,
    workbook: Workbook,
) -> tuple[int, int]:
    """Return persisted processing bounds, with upload metadata for old rows."""

    sheet = _metadata_sheet(workbook, editing_session.selected_sheet_name)
    max_row = editing_session.meaningful_max_row or sheet.get("max_row")
    max_column = editing_session.meaningful_max_column or sheet.get("max_column")
    if (
        isinstance(max_row, bool)
        or not isinstance(max_row, int)
        or max_row < 1
        or isinstance(max_column, bool)
        or not isinstance(max_column, int)
        or max_column < 1
    ):
        raise WorkbookServiceError(
            "INVALID_XLSX", 422, "Worksheet processing bounds are unavailable."
        )
    return max_row, max_column


def _get_workbook(db: Session, workbook_id: uuid.UUID, actor: User) -> Workbook:
    workbook = db.get(Workbook, workbook_id)
    if workbook is None:
        raise WorkbookServiceError("WORKBOOK_NOT_FOUND", 404, "Workbook was not found.")
    if actor.role != UserRole.ADMIN and workbook.created_by != actor.id:
        raise WorkbookServiceError("WORKBOOK_NOT_FOUND", 404, "Workbook was not found.")
    return workbook


def _get_session(db: Session, session_id: uuid.UUID, actor: User) -> WorkbookSession:
    editing_session = db.get(WorkbookSession, session_id)
    if editing_session is None:
        raise WorkbookServiceError("SESSION_NOT_FOUND", 404, "Session was not found.")
    _authorize_owner(actor=actor, owner_id=editing_session.created_by)
    return editing_session


def _get_version(
    db: Session, *, session_id: uuid.UUID, version_number: int
) -> WorkbookVersion:
    version = db.exec(
        select(WorkbookVersion).where(
            WorkbookVersion.session_id == session_id,
            WorkbookVersion.version_number == version_number,
        )
    ).one_or_none()
    if version is None:
        raise WorkbookServiceError(
            "STORAGE_OBJECT_MISSING", 500, "Workbook version is unavailable."
        )
    return version


@contextmanager
def _materialized_object(
    storage: WorkbookStorage,
    *,
    key: str,
    checksum: str,
    expected_size: int,
) -> Iterator[Path]:
    local_read_path = getattr(storage, "local_read_path", None)
    if callable(local_read_path):
        try:
            local_path = Path(local_read_path(key=key))
            stat = local_path.stat()
            identity = (
                stat.st_dev,
                stat.st_ino,
                stat.st_size,
                stat.st_mtime_ns,
                checksum,
            )
            with _verified_local_objects_guard:
                verified = identity in _verified_local_objects
            if not verified:
                digest = hashlib.sha256()
                with local_path.open("rb") as source:
                    while chunk := source.read(_COPY_CHUNK_SIZE):
                        digest.update(chunk)
                if stat.st_size != expected_size or digest.hexdigest() != checksum:
                    raise WorkbookServiceError(
                        "STORAGE_OBJECT_MISSING",
                        500,
                        "Workbook file failed integrity checking.",
                    )
                with _verified_local_objects_guard:
                    if len(_verified_local_objects) >= 1024:
                        _verified_local_objects.clear()
                    _verified_local_objects.add(identity)
            elif stat.st_size != expected_size:
                raise WorkbookServiceError(
                    "STORAGE_OBJECT_MISSING",
                    500,
                    "Workbook file failed integrity checking.",
                )
            yield local_path
            return
        except WorkbookServiceError:
            raise
        except (OSError, ValueError) as exc:
            raise WorkbookServiceError(
                "STORAGE_OBJECT_MISSING", 500, "Workbook file is unavailable."
            ) from exc

    temporary_path = _materialize_verified_object(
        storage,
        key=key,
        checksum=checksum,
        expected_size=expected_size,
    )
    try:
        yield temporary_path
    finally:
        try:
            temporary_path.unlink(missing_ok=True)
        except OSError:
            pass


def _materialize_verified_object(
    storage: WorkbookStorage,
    *,
    key: str,
    checksum: str,
    expected_size: int,
) -> Path:
    temporary_path: Path | None = None
    try:
        digest = hashlib.sha256()
        size = 0
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temporary:
            temporary_path = Path(temporary.name)
            try:
                source = storage.open_read(key=key)
            except Exception as exc:
                raise WorkbookServiceError(
                    "STORAGE_OBJECT_MISSING", 500, "Workbook file is unavailable."
                ) from exc
            with source:
                while True:
                    chunk = source.read(_COPY_CHUNK_SIZE)
                    if not chunk:
                        break
                    if not isinstance(chunk, bytes):
                        raise WorkbookServiceError(
                            "STORAGE_OBJECT_MISSING", 500, "Workbook file is invalid."
                        )
                    temporary.write(chunk)
                    digest.update(chunk)
                    size += len(chunk)
        if size != expected_size or digest.hexdigest() != checksum:
            raise WorkbookServiceError(
                "STORAGE_OBJECT_MISSING", 500, "Workbook file failed integrity checking."
            )
        result = temporary_path
        temporary_path = None
        return result
    finally:
        if temporary_path is not None:
            try:
                temporary_path.unlink(missing_ok=True)
            except OSError:
                pass


@contextmanager
def _serialize_local_save(db: Session, session_id: uuid.UUID) -> Iterator[None]:
    if db.get_bind().dialect.name != "sqlite":
        yield
        return

    with _sqlite_save_locks_guard:
        session_lock = _sqlite_save_locks.setdefault(session_id, threading.RLock())
    with session_lock:
        yield


def upload_workbook(
    db: Session,
    storage: WorkbookStorage,
    *,
    actor: User,
    filename: str,
    mime_type: str,
    source: BinaryIO,
    max_upload_bytes: int,
    max_rows: int,
    max_columns: int,
) -> WorkbookUploadResult:
    """Validate a bounded upload, publish its original, and persist metadata."""

    original_filename = _safe_original_filename(filename)
    if mime_type not in _SUPPORTED_UPLOAD_MIME_TYPES:
        raise WorkbookServiceError(
            "UNSUPPORTED_FILE_TYPE", 415, "Only .xlsx and .xls workbooks are supported."
        )
    if max_upload_bytes < 1:
        raise ValueError("Maximum upload size must be positive.")

    temporary_path: Path | None = None
    normalized_path: Path | None = None
    try:
        size = 0
        source_suffix = Path(original_filename).suffix.casefold()
        with tempfile.NamedTemporaryFile(suffix=source_suffix, delete=False) as temporary:
            temporary_path = Path(temporary.name)
            while True:
                chunk = source.read(_COPY_CHUNK_SIZE)
                if not chunk:
                    break
                if not isinstance(chunk, bytes):
                    raise WorkbookServiceError(
                        "INVALID_XLSX", 422, "Workbook upload must contain binary data."
                    )
                size += len(chunk)
                if size > max_upload_bytes:
                    raise WorkbookServiceError(
                        "FILE_TOO_LARGE", 413, "Workbook exceeds the upload size limit."
                    )
                temporary.write(chunk)

        try:
            normalized_path = temporary_path
            if source_suffix == ".xls":
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as converted:
                    normalized_path = Path(converted.name)
                convert_xls_to_xlsx(
                    temporary_path,
                    normalized_path,
                    max_rows=max_rows,
                    max_columns=max_columns,
                )
            inspection = validate_and_inspect_workbook(
                normalized_path,
                max_rows=max_rows,
                max_columns=max_columns,
            )
        except WorkbookValidationError as exc:
            raise _translate_domain_error(exc) from exc

        workbook_id = uuid.uuid4()
        key = f"originals/{workbook_id}/source.xlsx"
        try:
            with normalized_path.open("rb") as validated_source:
                stored = storage.put_immutable(key=key, source=validated_source)
        except Exception as exc:
            raise WorkbookServiceError(
                "STORAGE_WRITE_FAILED", 500, "Workbook original could not be stored."
            ) from exc

        workbook = Workbook(
            id=workbook_id,
            original_filename=original_filename,
            original_relative_path=stored.key,
            original_checksum=stored.checksum,
            mime_type=XLSX_MIME_TYPE,
            file_size=stored.size,
            sheet_count=inspection.sheet_count,
            sheet_metadata=[_sheet_metadata(sheet) for sheet in inspection.sheets],
            created_by=_id(actor.id),
        )
        db.add(workbook)
        db.commit()
        return WorkbookUploadResult(
            id=workbook_id,
            original_filename=original_filename,
            mime_type=XLSX_MIME_TYPE,
            file_size=stored.size,
            checksum=stored.checksum,
            sheet_count=inspection.sheet_count,
            sheets=inspection.sheets,
            created_at=_utc_datetime(workbook.created_at),
        )
    except WorkbookServiceError:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        raise WorkbookServiceError(
            "STORAGE_WRITE_FAILED", 500, "Workbook upload could not be recorded."
        ) from exc
    finally:
        for path in {temporary_path, normalized_path}:
            if path is not None:
                try:
                    path.unlink(missing_ok=True)
                except OSError:
                    pass


def _editor_primary_column_end(
    path: Path,
    worksheet: Any,
    header_structure: dict[int, tuple[str, str | None, int]],
    *,
    legacy_source: bool,
    max_column: int | None = None,
) -> int:
    """Keep a legacy main header band visible while retaining its hidden tail."""

    max_column = max_column or int(worksheet.max_column or 0)
    if not legacy_source or max_column < 1:
        return max_column

    grouped_columns = [
        column_number
        for column_number, (_label, group_label, _row_span) in header_structure.items()
        if group_label
    ]
    if not grouped_columns:
        return max_column

    group_end = max(grouped_columns)
    trailing_columns = set(range(group_end + 1, max_column + 1))
    if not trailing_columns:
        return max_column

    hidden_trailing = trailing_columns & read_hidden_columns(path, worksheet)
    if len(hidden_trailing) * 2 < len(trailing_columns):
        return max_column
    return group_end


def create_editing_session(
    db: Session,
    storage: WorkbookStorage,
    *,
    actor: User,
    workbook_id: uuid.UUID,
    sheet_name: str,
) -> EditingSessionDescriptor:
    """Create an editing branch using the automatically detected header."""

    try:
        workbook = _get_workbook(db, workbook_id, actor)
        sheet = _metadata_sheet(workbook, sheet_name)
        header_row = sheet.get("header_row_number")
        mapping = sheet.get("column_mapping", {})
        headers = sheet.get("detected_headers", [])
        if not isinstance(header_row, int) or not isinstance(mapping, dict):
            raise WorkbookServiceError(
                "INVALID_ROW", 422, "Worksheet header row is unavailable."
            )

        session_id = uuid.uuid4()
        version_id = uuid.uuid4()
        version_key = f"sessions/{session_id}/000001-{version_id}.xlsx"
        try:
            original_stream = storage.open_read(key=workbook.original_relative_path)
            with original_stream:
                stored = storage.put_immutable(key=version_key, source=original_stream)
        except Exception as exc:
            raise WorkbookServiceError(
                "STORAGE_OBJECT_MISSING", 500, "Workbook original is unavailable."
            ) from exc
        if (
            stored.checksum != workbook.original_checksum
            or stored.size != workbook.file_size
        ):
            raise WorkbookServiceError(
                "STORAGE_OBJECT_MISSING", 500, "Workbook original failed integrity checking."
            )

        semantic_by_column = {int(column): field for field, column in mapping.items()}
        inferred_types: dict[int, str] = {}
        with _materialized_object(storage, key=stored.key, checksum=stored.checksum, expected_size=stored.size) as session_path:
            source_workbook = load_workbook(session_path, read_only=True, data_only=False)
            try:
                source_sheet = source_workbook[sheet_name]
                meaningful_max_row = int(sheet.get("max_row", 0))
                meaningful_max_column = int(sheet.get("max_column", 0))
                header_end_row, header_structure = read_header_structure(
                    source_sheet,
                    header_row,
                    max_row=meaningful_max_row,
                    max_column=meaningful_max_column,
                )
                primary_column_end = _editor_primary_column_end(
                    session_path,
                    source_sheet,
                    header_structure,
                    legacy_source=Path(workbook.original_filename).suffix.casefold()
                    == ".xls",
                    max_column=meaningful_max_column,
                )
                inferred_types = {
                    column_number: "text"
                    for column_number in range(1, meaningful_max_column + 1)
                }
                unresolved_columns = set(inferred_types)
                for row in source_sheet.iter_rows(
                    min_row=header_end_row + 1,
                    max_row=meaningful_max_row,
                    min_col=1,
                    max_col=meaningful_max_column,
                ):
                    for column_number in tuple(unresolved_columns):
                        cell = row[column_number - 1]
                        if cell.value is None:
                            continue
                        if isinstance(cell.value, bool):
                            inferred_types[column_number] = "boolean"
                        elif isinstance(
                            cell.value, (date, datetime, time, timedelta)
                        ) or cell.is_date:
                            inferred_types[column_number] = "date"
                        elif isinstance(cell.value, (int, float)):
                            format_text = str(cell.number_format).casefold()
                            inferred_types[column_number] = (
                                "currency"
                                if any(
                                    token in format_text
                                    for token in (
                                        "₫",
                                        "vnd",
                                        "[$₫",
                                        "$",
                                        "€",
                                        "£",
                                        "¥",
                                    )
                                )
                                else "number"
                            )
                        unresolved_columns.remove(column_number)
                    if not unresolved_columns:
                        break
            finally:
                source_workbook.close()
        column_config = [
            {
                "id": f"source-{uuid.uuid5(uuid.NAMESPACE_URL, f'{workbook_id}:{sheet_name}:{column_number}')}",
                "label": (
                    str(headers[column_number - 1])
                    if column_number <= len(headers)
                    and headers[column_number - 1] is not None
                    else ""
                ),
                "column_number": column_number,
                "origin": "source",
                "data_type": "currency" if semantic_by_column.get(column_number) in {"net_price", "selling_price"} else inferred_types.get(column_number, "text"),
                "hidden": column_number > primary_column_end,
                "sticky": False,
                "semantic_field": semantic_by_column.get(column_number),
            }
            for column_number in range(1, int(sheet.get("max_column", len(headers))) + 1)
        ]
        editing_session = WorkbookSession(
            id=session_id,
            workbook_id=workbook_id,
            selected_sheet_name=sheet_name,
            header_row_number=header_row,
            column_mapping=dict(mapping),
            column_config=column_config,
            meaningful_max_row=int(sheet["max_row"]),
            meaningful_max_column=int(sheet["max_column"]),
            current_version=1,
            status=WorkbookSessionStatus.DRAFT,
            created_by=_id(actor.id),
        )
        version = WorkbookVersion(
            id=version_id,
            session_id=session_id,
            version_number=1,
            relative_path=stored.key,
            checksum=stored.checksum,
            file_size=stored.size,
            change_summary={"type": "SESSION_CREATED"},
            created_by=_id(actor.id),
        )
        db.add(editing_session)
        db.add(version)
        db.commit()
        return _describe_session(editing_session, workbook.original_filename)
    except WorkbookServiceError:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        raise WorkbookServiceError(
            "STORAGE_WRITE_FAILED", 500, "Editing session could not be created."
        ) from exc


def _normalized_column_config(
    editing_session: WorkbookSession,
) -> list[dict[str, Any]]:
    """Return API-safe metadata without changing source header text."""

    normalized: list[dict[str, Any]] = []
    for position, raw_item in enumerate(editing_session.column_config, start=1):
        item = dict(raw_item)
        column_number = item.get("column_number")
        if not isinstance(column_number, int) or column_number < 1:
            column_number = position
            item["column_number"] = column_number
        if item.get("label") is None:
            item["label"] = ""
        normalized.append(item)
    try:
        normalized, _order = normalize_column_formulas(normalized)
    except WorkbookFormulaError as exc:
        raise WorkbookServiceError(
            exc.code,
            _error_status(exc.code),
            str(exc),
            details=exc.details,
        ) from exc
    return normalized


def _describe_session(
    editing_session: WorkbookSession, original_filename: str
) -> EditingSessionDescriptor:
    return EditingSessionDescriptor(
        id=_id(editing_session.id),
        workbook_id=editing_session.workbook_id,
        original_filename=original_filename,
        selected_sheet_name=editing_session.selected_sheet_name,
        header_row_number=editing_session.header_row_number,
        column_mapping={
            field: int(column)
            for field, column in editing_session.column_mapping.items()
        },
        column_config=_normalized_column_config(editing_session),
        current_version=editing_session.current_version,
        status=editing_session.status,
        created_at=_utc_datetime(editing_session.created_at),
        updated_at=_utc_datetime(editing_session.updated_at),
    )


def _ensure_legacy_session_visibility(
    db: Session,
    storage: WorkbookStorage,
    editing_session: WorkbookSession,
    workbook: Workbook,
) -> None:
    """Upgrade untouched legacy column visibility without changing workbook data."""

    if Path(workbook.original_filename).suffix.casefold() != ".xls":
        return
    config = _normalized_column_config(editing_session)
    source_columns = [item for item in config if item.get("origin") == "source"]
    hidden_source_columns = [item for item in source_columns if item.get("hidden")]
    if any(
        str(item.get("label", "")).strip().casefold()
        not in {"no.", "no", "stt", "số tt", "số thứ tự"}
        for item in hidden_source_columns
    ):
        return

    version = _get_version(
        db,
        session_id=_id(editing_session.id),
        version_number=editing_session.current_version,
    )
    with _materialized_object(
        storage,
        key=version.relative_path,
        checksum=version.checksum,
        expected_size=version.file_size,
    ) as path:
        source_workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            source_sheet = source_workbook[editing_session.selected_sheet_name]
            meaningful_max_row, meaningful_max_column = _session_meaningful_bounds(
                editing_session, workbook
            )
            _header_end_row, header_structure = read_header_structure(
                source_sheet,
                editing_session.header_row_number,
                max_row=meaningful_max_row,
                max_column=meaningful_max_column,
            )
            primary_column_end = _editor_primary_column_end(
                path,
                source_sheet,
                header_structure,
                legacy_source=True,
                max_column=meaningful_max_column,
            )
            source_max_column = meaningful_max_column
            if primary_column_end == source_max_column:
                grouped_columns = [
                    column_number
                    for column_number, (
                        _label,
                        group_label,
                        _row_span,
                    ) in header_structure.items()
                    if group_label
                ]
                legacy_group_end = max(grouped_columns, default=source_max_column)
                # Versions created before XLS visibility preservation have no
                # hidden-column metadata. Only repair a large untouched tail.
                if source_max_column - legacy_group_end >= 8:
                    primary_column_end = legacy_group_end
        finally:
            source_workbook.close()
    if primary_column_end >= max(
        (int(item["column_number"]) for item in source_columns),
        default=0,
    ):
        return

    updated = [
        {
            **item,
            "hidden": (
                int(item["column_number"]) > primary_column_end
                if item.get("origin") == "source"
                else bool(item.get("hidden", False))
            ),
        }
        for item in config
    ]
    if updated == config:
        return
    editing_session.column_config = updated
    db.add(editing_session)
    db.commit()
    db.refresh(editing_session)


def _session_summary(
    editing_session: WorkbookSession,
    original_filename: str,
) -> SessionSummaryDescriptor:
    return SessionSummaryDescriptor(
        id=_id(editing_session.id),
        display_name=(editing_session.display_name or "").strip() or original_filename,
        original_filename=original_filename,
        selected_sheet_name=editing_session.selected_sheet_name,
        current_version=editing_session.current_version,
        status=editing_session.status,
        created_at=_utc_datetime(editing_session.created_at),
        updated_at=_utc_datetime(editing_session.updated_at),
    )


def get_editing_session(
    db: Session,
    *,
    actor: User,
    session_id: uuid.UUID,
    storage: WorkbookStorage | None = None,
) -> EditingSessionDescriptor:
    editing_session = _get_session(db, session_id, actor)
    workbook = db.get(Workbook, editing_session.workbook_id)
    if workbook is None:
        raise WorkbookServiceError(
            "WORKBOOK_NOT_FOUND", 404, "Workbook was not found."
        )
    if storage is not None:
        _ensure_legacy_session_visibility(db, storage, editing_session, workbook)
    return _describe_session(editing_session, workbook.original_filename)


def list_editing_sessions(
    db: Session,
    *,
    actor: User,
    page: int = 1,
    page_size: int = 10,
    search: str | None = None,
    session_status: WorkbookSessionStatus | None = None,
) -> SessionListResult:
    """List the authenticated user's sessions, newest activity first."""

    if page < 1 or page_size < 1 or page_size > 100:
        raise WorkbookServiceError(
            "INVALID_PAGINATION", 422, "Session pagination is invalid."
        )

    conditions = []
    if actor.role != UserRole.ADMIN:
        conditions.append(WorkbookSession.created_by == _id(actor.id))
    if session_status is None:
        conditions.append(WorkbookSession.status != WorkbookSessionStatus.DISCARDED)
    else:
        conditions.append(WorkbookSession.status == session_status)

    normalized_search = search.strip() if search else ""
    if normalized_search:
        pattern = f"%{normalized_search}%"
        conditions.append(
            or_(
                WorkbookSession.display_name.ilike(pattern),
                Workbook.original_filename.ilike(pattern),
                WorkbookSession.selected_sheet_name.ilike(pattern),
            )
        )

    total = db.exec(
        select(func.count())
        .select_from(WorkbookSession)
        .join(Workbook, WorkbookSession.workbook_id == Workbook.id)
        .where(*conditions)
    ).one()
    rows = db.exec(
        select(WorkbookSession, Workbook)
        .join(Workbook, WorkbookSession.workbook_id == Workbook.id)
        .where(*conditions)
        .order_by(
            WorkbookSession.updated_at.desc(),
            WorkbookSession.created_at.desc(),
        )
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()
    return SessionListResult(
        items=tuple(
            _session_summary(editing_session, workbook.original_filename)
            for editing_session, workbook in rows
        ),
        page=page,
        page_size=page_size,
        total=total,
        total_pages=math.ceil(total / page_size) if total else 0,
    )


def rename_editing_session(
    db: Session,
    *,
    actor: User,
    session_id: uuid.UUID,
    display_name: str,
) -> SessionSummaryDescriptor:
    """Rename an active session without changing workbook contents."""

    normalized_name = display_name.strip()
    if not normalized_name or len(normalized_name) > 255:
        raise WorkbookServiceError(
            "INVALID_SESSION_NAME", 422, "Session name is invalid."
        )
    try:
        editing_session = _get_session(db, session_id, actor)
        if editing_session.status != WorkbookSessionStatus.DRAFT:
            raise WorkbookServiceError(
                "SESSION_NOT_ACTIVE", 409, "Editing session is not active."
            )
        workbook = db.get(Workbook, editing_session.workbook_id)
        if workbook is None:
            raise WorkbookServiceError(
                "WORKBOOK_NOT_FOUND", 404, "Workbook was not found."
            )
        editing_session.display_name = normalized_name
        editing_session.updated_at = utc_now()
        db.add(editing_session)
        db.commit()
        return _session_summary(editing_session, workbook.original_filename)
    except WorkbookServiceError:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        raise WorkbookServiceError(
            "SESSION_UPDATE_FAILED", 500, "Session could not be renamed."
        ) from exc


def discard_editing_session(
    db: Session,
    *,
    actor: User,
    session_id: uuid.UUID,
) -> SessionSummaryDescriptor:
    """Soft-discard an active session while preserving its audit history."""

    try:
        editing_session = _get_session(db, session_id, actor)
        if editing_session.status != WorkbookSessionStatus.DRAFT:
            raise WorkbookServiceError(
                "SESSION_NOT_ACTIVE", 409, "Editing session is not active."
            )
        workbook = db.get(Workbook, editing_session.workbook_id)
        if workbook is None:
            raise WorkbookServiceError(
                "WORKBOOK_NOT_FOUND", 404, "Workbook was not found."
            )
        discarded_at = utc_now()
        editing_session.status = WorkbookSessionStatus.DISCARDED
        editing_session.discarded_at = discarded_at
        editing_session.updated_at = discarded_at
        db.add(editing_session)
        db.commit()
        return _session_summary(editing_session, workbook.original_filename)
    except WorkbookServiceError:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        raise WorkbookServiceError(
            "SESSION_UPDATE_FAILED", 500, "Session could not be discarded."
        ) from exc


def get_latest_editing_session(
    db: Session,
    *,
    actor: User,
) -> EditingSessionDescriptor:
    """Return the user's most recently active workbook editing session."""

    editing_session = db.exec(
        select(WorkbookSession)
        .where(
            WorkbookSession.created_by == _id(actor.id),
            WorkbookSession.status == WorkbookSessionStatus.DRAFT,
        )
        .order_by(WorkbookSession.updated_at.desc())
    ).first()
    if editing_session is None:
        raise WorkbookServiceError(
            "SESSION_NOT_FOUND", 404, "No active workbook session was found."
        )
    workbook = db.get(Workbook, editing_session.workbook_id)
    if workbook is None:
        raise WorkbookServiceError(
            "WORKBOOK_NOT_FOUND", 404, "Workbook was not found."
        )
    return _describe_session(editing_session, workbook.original_filename)


def read_session_records(
    db: Session,
    storage: WorkbookStorage,
    *,
    actor: User,
    session_id: uuid.UUID,
    page: int = 1,
    page_size: int = 50,
    max_page_size: int = 200,
    search: str | None = None,
    sort_by: str | None = None,
    sort_direction: str = "asc",
) -> SessionRecordsResult:
    editing_session = _get_session(db, session_id, actor)
    workbook = db.get(Workbook, editing_session.workbook_id)
    if workbook is None:
        raise WorkbookServiceError(
            "WORKBOOK_NOT_FOUND", 404, "Workbook was not found."
        )
    _ensure_legacy_session_visibility(db, storage, editing_session, workbook)
    meaningful_max_row, meaningful_max_column = _session_meaningful_bounds(
        editing_session, workbook
    )
    version = _get_version(
        db,
        session_id=_id(editing_session.id),
        version_number=editing_session.current_version,
    )
    try:
        with _materialized_object(
            storage,
            key=version.relative_path,
            checksum=version.checksum,
            expected_size=version.file_size,
        ) as path:
            page_result = read_workbook_records(
                path,
                sheet_name=editing_session.selected_sheet_name,
                header_row_number=editing_session.header_row_number,
                column_mapping={
                    field: int(column)
                    for field, column in editing_session.column_mapping.items()
                },
                column_config=_normalized_column_config(editing_session),
                page=page,
                page_size=page_size,
                max_page_size=max_page_size,
                search=search,
                sort_by=sort_by,
                sort_direction=sort_direction,
                meaningful_max_row=meaningful_max_row,
                meaningful_max_column=meaningful_max_column,
            )
            return SessionRecordsResult(
                session_id=session_id,
                version=version.version_number,
                sheet_name=editing_session.selected_sheet_name,
                page=page_result,
            )
    except WorkbookReadError as exc:
        raise _translate_domain_error(exc) from exc


def lookup_session_cell_values(
    db: Session,
    storage: WorkbookStorage,
    *,
    actor: User,
    session_id: uuid.UUID,
    base_version: int,
    cells: Sequence[WorkbookCellReference],
    max_cells: int = 500,
) -> SessionCellValuesResult:
    """Read current sparse cells for safe client-side draft reconciliation."""

    editing_session = _get_session(db, session_id, actor)
    if base_version != editing_session.current_version:
        raise WorkbookServiceError(
            "VERSION_CONFLICT",
            409,
            "Workbook session has a newer version.",
            details={"current_version": editing_session.current_version},
        )
    version = _get_version(
        db,
        session_id=_id(editing_session.id),
        version_number=base_version,
    )
    workbook = db.get(Workbook, editing_session.workbook_id)
    if workbook is None:
        raise WorkbookServiceError(
            "WORKBOOK_NOT_FOUND", 404, "Workbook was not found."
        )
    meaningful_max_row, meaningful_max_column = _session_meaningful_bounds(
        editing_session, workbook
    )
    try:
        with _materialized_object(
            storage,
            key=version.relative_path,
            checksum=version.checksum,
            expected_size=version.file_size,
        ) as path:
            values = read_workbook_cell_values(
                path,
                sheet_name=editing_session.selected_sheet_name,
                header_row_number=editing_session.header_row_number,
                column_config=_normalized_column_config(editing_session),
                cells=cells,
                max_cells=max_cells,
                meaningful_max_row=meaningful_max_row,
                meaningful_max_column=meaningful_max_column,
            )
        return SessionCellValuesResult(
            session_id=session_id,
            version=version.version_number,
            cells=values,
        )
    except WorkbookReadError as exc:
        raise _translate_domain_error(exc) from exc


def preview_session_formula(
    db: Session,
    storage: WorkbookStorage,
    *,
    actor: User,
    session_id: uuid.UUID,
    base_version: int,
    formula: dict[str, Any],
    output_type: str,
    output_column_id: str | None = None,
    sample_rows: Sequence[int] | None = None,
) -> FormulaPreviewResult:
    """Validate and evaluate a proposed formula without mutating the session."""

    editing_session = _get_session(db, session_id, actor)
    if base_version != editing_session.current_version:
        raise WorkbookServiceError(
            "VERSION_CONFLICT",
            409,
            "Workbook session has a newer version.",
            details={"current_version": editing_session.current_version},
        )
    config = _normalized_column_config(editing_session)
    workbook_record = db.get(Workbook, editing_session.workbook_id)
    if workbook_record is None:
        raise WorkbookServiceError(
            "WORKBOOK_NOT_FOUND", 404, "Workbook was not found."
        )
    meaningful_max_row, meaningful_max_column = _session_meaningful_bounds(
        editing_session, workbook_record
    )
    preview_column_id = output_column_id or "__formula_preview__"
    target = next(
        (item for item in config if item.get("id") == preview_column_id),
        None,
    )
    if output_column_id is not None and target is None:
        raise WorkbookServiceError("COLUMN_NOT_FOUND", 404, "Column was not found.")
    proposed_column = {
        **(target or {}),
        "id": preview_column_id,
        "label": (target or {}).get("label") or "Formula preview",
        "column_number": int((target or {}).get("column_number") or (max(int(item["column_number"]) for item in config) + 1)),
        "origin": "user",
        "data_type": output_type,
        "formula": formula,
    }
    proposed_config = [
        proposed_column if item.get("id") == preview_column_id else item
        for item in config
    ]
    if target is None:
        proposed_config.append(proposed_column)

    try:
        normalized_config, formula_order = normalize_column_formulas(proposed_config)
        normalized_proposed = next(
            item for item in normalized_config if item["id"] == preview_column_id
        )["formula"]
        expression = render_readable_expression(normalized_proposed, normalized_config)
        references = referenced_column_ids(normalized_proposed)
    except WorkbookFormulaError as exc:
        return FormulaPreviewResult(
            valid=False,
            normalized_formula=None,
            readable_expression=None,
            referenced_column_ids=(),
            results=(),
            errors=(
                {
                    "code": exc.code,
                    "message": str(exc),
                    "details": exc.details,
                },
            ),
        )

    version = _get_version(db, session_id=session_id, version_number=base_version)
    try:
        with _materialized_object(
            storage,
            key=version.relative_path,
            checksum=version.checksum,
            expected_size=version.file_size,
        ) as path:
            workbook = load_workbook(path, read_only=True, data_only=False)
            try:
                if editing_session.selected_sheet_name not in workbook.sheetnames:
                    raise WorkbookServiceError(
                        "SHEET_NOT_FOUND", 404, "Worksheet was not found."
                    )
                worksheet = workbook[editing_session.selected_sheet_name]
                header_end_row, _structure = read_header_structure(
                    worksheet,
                    editing_session.header_row_number,
                    max_row=meaningful_max_row,
                    max_column=meaningful_max_column,
                )
                requested_rows = list(sample_rows or [])
                if requested_rows:
                    if any(
                        isinstance(row_number, bool)
                        or row_number <= header_end_row
                        or row_number > meaningful_max_row
                        for row_number in requested_rows
                    ):
                        raise WorkbookServiceError(
                            "INVALID_ROW", 422, "A preview row is outside the workbook data area."
                        )
                else:
                    requested_rows = [
                        row_number
                        for row_number in range(
                            header_end_row + 1, meaningful_max_row + 1
                        )
                        if any(
                            worksheet.cell(row_number, int(item["column_number"])).value
                            not in (None, "")
                            for item in config
                            if not item.get("formula")
                        )
                    ][:5]

                results: list[FormulaPreviewRow] = []
                for row_number in requested_rows:
                    values = {
                        str(item["id"]): worksheet.cell(
                            row_number, int(item["column_number"])
                        ).value
                        for item in config
                    }
                    values, evaluations = evaluate_normalized_formula_columns(
                        normalized_config, formula_order, values
                    )
                    evaluation = evaluations[preview_column_id]
                    results.append(
                        FormulaPreviewRow(
                            row_number=row_number,
                            value=evaluation.value,
                            error_code=evaluation.error_code,
                            error_message=evaluation.error_message,
                        )
                    )
            finally:
                workbook.close()
    except WorkbookServiceError:
        raise
    except (WorkbookReadError, OSError, ValueError, KeyError) as exc:
        raise _translate_domain_error(exc) from exc

    warnings: list[dict[str, Any]] = []
    if not results:
        warnings.append(
            {
                "code": "NO_PREVIEW_ROWS",
                "message": "No populated workbook rows are available for preview.",
                "details": {},
            }
        )
    row_errors = [result for result in results if result.error_code]
    if row_errors:
        warnings.append(
            {
                "code": "PREVIEW_ROW_ERRORS",
                "message": "Formula could not be evaluated for one or more preview rows.",
                "details": {"row_numbers": [result.row_number for result in row_errors]},
            }
        )
    return FormulaPreviewResult(
        valid=not row_errors,
        normalized_formula=normalized_proposed,
        readable_expression=expression,
        referenced_column_ids=references,
        results=tuple(results),
        warnings=tuple(warnings),
    )


def _json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return str(value)
        return int(value) if value.is_integer() else str(value)
    return value


def _canonical_save_payload(
    *, base_version: int, changes: Sequence[PriceChange]
) -> tuple[dict[str, Any], str]:
    canonical_changes = [
        {
            "row_number": change.row_number,
            "values": {
                field: _json_value(value)
                for field, value in (
                    ("net_price", change.net_price),
                    ("selling_price", change.selling_price),
                )
                if value is not None
            } | {field: _json_value(value) for field, value in (change.values or {}).items()},
        }
        for change in changes
    ]
    canonical_changes.sort(
        key=lambda item: (
            str(item["row_number"]),
            json.dumps(item["values"], sort_keys=True, ensure_ascii=False),
        )
    )
    payload = {"base_version": base_version, "changes": canonical_changes}
    try:
        encoded = json.dumps(
            payload,
            sort_keys=True,
            separators=(",", ":"),
            ensure_ascii=False,
            allow_nan=False,
        ).encode("utf-8")
    except (TypeError, ValueError) as exc:
        raise WorkbookServiceError(
            "INVALID_CELL_VALUE", 422, "Save payload contains an invalid value."
        ) from exc
    return payload, hashlib.sha256(encoded).hexdigest()


def _find_operation(
    db: Session, *, session_id: uuid.UUID, request_id: uuid.UUID
) -> WorkbookOperation | None:
    return db.exec(
        select(WorkbookOperation).where(
            WorkbookOperation.session_id == session_id,
            WorkbookOperation.request_id == request_id,
        )
    ).one_or_none()


def _replay_result(
    db: Session,
    *,
    operation: WorkbookOperation,
    payload_checksum: str,
) -> WorkbookSaveResult:
    if operation.payload_checksum != payload_checksum:
        raise WorkbookServiceError(
            "IDEMPOTENCY_KEY_REUSED",
            409,
            "Request ID was already used with a different save payload.",
        )
    version = _get_version(
        db,
        session_id=operation.session_id,
        version_number=operation.to_version,
    )
    return WorkbookSaveResult(
        operation_id=_id(operation.id),
        request_id=operation.request_id,
        previous_version=operation.from_version,
        current_version=operation.to_version,
        changed_cells=operation.changed_cells,
        saved_at=_utc_datetime(operation.created_at),
        checksum=version.checksum,
        file_size=version.file_size,
        replayed=True,
    )


def _save_session_changes_locked(
    db: Session,
    storage: WorkbookStorage,
    *,
    actor: User,
    session_id: uuid.UUID,
    request_id: uuid.UUID,
    base_version: int,
    changes: Sequence[PriceChange],
    max_changes: int = 500,
) -> WorkbookSaveResult:
    """Save approved price changes as one immutable session version."""

    payload, payload_checksum = _canonical_save_payload(
        base_version=base_version,
        changes=changes,
    )
    try:
        editing_session = db.exec(
            select(WorkbookSession)
            .where(WorkbookSession.id == session_id)
            .with_for_update()
        ).one_or_none()
        if editing_session is None:
            raise WorkbookServiceError("SESSION_NOT_FOUND", 404, "Session was not found.")
        _authorize_owner(actor=actor, owner_id=editing_session.created_by)

        existing = _find_operation(
            db, session_id=session_id, request_id=request_id
        )
        if existing is not None:
            result = _replay_result(
                db, operation=existing, payload_checksum=payload_checksum
            )
            db.rollback()
            return result
        if editing_session.status != WorkbookSessionStatus.DRAFT:
            raise WorkbookServiceError(
                "SESSION_NOT_ACTIVE", 409, "Editing session is not active."
            )
        if base_version != editing_session.current_version:
            raise WorkbookServiceError(
                "VERSION_CONFLICT",
                409,
                "Workbook session has a newer version.",
                details={"current_version": editing_session.current_version},
            )

        current = _get_version(
            db, session_id=session_id, version_number=editing_session.current_version
        )
        workbook_record = db.get(Workbook, editing_session.workbook_id)
        if workbook_record is None:
            raise WorkbookServiceError(
                "WORKBOOK_NOT_FOUND", 404, "Workbook was not found."
            )
        meaningful_max_row, meaningful_max_column = _session_meaningful_bounds(
            editing_session, workbook_record
        )
        next_number = editing_session.current_version + 1
        next_version_id = uuid.uuid4()
        next_key = (
            f"sessions/{session_id}/{next_number:06d}-{next_version_id}.xlsx"
        )
        with _materialized_object(
            storage,
            key=current.relative_path,
            checksum=current.checksum,
            expected_size=current.file_size,
        ) as source_path, tempfile.TemporaryDirectory() as directory:
            output_path = Path(directory) / "output.xlsx"
            try:
                mutation = apply_price_changes(
                    source_path,
                    output_path,
                    sheet_name=editing_session.selected_sheet_name,
                    header_row_number=editing_session.header_row_number,
                    column_mapping={
                        field: int(column)
                        for field, column in editing_session.column_mapping.items()
                    },
                    changes=changes,
                    max_changes=max_changes,
                    column_config=_normalized_column_config(editing_session),
                    meaningful_max_row=meaningful_max_row,
                    meaningful_max_column=meaningful_max_column,
                )
            except WorkbookMutationError as exc:
                raise _translate_domain_error(exc) from exc
            try:
                with output_path.open("rb") as generated:
                    stored = storage.put_immutable(key=next_key, source=generated)
            except Exception as exc:
                raise WorkbookServiceError(
                    "STORAGE_WRITE_FAILED", 500, "Workbook version could not be stored."
                ) from exc

        audited_changes = [
            {key: _json_value(value) for key, value in asdict(change).items()}
            for change in mutation.changes
        ]
        version = WorkbookVersion(
            id=next_version_id,
            session_id=session_id,
            version_number=next_number,
            relative_path=stored.key,
            checksum=stored.checksum,
            file_size=stored.size,
            change_summary={"changed_cells": mutation.changed_cell_count},
            created_by=_id(actor.id),
        )
        operation_id = uuid.uuid4()
        saved_at = utc_now()
        operation = WorkbookOperation(
            id=operation_id,
            session_id=session_id,
            from_version=base_version,
            to_version=next_number,
            request_id=request_id,
            operation_type=WorkbookOperationType.UPDATE_PRICES,
            operation_payload={"request": payload, "changes": audited_changes},
            payload_checksum=payload_checksum,
            changed_cells=mutation.changed_cell_count,
            created_by=_id(actor.id),
            created_at=_utc_datetime(saved_at),
        )
        editing_session.current_version = next_number
        editing_session.updated_at = utc_now()
        db.add(version)
        db.add(operation)
        db.add(editing_session)
        db.commit()
        return WorkbookSaveResult(
            operation_id=operation_id,
            request_id=request_id,
            previous_version=base_version,
            current_version=next_number,
            changed_cells=mutation.changed_cell_count,
            saved_at=_utc_datetime(saved_at),
            checksum=stored.checksum,
            file_size=stored.size,
        )
    except WorkbookServiceError:
        db.rollback()
        raise
    except IntegrityError as exc:
        db.rollback()
        raise WorkbookServiceError(
            "VERSION_CONFLICT", 409, "Workbook session was updated concurrently."
        ) from exc
    except Exception as exc:
        db.rollback()
        raise WorkbookServiceError(
            "STORAGE_WRITE_FAILED", 500, "Workbook changes could not be saved."
        ) from exc


def save_session_changes(
    db: Session,
    storage: WorkbookStorage,
    *,
    actor: User,
    session_id: uuid.UUID,
    request_id: uuid.UUID,
    base_version: int,
    changes: Sequence[PriceChange],
    max_changes: int = 500,
) -> WorkbookSaveResult:
    """Serialize local saves and persist one immutable workbook version."""

    with _serialize_local_save(db, session_id):
        return _save_session_changes_locked(
            db,
            storage,
            actor=actor,
            session_id=session_id,
            request_id=request_id,
            base_version=base_version,
            changes=changes,
            max_changes=max_changes,
        )


def add_session_column(
    db: Session,
    storage: WorkbookStorage,
    *,
    actor: User,
    session_id: uuid.UUID,
    base_version: int,
    label: str,
    data_type: str,
    formula: dict[str, Any] | None = None,
) -> EditingSessionDescriptor:
    """Append an arbitrary user-owned column as an immutable version."""

    normalized_label = label.strip()
    if not normalized_label:
        raise WorkbookServiceError(
            "INVALID_COLUMN", 422, "Column label cannot be blank."
        )
    with _serialize_local_save(db, session_id):
        try:
            editing_session = db.exec(
                select(WorkbookSession)
                .where(WorkbookSession.id == session_id)
                .with_for_update()
            ).one_or_none()
            if editing_session is None:
                raise WorkbookServiceError(
                    "SESSION_NOT_FOUND", 404, "Session was not found."
                )
            _authorize_owner(actor=actor, owner_id=editing_session.created_by)
            if editing_session.status != WorkbookSessionStatus.DRAFT:
                raise WorkbookServiceError(
                    "SESSION_NOT_ACTIVE", 409, "Editing session is not active."
                )
            if base_version != editing_session.current_version:
                raise WorkbookServiceError(
                    "VERSION_CONFLICT",
                    409,
                    "Workbook session has a newer version.",
                    details={"current_version": editing_session.current_version},
                )
            current = _get_version(
                db, session_id=session_id, version_number=base_version
            )
            workbook_record = db.get(Workbook, editing_session.workbook_id)
            if workbook_record is None:
                raise WorkbookServiceError(
                    "WORKBOOK_NOT_FOUND", 404, "Workbook was not found."
                )
            meaningful_max_row, meaningful_max_column = _session_meaningful_bounds(
                editing_session, workbook_record
            )
            config = _normalized_column_config(editing_session)
            column_id = f"user-{uuid.uuid4()}"
            new_column = {
                "id": column_id,
                "label": normalized_label,
                "column_number": max(int(item["column_number"]) for item in config) + 1,
                "origin": "user",
                "data_type": data_type,
                "hidden": False,
                "sticky": False,
                "semantic_field": None,
                "formula": formula,
            }
            proposed_config = config + [new_column]
            try:
                proposed_config, _formula_order = normalize_column_formulas(
                    proposed_config
                )
            except WorkbookFormulaError as exc:
                raise WorkbookServiceError(
                    exc.code,
                    _error_status(exc.code),
                    str(exc),
                    details=exc.details,
                ) from exc
            next_number = base_version + 1
            version_id = uuid.uuid4()
            key = f"sessions/{session_id}/{next_number:06d}-{version_id}.xlsx"
            with _materialized_object(
                storage,
                key=current.relative_path,
                checksum=current.checksum,
                expected_size=current.file_size,
            ) as source_path, tempfile.TemporaryDirectory() as directory:
                output_path = Path(directory) / "output.xlsx"
                mutation = add_workbook_column(
                    source_path,
                    output_path,
                    sheet_name=editing_session.selected_sheet_name,
                    header_row_number=editing_session.header_row_number,
                    label=normalized_label,
                    column_config=proposed_config,
                    meaningful_max_row=meaningful_max_row,
                    meaningful_max_column=meaningful_max_column,
                )
                with output_path.open("rb") as generated:
                    stored = storage.put_immutable(key=key, source=generated)
            if mutation.column_number != int(new_column["column_number"]):
                raise WorkbookServiceError(
                    "STORAGE_WRITE_FAILED", 500, "Workbook column position is invalid."
                )
            editing_session.column_config = proposed_config
            editing_session.meaningful_max_column = meaningful_max_column + 1
            editing_session.current_version = next_number
            editing_session.updated_at = utc_now()
            db.add(
                WorkbookVersion(
                    id=version_id,
                    session_id=session_id,
                    version_number=next_number,
                    relative_path=stored.key,
                    checksum=stored.checksum,
                    file_size=stored.size,
                    change_summary={
                        "type": "ADD_COLUMN",
                        "column_id": column_id,
                        "label": normalized_label,
                    },
                    created_by=_id(actor.id),
                )
            )
            db.add(editing_session)
            db.commit()
            workbook = db.get(Workbook, editing_session.workbook_id)
            if workbook is None:
                raise WorkbookServiceError("WORKBOOK_NOT_FOUND", 404, "Workbook was not found.")
            return _describe_session(editing_session, workbook.original_filename)
        except WorkbookServiceError:
            db.rollback()
            raise
        except WorkbookMutationError as exc:
            db.rollback()
            raise _translate_domain_error(exc) from exc
        except Exception as exc:
            db.rollback()
            raise WorkbookServiceError(
                "STORAGE_WRITE_FAILED", 500, "Workbook column could not be added."
            ) from exc


def update_session_column(
    db: Session,
    storage: WorkbookStorage,
    *,
    actor: User,
    session_id: uuid.UUID,
    column_id: str,
    base_version: int,
    label: str | None = None,
    data_type: str | None = None,
    formula: dict[str, Any] | None = None,
    formula_was_provided: bool = False,
) -> EditingSessionDescriptor:
    """Rename or edit a user formula column as one immutable version."""

    with _serialize_local_save(db, session_id):
        try:
            editing_session = db.exec(
                select(WorkbookSession)
                .where(WorkbookSession.id == session_id)
                .with_for_update()
            ).one_or_none()
            if editing_session is None:
                raise WorkbookServiceError("SESSION_NOT_FOUND", 404, "Session was not found.")
            _authorize_owner(actor=actor, owner_id=editing_session.created_by)
            if editing_session.status != WorkbookSessionStatus.DRAFT:
                raise WorkbookServiceError(
                    "SESSION_NOT_ACTIVE", 409, "Editing session is not active."
                )
            if base_version != editing_session.current_version:
                raise WorkbookServiceError(
                    "VERSION_CONFLICT",
                    409,
                    "Workbook session has a newer version.",
                    details={"current_version": editing_session.current_version},
                )
            config = _normalized_column_config(editing_session)
            workbook_record = db.get(Workbook, editing_session.workbook_id)
            if workbook_record is None:
                raise WorkbookServiceError(
                    "WORKBOOK_NOT_FOUND", 404, "Workbook was not found."
                )
            meaningful_max_row, meaningful_max_column = _session_meaningful_bounds(
                editing_session, workbook_record
            )
            target = next((item for item in config if item.get("id") == column_id), None)
            if target is None:
                raise WorkbookServiceError("COLUMN_NOT_FOUND", 404, "Column was not found.")
            if target.get("origin") != "user":
                raise WorkbookServiceError(
                    "SOURCE_COLUMN_IMMUTABLE",
                    422,
                    "Columns from the uploaded workbook cannot be renamed or assigned formulas.",
                )

            next_label = label.strip() if label is not None else str(target.get("label") or "")
            if not next_label:
                raise WorkbookServiceError(
                    "INVALID_COLUMN", 422, "Column label cannot be blank."
                )
            next_type = data_type or str(target.get("data_type", "text"))
            next_formula = formula if formula_was_provided else target.get("formula")
            if next_formula is not None and next_type not in {"number", "currency"}:
                raise WorkbookServiceError(
                    "INVALID_FORMULA_OUTPUT_TYPE",
                    422,
                    "Formula output must be number or currency.",
                )
            updated_target = {
                **target,
                "label": next_label,
                "data_type": next_type,
                "formula": next_formula,
            }
            proposed_config = [
                updated_target if item.get("id") == column_id else item
                for item in config
            ]
            try:
                proposed_config, _formula_order = normalize_column_formulas(
                    proposed_config
                )
            except WorkbookFormulaError as exc:
                raise WorkbookServiceError(
                    exc.code,
                    _error_status(exc.code),
                    str(exc),
                    details=exc.details,
                ) from exc
            if proposed_config == config:
                raise WorkbookServiceError(
                    "NO_CHANGES", 422, "Column update does not change the workbook."
                )

            current = _get_version(db, session_id=session_id, version_number=base_version)
            next_number, version_id = base_version + 1, uuid.uuid4()
            key = f"sessions/{session_id}/{next_number:06d}-{version_id}.xlsx"
            with _materialized_object(
                storage,
                key=current.relative_path,
                checksum=current.checksum,
                expected_size=current.file_size,
            ) as source_path, tempfile.TemporaryDirectory() as directory:
                output_path = Path(directory) / "output.xlsx"
                if data_type is not None and next_formula is None:
                    validate_workbook_column_values(
                        source_path,
                        sheet_name=editing_session.selected_sheet_name,
                        header_row_number=editing_session.header_row_number,
                        column_number=int(target["column_number"]),
                        data_type=next_type,
                        semantic_field=target.get("semantic_field"),
                        meaningful_max_row=meaningful_max_row,
                        meaningful_max_column=meaningful_max_column,
                    )
                update_workbook_column(
                    source_path,
                    output_path,
                    sheet_name=editing_session.selected_sheet_name,
                    header_row_number=editing_session.header_row_number,
                    column_number=int(target["column_number"]),
                    label=next_label,
                    column_config=proposed_config,
                    clear_column_values=bool(target.get("formula")) and next_formula is None,
                    meaningful_max_row=meaningful_max_row,
                    meaningful_max_column=meaningful_max_column,
                )
                with output_path.open("rb") as generated:
                    stored = storage.put_immutable(key=key, source=generated)

            editing_session.column_config = proposed_config
            editing_session.current_version = next_number
            editing_session.updated_at = utc_now()
            db.add(
                WorkbookVersion(
                    id=version_id,
                    session_id=session_id,
                    version_number=next_number,
                    relative_path=stored.key,
                    checksum=stored.checksum,
                    file_size=stored.size,
                    change_summary={
                        "type": "UPDATE_COLUMN",
                        "column_id": column_id,
                        "formula_updated": formula_was_provided,
                    },
                    created_by=_id(actor.id),
                )
            )
            db.add(editing_session)
            db.commit()
            workbook = db.get(Workbook, editing_session.workbook_id)
            if workbook is None:
                raise WorkbookServiceError("WORKBOOK_NOT_FOUND", 404, "Workbook was not found.")
            return _describe_session(editing_session, workbook.original_filename)
        except WorkbookServiceError:
            db.rollback()
            raise
        except WorkbookMutationError as exc:
            db.rollback()
            raise _translate_domain_error(exc) from exc
        except Exception as exc:
            db.rollback()
            raise WorkbookServiceError(
                "STORAGE_WRITE_FAILED", 500, "Workbook column could not be updated."
            ) from exc


def remove_session_column(
    db: Session,
    storage: WorkbookStorage,
    *,
    actor: User,
    session_id: uuid.UUID,
    column_id: str,
    base_version: int,
) -> EditingSessionDescriptor:
    """Remove a user column; source columns are permanent."""
    with _serialize_local_save(db, session_id):
        try:
            editing_session = db.exec(
                select(WorkbookSession)
                .where(WorkbookSession.id == session_id)
                .with_for_update()
            ).one_or_none()
            if editing_session is None:
                raise WorkbookServiceError(
                    "SESSION_NOT_FOUND", 404, "Session was not found."
                )
            _authorize_owner(actor=actor, owner_id=editing_session.created_by)
            if editing_session.status != WorkbookSessionStatus.DRAFT:
                raise WorkbookServiceError(
                    "SESSION_NOT_ACTIVE", 409, "Editing session is not active."
                )
            if base_version != editing_session.current_version:
                raise WorkbookServiceError(
                    "VERSION_CONFLICT",
                    409,
                    "Workbook session has a newer version.",
                    details={"current_version": editing_session.current_version},
                )
            config = _normalized_column_config(editing_session)
            workbook_record = db.get(Workbook, editing_session.workbook_id)
            if workbook_record is None:
                raise WorkbookServiceError(
                    "WORKBOOK_NOT_FOUND", 404, "Workbook was not found."
                )
            meaningful_max_row, meaningful_max_column = _session_meaningful_bounds(
                editing_session, workbook_record
            )
            target = next((item for item in config if item.get("id") == column_id), None)
            if target is None:
                raise WorkbookServiceError(
                    "COLUMN_NOT_FOUND", 404, "Column was not found."
                )
            if target.get("origin") != "user":
                raise WorkbookServiceError(
                    "SOURCE_COLUMN_IMMUTABLE",
                    422,
                    "Columns from the uploaded workbook cannot be removed.",
                )
            dependents = dependent_formula_columns(config, column_id)
            if dependents:
                raise WorkbookServiceError(
                    "COLUMN_IN_USE",
                    422,
                    "Column is referenced by a formula and cannot be removed.",
                    details={"dependent_column_ids": list(dependents)},
                )
            column_number = int(target["column_number"])
            next_config = [
                {
                    **item,
                    "column_number": int(item["column_number"])
                    - (1 if int(item["column_number"]) > column_number else 0),
                }
                for item in config
                if item.get("id") != column_id
            ]
            try:
                next_config, _formula_order = normalize_column_formulas(next_config)
            except WorkbookFormulaError as exc:
                raise WorkbookServiceError(
                    exc.code,
                    _error_status(exc.code),
                    str(exc),
                    details=exc.details,
                ) from exc
            current = _get_version(
                db, session_id=session_id, version_number=base_version
            )
            next_number, version_id = base_version + 1, uuid.uuid4()
            key = f"sessions/{session_id}/{next_number:06d}-{version_id}.xlsx"
            with _materialized_object(
                storage,
                key=current.relative_path,
                checksum=current.checksum,
                expected_size=current.file_size,
            ) as source_path, tempfile.TemporaryDirectory() as directory:
                output_path = Path(directory) / "output.xlsx"
                remove_workbook_column(
                    source_path,
                    output_path,
                    sheet_name=editing_session.selected_sheet_name,
                    column_number=column_number,
                    header_row_number=editing_session.header_row_number,
                    column_config=next_config,
                    meaningful_max_row=meaningful_max_row,
                    meaningful_max_column=meaningful_max_column,
                )
                with output_path.open("rb") as generated:
                    stored = storage.put_immutable(key=key, source=generated)
            editing_session.column_config = next_config
            editing_session.meaningful_max_column = meaningful_max_column - 1
            editing_session.column_mapping = {
                field: int(number) - (1 if int(number) > column_number else 0)
                for field, number in editing_session.column_mapping.items()
                if int(number) != column_number
            }
            editing_session.current_version = next_number
            editing_session.updated_at = utc_now()
            db.add(
                WorkbookVersion(
                    id=version_id,
                    session_id=session_id,
                    version_number=next_number,
                    relative_path=stored.key,
                    checksum=stored.checksum,
                    file_size=stored.size,
                    change_summary={
                        "type": "REMOVE_COLUMN",
                        "column_id": column_id,
                    },
                    created_by=_id(actor.id),
                )
            )
            db.add(editing_session)
            db.commit()
            workbook = db.get(Workbook, editing_session.workbook_id)
            if workbook is None:
                raise WorkbookServiceError("WORKBOOK_NOT_FOUND", 404, "Workbook was not found.")
            return _describe_session(editing_session, workbook.original_filename)
        except WorkbookServiceError:
            db.rollback()
            raise
        except Exception as exc:
            db.rollback()
            raise WorkbookServiceError(
                "STORAGE_WRITE_FAILED", 500, "Workbook column could not be removed."
            ) from exc


def update_session_column_configuration(
    db: Session,
    *,
    actor: User,
    session_id: uuid.UUID,
    base_version: int,
    hidden_column_ids: Sequence[str],
    sticky_column_ids: Sequence[str],
) -> EditingSessionDescriptor:
    """Persist display-only preferences without creating a workbook version."""
    with _serialize_local_save(db, session_id):
        try:
            editing_session = db.exec(
                select(WorkbookSession)
                .where(WorkbookSession.id == session_id)
                .with_for_update()
            ).one_or_none()
            if editing_session is None:
                raise WorkbookServiceError(
                    "SESSION_NOT_FOUND", 404, "Session was not found."
                )
            _authorize_owner(actor=actor, owner_id=editing_session.created_by)
            if editing_session.status != WorkbookSessionStatus.DRAFT:
                raise WorkbookServiceError(
                    "SESSION_NOT_ACTIVE", 409, "Editing session is not active."
                )
            if base_version != editing_session.current_version:
                raise WorkbookServiceError(
                    "VERSION_CONFLICT",
                    409,
                    "Workbook session has a newer version.",
                    details={"current_version": editing_session.current_version},
                )
            config = _normalized_column_config(editing_session)
            known = {str(item.get("id")) for item in config}
            requested = set(hidden_column_ids) | set(sticky_column_ids)
            if requested - known:
                raise WorkbookServiceError(
                    "COLUMN_NOT_FOUND",
                    422,
                    "Column configuration contains an unknown column ID.",
                    details={"column_ids": sorted(requested - known)},
                )
            hidden, sticky = set(hidden_column_ids), set(sticky_column_ids)
            editing_session.column_config = [
                {
                    **item,
                    "hidden": str(item.get("id")) in hidden,
                    "sticky": str(item.get("id")) in sticky,
                }
                for item in config
            ]
            editing_session.updated_at = utc_now()
            db.add(editing_session)
            db.commit()
            workbook = db.get(Workbook, editing_session.workbook_id)
            if workbook is None:
                raise WorkbookServiceError(
                    "WORKBOOK_NOT_FOUND", 404, "Workbook was not found."
                )
            return _describe_session(editing_session, workbook.original_filename)
        except WorkbookServiceError:
            db.rollback()
            raise
        except Exception as exc:
            db.rollback()
            raise WorkbookServiceError(
                "SESSION_UPDATE_FAILED",
                500,
                "Column configuration could not be updated.",
            ) from exc


def get_current_download(
    db: Session,
    storage: WorkbookStorage,
    *,
    actor: User,
    session_id: uuid.UUID,
) -> WorkbookDownloadDescriptor:
    editing_session = _get_session(db, session_id, actor)
    workbook = db.get(Workbook, editing_session.workbook_id)
    if workbook is None:
        raise WorkbookServiceError(
            "WORKBOOK_NOT_FOUND", 404, "Workbook was not found."
        )
    version = _get_version(
        db, session_id=session_id, version_number=editing_session.current_version
    )
    temporary_path = _materialize_verified_object(
        storage,
        key=version.relative_path,
        checksum=version.checksum,
        expected_size=version.file_size,
    )
    try:
        stream = _DeletingBinaryStream(temporary_path)
    except OSError as exc:
        try:
            temporary_path.unlink(missing_ok=True)
        except OSError:
            pass
        raise WorkbookServiceError(
            "STORAGE_OBJECT_MISSING", 500, "Workbook file is unavailable."
        ) from exc

    stem = Path(workbook.original_filename).stem
    safe_stem = re.sub(r"[^\w.-]+", "-", stem, flags=re.UNICODE).strip("-._")
    safe_stem = safe_stem[:180] or "workbook"
    return WorkbookDownloadDescriptor(
        stream=stream,
        filename=f"{safe_stem}-edited-v{version.version_number}.xlsx",
        mime_type=XLSX_MIME_TYPE,
        checksum=version.checksum,
        file_size=version.file_size,
        version=version.version_number,
    )

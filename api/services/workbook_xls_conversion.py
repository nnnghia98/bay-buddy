"""Safe normalization of legacy BIFF ``.xls`` workbooks for the editor."""

from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError

from services.workbook_validation import WorkbookValidationError


def convert_xls_to_xlsx(
    source_path: str | Path,
    target_path: str | Path,
    *,
    max_rows: int,
    max_columns: int,
) -> None:
    """Convert a legacy XLS file into a macro-free XLSX editing source.

    XLS formulas are intentionally imported as their last calculated values.
    The workbook editor does not support editing formulas, and writing values
    avoids carrying legacy macro or formula payloads into the generated XLSX.
    """

    source = Path(source_path)
    target = Path(target_path)
    if source.suffix.casefold() != ".xls":
        raise ValueError("Legacy workbook source must use the .xls extension.")
    if target.suffix.casefold() != ".xlsx":
        raise ValueError("Converted workbook target must use the .xlsx extension.")
    if max_rows <= 0 or max_columns <= 0:
        raise ValueError("Workbook row and column limits must be positive.")

    try:
        legacy_workbook = xlrd.open_workbook(
            filename=str(source),
            formatting_info=True,
            on_demand=True,
        )
    except (OSError, ValueError, xlrd.XLRDError) as exc:
        raise WorkbookValidationError(
            "INVALID_XLS",
            "Legacy Excel workbook is corrupt or unreadable.",
            details={"reason": "UNREADABLE_LEGACY_WORKBOOK"},
        ) from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    try:
        for sheet_index in range(legacy_workbook.nsheets):
            legacy_sheet = legacy_workbook.sheet_by_index(sheet_index)
            effective_rows, effective_columns = _meaningful_bounds(legacy_sheet)
            if effective_rows > max_rows or effective_columns > max_columns:
                raise WorkbookValidationError(
                    "WORKBOOK_LIMIT_EXCEEDED",
                    "Workbook exceeds the supported worksheet size.",
                    details={
                        "sheet_name": legacy_sheet.name,
                        "rows": effective_rows,
                        "columns": effective_columns,
                        "max_rows": max_rows,
                        "max_columns": max_columns,
                    },
                )
            worksheet = workbook.create_sheet(title=legacy_sheet.name)
            _copy_sheet(
                legacy_workbook,
                legacy_sheet,
                worksheet,
                row_count=effective_rows,
                column_count=effective_columns,
            )
            if getattr(legacy_sheet, "visibility", 0) != 0:
                worksheet.sheet_state = "hidden"

        workbook.save(target)
    except WorkbookValidationError:
        raise
    except (IllegalCharacterError, OSError, ValueError) as exc:
        raise WorkbookValidationError(
            "INVALID_XLS",
            "Legacy Excel workbook could not be converted safely.",
            details={"reason": "LEGACY_CONVERSION_FAILED"},
        ) from exc
    finally:
        workbook.close()
        legacy_workbook.release_resources()


def _copy_sheet(
    legacy_workbook: xlrd.book.Book,
    legacy_sheet: xlrd.sheet.Sheet,
    worksheet: Any,
    *,
    row_count: int,
    column_count: int,
) -> None:
    for row_index in range(row_count):
        for column_index in range(column_count):
            cell = legacy_sheet.cell(row_index, column_index)
            value = _cell_value(legacy_workbook, cell)
            number_format = _cell_number_format(legacy_workbook, cell)
            if value is None and number_format in {None, "General"}:
                continue
            target_cell = worksheet.cell(
                row=row_index + 1,
                column=column_index + 1,
                value=value,
            )
            if number_format:
                target_cell.number_format = number_format

    for column_index, column_info in legacy_sheet.colinfo_map.items():
        if column_index >= column_count:
            continue
        # XLS stores widths in units of 1/256 of a character; openpyxl uses
        # character widths. The cap prevents pathological column dimensions.
        width = max(0.5, min(column_info.width / 256, 255))
        dimension = worksheet.column_dimensions[get_column_letter(column_index + 1)]
        dimension.width = width
        dimension.hidden = bool(column_info.hidden)
        dimension.outline_level = min(
            max(int(getattr(column_info, "outline_level", 0)), 0),
            7,
        )
        dimension.collapsed = bool(getattr(column_info, "collapsed", False))

    for row_index, row_info in legacy_sheet.rowinfo_map.items():
        if row_index >= row_count:
            continue
        dimension = worksheet.row_dimensions[row_index + 1]
        dimension.hidden = bool(row_info.hidden)
        dimension.outline_level = min(
            max(int(getattr(row_info, "outline_level", 0)), 0),
            7,
        )
        dimension.collapsed = bool(getattr(row_info, "collapsed", False))
        if getattr(row_info, "has_default_height", False) and row_info.height > 0:
            dimension.height = row_info.height / 20

    for row_low, row_high, column_low, column_high in legacy_sheet.merged_cells:
        if row_low >= row_count or column_low >= column_count:
            continue
        worksheet.merge_cells(
            start_row=row_low + 1,
            end_row=min(row_high, row_count),
            start_column=column_low + 1,
            end_column=min(column_high, column_count),
        )


def _meaningful_bounds(legacy_sheet: xlrd.sheet.Sheet) -> tuple[int, int]:
    """Ignore formatting-only blank cells when enforcing worksheet limits."""

    max_row = 0
    max_column = 0
    meaningful_cells: set[tuple[int, int]] = set()
    for row_index in range(legacy_sheet.nrows):
        for column_index in range(legacy_sheet.ncols):
            cell = legacy_sheet.cell(row_index, column_index)
            if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
                continue
            meaningful_cells.add((row_index, column_index))
            max_row = max(max_row, row_index + 1)
            max_column = max(max_column, column_index + 1)

    for row_low, row_high, column_low, column_high in legacy_sheet.merged_cells:
        if (row_low, column_low) not in meaningful_cells:
            continue
        max_row = max(max_row, row_high)
        max_column = max(max_column, column_high)

    return max_row, max_column


def _cell_value(legacy_workbook: xlrd.book.Book, cell: xlrd.sheet.Cell) -> Any:
    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
        return None
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_DATE:
        value = xlrd.xldate_as_datetime(cell.value, legacy_workbook.datemode)
        return _normalize_excel_date(value)
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return xlrd.error_text_from_code.get(cell.value, "#VALUE!")
    return cell.value


def _cell_number_format(
    legacy_workbook: xlrd.book.Book,
    cell: xlrd.sheet.Cell,
) -> str | None:
    """Return the XLS display format without changing the stored cell value."""

    if cell.xf_index < 0 or cell.xf_index >= len(legacy_workbook.xf_list):
        return None
    cell_format = legacy_workbook.xf_list[cell.xf_index]
    number_format = legacy_workbook.format_map.get(cell_format.format_key)
    if number_format is None or not number_format.format_str:
        return None
    return str(number_format.format_str)


def _normalize_excel_date(value: datetime) -> date | datetime | time:
    if value.time() == time.min:
        return value.date()
    if value.date() == date(1899, 12, 31):
        return value.time()
    return value

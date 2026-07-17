"""Server-side record reads for Workbook Editor V2.

The reader exposes semantic workbook rows without leaking filesystem details or
loading workbook contents into a client. Physical Excel row numbers are the
stable record identity across filtering, sorting, and pagination.
"""

from __future__ import annotations

import math
import unicodedata
import xml.etree.ElementTree as ET
from zipfile import BadZipFile, ZipFile
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Final, Sequence

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils.cell import range_boundaries

from services.workbook_formula import (
    WorkbookFormulaError,
    evaluate_normalized_formula_columns,
    normalize_column_formulas,
)


PRICE_FIELDS: Final = frozenset({"net_price", "selling_price"})
IDENTITY_FIELDS: Final = frozenset(
    {"passenger_name", "pnr", "ticket_number"}
)
SUPPORTED_FIELDS: Final = PRICE_FIELDS | IDENTITY_FIELDS
SORT_DIRECTIONS: Final = frozenset({"asc", "desc"})


class WorkbookReadError(ValueError):
    """Safe reader failure with a stable machine-readable error code."""

    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code


@dataclass(frozen=True)
class WorkbookColumn:
    """A mapped semantic column returned to the records workbench."""

    field: str
    header: str
    column_number: int
    editable: bool
    semantic_field: str | None = None
    id: str = ""
    origin: str = "source"
    data_type: str = "text"
    hidden: bool = False
    sticky: bool = False
    group_label: str | None = None
    header_row_span: int = 1
    formula: dict[str, Any] | None = None
    number_format: str | None = None


@dataclass(frozen=True)
class WorkbookRecord:
    """One non-empty physical Excel row."""

    row_number: int
    values: dict[str, Any]
    editable: dict[str, bool]


@dataclass(frozen=True, slots=True)
class WorkbookCellReference:
    row_number: int
    column_id: str


@dataclass(frozen=True, slots=True)
class WorkbookCellValueResult:
    row_number: int
    column_id: str
    value: Any


@dataclass(frozen=True)
class WorkbookPagination:
    """Pagination metadata calculated after filtering and sorting."""

    page: int
    page_size: int
    total: int
    total_pages: int


@dataclass(frozen=True)
class WorkbookRecordPage:
    """Mapped columns, records, and pagination for one request."""

    columns: tuple[WorkbookColumn, ...]
    records: tuple[WorkbookRecord, ...]
    pagination: WorkbookPagination
    header_row_count: int = 1


def _source_header_text(value: Any) -> str:
    return "" if value is None else str(value)


def read_header_structure(
    worksheet: Any,
    header_row_number: int,
    *,
    max_row: int | None = None,
    max_column: int | None = None,
) -> tuple[int, dict[int, tuple[str, str | None, int]]]:
    """Read a one- or two-level Excel header without translating its labels."""

    max_column = max_column if max_column is not None else worksheet.max_column
    max_row = max_row if max_row is not None else worksheet.max_row
    top_values = {
        column: worksheet.cell(header_row_number, column).value
        for column in range(1, max_column + 1)
    }
    header_end_row = header_row_number

    merged_cells = getattr(worksheet, "merged_cells", None)
    for merged_range in getattr(merged_cells, "ranges", ()):
        if merged_range.min_row <= header_row_number <= merged_range.max_row:
            header_end_row = max(header_end_row, merged_range.max_row)

    header_end_row = min(header_end_row, max_row)
    if header_end_row == header_row_number and header_row_number < max_row:
        next_values = [
            worksheet.cell(header_row_number + 1, column).value
            for column in range(1, max_column + 1)
        ]
        populated = [value for value in next_values if value is not None and value != ""]
        # Legacy XLS conversions created before merge preservation lost their
        # ranges. A sparse, text-only row directly below a header is the
        # surviving child-header band, not a business record.
        if (
            populated
            and len(populated) <= max(4, max_column // 2)
            and all(isinstance(value, str) for value in populated)
            and any(value is None or value == "" for value in top_values.values())
        ):
            header_end_row += 1

    structure = {
        column: (_source_header_text(top_values[column]), None, 1)
        for column in range(1, max_column + 1)
    }
    if header_end_row == header_row_number:
        return header_end_row, structure

    child_values = {
        column: worksheet.cell(header_row_number + 1, column).value
        for column in range(1, max_column + 1)
    }
    top_starts = [
        column
        for column in range(1, max_column + 1)
        if top_values[column] is not None and top_values[column] != ""
    ]
    for index, start_column in enumerate(top_starts):
        end_column = (
            top_starts[index + 1] - 1
            if index + 1 < len(top_starts)
            else max_column
        )
        has_children = any(
            child_values[column] is not None and child_values[column] != ""
            for column in range(start_column, end_column + 1)
        )
        if has_children:
            group_label = _source_header_text(top_values[start_column])
            for column in range(start_column, end_column + 1):
                structure[column] = (
                    _source_header_text(child_values[column]),
                    group_label,
                    1,
                )
        else:
            structure[start_column] = (
                _source_header_text(top_values[start_column]),
                None,
                2,
            )
    return header_end_row, structure


def normalize_workbook_cell_value(value: Any) -> Any:
    """Convert Excel-only temporal values to stable JSON-safe values."""

    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, timedelta):
        total_microseconds = (
            value.days * 86_400_000_000
            + value.seconds * 1_000_000
            + value.microseconds
        )
        sign = "-" if total_microseconds < 0 else ""
        total_microseconds = abs(total_microseconds)
        total_seconds, microseconds = divmod(total_microseconds, 1_000_000)
        days, remainder = divmod(total_seconds, 86_400)
        hours, remainder = divmod(remainder, 3_600)
        minutes, seconds = divmod(remainder, 60)
        seconds_text = str(seconds)
        if microseconds:
            seconds_text += f".{microseconds:06d}".rstrip("0")
        date_part = f"{days}D" if days else ""
        time_parts = ""
        if hours:
            time_parts += f"{hours}H"
        if minutes:
            time_parts += f"{minutes}M"
        if seconds or microseconds or not (date_part or time_parts):
            time_parts += f"{seconds_text}S"
        return f"{sign}P{date_part}T{time_parts}"
    return value


def _normalize_search_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        text = value.isoformat()
    else:
        text = str(value)
    text = text.translate(str.maketrans({"đ": "d", "Đ": "D"}))
    decomposed = unicodedata.normalize("NFKD", text)
    without_accents = "".join(
        character
        for character in decomposed
        if not unicodedata.combining(character)
    )
    return " ".join(without_accents.casefold().split())


def _validate_request(
    *,
    header_row_number: int,
    column_mapping: dict[str, int],
    page: int,
    page_size: int,
    max_page_size: int,
    sort_by: str | None,
    sort_direction: str,
) -> None:
    if page < 1 or page_size < 1 or max_page_size < 1 or page_size > max_page_size:
        raise WorkbookReadError(
            "INVALID_PAGINATION",
            "Page and page size must be positive and within the configured limit.",
        )
    if header_row_number < 1:
        raise WorkbookReadError(
            "INVALID_ROW",
            "Header row number must be a positive Excel row number.",
        )
    if not isinstance(column_mapping, dict):
        raise WorkbookReadError("INVALID_MAPPING", "Column mapping must be an object.")

    fields = set(column_mapping)
    if fields - SUPPORTED_FIELDS:
        raise WorkbookReadError(
            "INVALID_MAPPING",
            "Column mapping contains an unsupported semantic field.",
        )
    column_numbers = list(column_mapping.values())
    if any(
        isinstance(column_number, bool)
        or not isinstance(column_number, int)
        or column_number < 1
        for column_number in column_numbers
    ):
        raise WorkbookReadError(
            "INVALID_MAPPING",
            "Mapped columns must use positive Excel column numbers.",
        )
    if len(set(column_numbers)) != len(column_numbers):
        raise WorkbookReadError(
            "AMBIGUOUS_MAPPING",
            "Each semantic field must map to a different Excel column.",
        )
    if sort_by is not None and (
        not isinstance(sort_by, str) or not sort_by or len(sort_by) > 64
    ):
        raise WorkbookReadError(
            "INVALID_SORT",
            "Sort column ID must be a non-empty string of at most 64 characters.",
        )
    if sort_direction not in SORT_DIRECTIONS:
        raise WorkbookReadError(
            "INVALID_SORT",
            "Sort direction must be either 'asc' or 'desc'.",
        )


def _is_editable_cell(
    cell: Cell,
    *,
    row_number: int,
    column_number: int,
    sheet_protected: bool,
    merged_cells: set[tuple[int, int]],
) -> bool:
    if cell.data_type == "f":
        return False
    if (row_number, column_number) in merged_cells:
        return False
    return not (sheet_protected and cell.protection.locked)


def _read_merged_cells(
    path: Path,
    worksheet: Any,
    columns: set[int],
) -> set[tuple[int, int]]:
    worksheet_path = getattr(worksheet, "_worksheet_path", None)
    if not worksheet_path:
        return set()

    merged_cells: set[tuple[int, int]] = set()
    with ZipFile(path) as archive, archive.open(worksheet_path) as worksheet_xml:
        for _event, element in ET.iterparse(worksheet_xml, events=("end",)):
            if not element.tag.endswith("mergeCell"):
                element.clear()
                continue
            reference = element.attrib.get("ref")
            if reference:
                min_column, min_row, max_column, max_row = range_boundaries(reference)
                for column_number in columns:
                    if min_column <= column_number <= max_column:
                        merged_cells.update(
                            (row_number, column_number)
                            for row_number in range(min_row, max_row + 1)
                        )
            element.clear()
    return merged_cells


def read_hidden_columns(path: str | Path, worksheet: Any) -> set[int]:
    """Read hidden source columns without loading the worksheet into memory."""

    worksheet_path = getattr(worksheet, "_worksheet_path", None)
    if not worksheet_path:
        return set()

    hidden_columns: set[int] = set()
    with ZipFile(Path(path)) as archive, archive.open(worksheet_path) as worksheet_xml:
        for _event, element in ET.iterparse(worksheet_xml, events=("end",)):
            if not element.tag.endswith("}col") and element.tag != "col":
                element.clear()
                continue
            if element.attrib.get("hidden", "").casefold() not in {"1", "true"}:
                element.clear()
                continue
            min_column = int(element.attrib.get("min", "0"))
            max_column = int(element.attrib.get("max", "0"))
            if min_column > 0 and max_column >= min_column:
                hidden_columns.update(range(min_column, max_column + 1))
            element.clear()
    return hidden_columns


def _sort_value(value: Any) -> tuple[int, int, Any]:
    """Return a comparable key with blank values consistently placed last."""

    if value is None or value == "":
        return (1, 0, "")
    if isinstance(value, bool):
        return (0, 0, int(value))
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return (1, 0, "")
        return (0, 0, value)
    if isinstance(value, (datetime, date, time)):
        return (0, 1, value.isoformat())
    return (0, 1, _normalize_search_value(value))


def read_workbook_cell_values(
    path: str | Path,
    *,
    sheet_name: str,
    header_row_number: int,
    column_config: list[dict[str, Any]],
    cells: Sequence[WorkbookCellReference],
    max_cells: int = 500,
    meaningful_max_row: int | None = None,
    meaningful_max_column: int | None = None,
) -> tuple[WorkbookCellValueResult, ...]:
    """Read bounded sparse cell values using stable configured column IDs."""

    if max_cells < 1 or not 1 <= len(cells) <= max_cells:
        raise WorkbookReadError(
            "INVALID_CELL_LOOKUP",
            f"Cell lookup must contain between 1 and {max_cells} coordinates.",
        )
    references = [(cell.row_number, cell.column_id) for cell in cells]
    if len(references) != len(set(references)):
        raise WorkbookReadError(
            "INVALID_CELL_LOOKUP",
            "Workbook cell references must be unique.",
        )
    if header_row_number < 1 or any(
        isinstance(row_number, bool) or row_number < 1
        for row_number, _column_id in references
    ):
        raise WorkbookReadError("INVALID_ROW", "Workbook rows must be positive.")
    if not isinstance(sheet_name, str) or not sheet_name.strip():
        raise WorkbookReadError("SHEET_NOT_FOUND", "Worksheet was not found.")
    if not isinstance(column_config, list) or not column_config:
        raise WorkbookReadError("INVALID_MAPPING", "Column configuration is unavailable.")

    config_by_id: dict[str, dict[str, Any]] = {}
    column_numbers: set[int] = set()
    for item in column_config:
        if not isinstance(item, dict):
            raise WorkbookReadError("INVALID_MAPPING", "Column configuration is invalid.")
        column_id = item.get("id")
        column_number = item.get("column_number")
        if (
            not isinstance(column_id, str)
            or not column_id
            or len(column_id) > 64
            or column_id in config_by_id
            or isinstance(column_number, bool)
            or not isinstance(column_number, int)
            or column_number < 1
            or column_number in column_numbers
        ):
            raise WorkbookReadError("INVALID_MAPPING", "Column configuration is invalid.")
        config_by_id[column_id] = item
        column_numbers.add(column_number)

    try:
        normalized_config, formula_order = normalize_column_formulas(column_config)
    except WorkbookFormulaError as exc:
        raise WorkbookReadError(exc.code, str(exc)) from exc
    config_by_id = {str(item["id"]): item for item in normalized_config}

    unknown_ids = {column_id for _row_number, column_id in references} - set(config_by_id)
    if unknown_ids:
        raise WorkbookReadError(
            "COLUMN_NOT_FOUND",
            "A requested workbook column was not found.",
        )

    workbook = None
    try:
        workbook = load_workbook(filename=Path(path), read_only=True, data_only=False)
        if sheet_name not in workbook.sheetnames:
            raise WorkbookReadError("SHEET_NOT_FOUND", "Worksheet was not found.")
        worksheet = workbook[sheet_name]
        max_row = meaningful_max_row or int(worksheet.max_row or 0)
        max_column = meaningful_max_column or int(worksheet.max_column or 0)
        if header_row_number > max_row:
            raise WorkbookReadError(
                "INVALID_ROW",
                "Header row is outside the selected worksheet.",
            )
        if max(column_numbers) > max_column:
            raise WorkbookReadError(
                "INVALID_MAPPING",
                "A configured column is outside the selected worksheet.",
            )
        header_end_row, _header_structure = read_header_structure(
            worksheet,
            header_row_number,
            max_row=max_row,
            max_column=max_column,
        )
        requested_rows = {row_number for row_number, _column_id in references}
        if any(
            row_number <= header_end_row or row_number > max_row
            for row_number in requested_rows
        ):
            raise WorkbookReadError(
                "INVALID_ROW",
                "A requested workbook row is outside the data region.",
            )

        values_by_reference: dict[tuple[int, str], Any] = {}
        found_rows: set[int] = set()
        ordered_config = sorted(
            config_by_id.items(), key=lambda item: int(item[1]["column_number"])
        )
        max_requested_row = max(requested_rows)
        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=header_end_row + 1,
                max_row=max_requested_row,
            ),
            start=header_end_row + 1,
        ):
            if row_number not in requested_rows:
                continue
            if all(cell.value is None for cell in row):
                continue
            found_rows.add(row_number)
            values_by_id = {
                column_id: row[int(config["column_number"]) - 1].value
                for column_id, config in ordered_config
            }
            values_by_id, _evaluations = evaluate_normalized_formula_columns(
                normalized_config, formula_order, values_by_id
            )
            for requested_row, column_id in references:
                if requested_row == row_number:
                    values_by_reference[(requested_row, column_id)] = values_by_id[column_id]

        if found_rows != requested_rows:
            raise WorkbookReadError(
                "INVALID_ROW",
                "A requested workbook row is blank or unavailable.",
            )
        return tuple(
            WorkbookCellValueResult(
                row_number=row_number,
                column_id=column_id,
                value=normalize_workbook_cell_value(
                    values_by_reference[(row_number, column_id)]
                ),
            )
            for row_number, column_id in references
        )
    except WorkbookReadError:
        raise
    except (BadZipFile, InvalidFileException, OSError, ValueError, KeyError) as exc:
        raise WorkbookReadError(
            "INVALID_XLSX",
            "Workbook could not be read as a valid .xlsx file.",
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()


def read_workbook_records(
    path: str | Path,
    *,
    sheet_name: str,
    header_row_number: int,
    column_mapping: dict[str, int],
    column_config: list[dict[str, Any]] | None = None,
    page: int = 1,
    page_size: int = 50,
    max_page_size: int = 200,
    search: str | None = None,
    sort_by: str | None = None,
    sort_direction: str = "asc",
    meaningful_max_row: int | None = None,
    meaningful_max_column: int | None = None,
) -> WorkbookRecordPage:
    """Read a filtered, sorted page of semantic records from one worksheet."""

    _validate_request(
        header_row_number=header_row_number,
        column_mapping=column_mapping,
        page=page,
        page_size=page_size,
        max_page_size=max_page_size,
        sort_by=sort_by,
        sort_direction=sort_direction,
    )
    if not isinstance(sheet_name, str) or not sheet_name.strip():
        raise WorkbookReadError("SHEET_NOT_FOUND", "Worksheet was not found.")

    workbook = None
    try:
        workbook = load_workbook(
            filename=Path(path),
            read_only=True,
            data_only=False,
        )
        if sheet_name not in workbook.sheetnames:
            raise WorkbookReadError("SHEET_NOT_FOUND", "Worksheet was not found.")
        worksheet = workbook[sheet_name]
        max_row = meaningful_max_row or int(worksheet.max_row or 0)
        max_column = meaningful_max_column or int(worksheet.max_column or 0)
        if header_row_number > max_row:
            raise WorkbookReadError(
                "INVALID_ROW",
                "Header row is outside the selected worksheet.",
            )
        if column_mapping and max(column_mapping.values()) > max_column:
            raise WorkbookReadError(
                "INVALID_MAPPING",
                "A mapped column is outside the selected worksheet.",
            )

        semantic_by_column = {
            column_number: field for field, column_number in column_mapping.items()
        }
        try:
            normalized_config, formula_order = normalize_column_formulas(
                column_config or []
            ) if column_config else ([], ())
        except WorkbookFormulaError as exc:
            raise WorkbookReadError(exc.code, str(exc)) from exc
        config_by_number = {
            int(item["column_number"]): item
            for item in normalized_config
            if isinstance(item, dict) and isinstance(item.get("column_number"), int)
        }
        header_end_row, header_structure = read_header_structure(
            worksheet,
            header_row_number,
            max_row=max_row,
            max_column=max_column,
        )
        ordered_column_numbers = tuple(
            int(item["column_number"])
            for item in normalized_config
            if isinstance(item, dict)
            and isinstance(item.get("column_number"), int)
            and 1 <= int(item["column_number"]) <= max_column
        ) or tuple(range(1, max_column + 1))
        columns = tuple(
            WorkbookColumn(
                field=str(
                    config_by_number.get(column_number, {}).get("id")
                    or semantic_by_column.get(column_number)
                    or f"column_{column_number}"
                ),
                header=header_structure[column_number][0],
                column_number=column_number,
                editable=(
                    not bool(config_by_number.get(column_number, {}).get("formula"))
                    if column_config
                    else semantic_by_column.get(column_number) in PRICE_FIELDS
                ),
                semantic_field=semantic_by_column.get(column_number),
                id=str(config_by_number.get(column_number, {}).get("id") or f"source-{column_number}"),
                origin=str(config_by_number.get(column_number, {}).get("origin", "source")),
                data_type=str(config_by_number.get(column_number, {}).get("data_type", "text")),
                hidden=bool(config_by_number.get(column_number, {}).get("hidden", False)),
                sticky=bool(config_by_number.get(column_number, {}).get("sticky", False)),
                group_label=header_structure[column_number][1],
                header_row_span=header_structure[column_number][2],
                formula=config_by_number.get(column_number, {}).get("formula"),
                number_format=str(
                    worksheet.cell(
                        row=min(header_end_row + 1, max_row),
                        column=column_number,
                    ).number_format
                    or "General"
                ),
            )
            for column_number in ordered_column_numbers
        )
        records: list[WorkbookRecord] = []
        matched_total = 0
        page_start = (page - 1) * page_size
        page_end = page_start + page_size
        visible_columns = ordered_column_numbers
        sortable_fields = {
            key: column.field
            for column in columns
            for key in (column.id, column.field, column.semantic_field)
            if key is not None
        }
        if sort_by is not None and sort_by not in sortable_fields:
            raise WorkbookReadError(
                "INVALID_SORT",
                "Sort column ID was not found in the worksheet configuration.",
            )
        search_fields = tuple(column.field for column in columns)
        normalized_search = _normalize_search_value(search)
        protection = getattr(worksheet, "protection", None)
        sheet_protected = bool(protection and protection.sheet)
        merged_cells = _read_merged_cells(
            Path(path),
            worksheet,
            set(visible_columns),
        )

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=header_end_row + 1,
                max_row=max_row,
                max_col=max_column,
            ),
            start=header_end_row + 1,
        ):
            if all(cell.value is None for cell in row):
                continue
            cell_by_column = {
                column_number: row[column_number - 1]
                for column_number in visible_columns
            }
            values = {
                column.field: cell_by_column[column.column_number].value
                for column in columns
            }
            values_by_id = {column.id: values[column.field] for column in columns}
            values_by_id, _evaluations = evaluate_normalized_formula_columns(
                normalized_config, formula_order, values_by_id
            )
            for column in columns:
                if column.formula:
                    values[column.field] = values_by_id[column.id]
            if normalized_search and not any(
                normalized_search in _normalize_search_value(values[field])
                for field in search_fields
            ):
                continue
            matched_total += 1
            if sort_by is None and not page_start < matched_total <= page_end:
                continue
            editable = {
                column.field: (
                    _is_editable_cell(
                        cell_by_column[column.column_number],
                        row_number=row_number,
                        column_number=column.column_number,
                        sheet_protected=sheet_protected,
                        merged_cells=merged_cells,
                    )
                    and not (
                        column.data_type == "date"
                        and isinstance(
                            cell_by_column[column.column_number].value,
                            (datetime, time, timedelta),
                        )
                        and not (
                            isinstance(
                                cell_by_column[column.column_number].value,
                                datetime,
                            )
                            and cell_by_column[column.column_number].value.time()
                            == time.min
                        )
                    )
                )
                for column in columns
                if (column_config and not column.formula)
                or (not column_config and column.semantic_field in PRICE_FIELDS)
            }
            values = {
                field: normalize_workbook_cell_value(value)
                for field, value in values.items()
            }
            records.append(
                WorkbookRecord(
                    row_number=row_number,
                    values=values,
                    editable=editable,
                )
            )

        if sort_by is not None:
            sort_field = sortable_fields[sort_by]
            records.sort(key=lambda record: record.row_number)
            populated_records = [
                record
                for record in records
                if _sort_value(record.values[sort_field])[0] == 0
            ]
            blank_records = [
                record
                for record in records
                if _sort_value(record.values[sort_field])[0] == 1
            ]
            populated_records.sort(
                key=lambda record: _sort_value(record.values[sort_field]),
                reverse=sort_direction == "desc",
            )
            records = populated_records + blank_records

        total = len(records) if sort_by is not None else matched_total
        total_pages = math.ceil(total / page_size) if total else 0
        page_records = (
            tuple(records[page_start:page_end])
            if sort_by is not None
            else tuple(records)
        )
        return WorkbookRecordPage(
            columns=columns,
            records=page_records,
            pagination=WorkbookPagination(
                page=page,
                page_size=page_size,
                total=total,
                total_pages=total_pages,
            ),
            header_row_count=header_end_row - header_row_number + 1,
        )
    except WorkbookReadError:
        raise
    except (BadZipFile, InvalidFileException, OSError, ValueError, KeyError) as exc:
        raise WorkbookReadError(
            "INVALID_XLSX",
            "Workbook could not be read as a valid .xlsx file.",
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()
